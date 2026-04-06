import os
import re
import json
import time
import uuid
import random
import logging
import html
from io import BytesIO
from datetime import datetime, timezone, timedelta
from typing import Optional, Tuple, List, Dict

import requests
from flask import Flask, request, jsonify

import sqlalchemy as sa
from sqlalchemy import text
from sqlalchemy.exc import IntegrityError

from openpyxl import Workbook

# Timezone (Python 3.9+)
try:
    from zoneinfo import ZoneInfo
except Exception:
    ZoneInfo = None

# Cloud Tasks (optional)
try:
    from google.cloud import tasks_v2
    _HAS_CLOUD_TASKS = True
except Exception:
    tasks_v2 = None
    _HAS_CLOUD_TASKS = False


# ---------------------------
# CONFIG
# ---------------------------
LOG_LEVEL = os.getenv("LOG_LEVEL", "INFO").upper()
logging.basicConfig(level=LOG_LEVEL, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger("boda8")

DATABASE_URL = os.getenv("DATABASE_URL", "").strip()
if not DATABASE_URL:
    raise RuntimeError("❌ DATABASE_URL env var required")

APP_TZ_NAME = os.getenv("TZ", "Asia/Kuala_Lumpur")
LOCAL_TZ = ZoneInfo(APP_TZ_NAME) if ZoneInfo else None

DB_POOL_SIZE = int(os.getenv("DB_POOL_SIZE", "3"))
DB_MAX_OVERFLOW = int(os.getenv("DB_MAX_OVERFLOW", "7"))

AFFILIATE_AMOUNT = float(os.getenv("AFFILIATE_AMOUNT", "1.00"))

MIN_WITHDRAW_DEFAULT = float(os.getenv("MIN_WITHDRAW_DEFAULT", "30.00"))

# Broadcast tuning
BROADCAST_BATCH = int(os.getenv("BROADCAST_BATCH", "200"))
BROADCAST_SLEEP = float(os.getenv("BROADCAST_SLEEP", "0.04"))

# Cloud Tasks env
GCP_PROJECT = os.getenv("GCP_PROJECT", "").strip()
TASKS_LOCATION = os.getenv("TASKS_LOCATION", "").strip()
TASKS_QUEUE = os.getenv("TASKS_QUEUE", "").strip()
TASKS_HANDLER_URL = os.getenv("TASKS_HANDLER_URL", "").strip()
TASKS_SECRET = os.getenv("TASKS_SECRET", "").strip()

SERVICE_NAME = os.getenv("SERVICE_NAME", "boda8-bot")

# untuk /addbot auto setWebhook
PUBLIC_BASE_URL = os.getenv("PUBLIC_BASE_URL", "").strip().rstrip("/")

# UI /settings
SETTINGS_CB_PAGE_SIZE = int(os.getenv("SETTINGS_CB_PAGE_SIZE", "12"))

# Admin management defaults
ADMIN_DEFAULT_DAYS = int(os.getenv("ADMIN_DEFAULT_DAYS", "30"))

# Telegram API limits (visible text after HTML parsing)
TG_MAX_TEXT = int(os.getenv("TG_MAX_TEXT", "4096"))
TG_MAX_CAPTION = int(os.getenv("TG_MAX_CAPTION", "1024"))


# ---------------------------
# FLASK & DB
# ---------------------------
app = Flask(__name__)
app.url_map.strict_slashes = False  # /healthz dan /healthz/ sama-sama ok

engine = sa.create_engine(
    DATABASE_URL,
    pool_pre_ping=True,
    pool_size=DB_POOL_SIZE,
    max_overflow=DB_MAX_OVERFLOW,
    pool_recycle=1800,
    future=True,
)

TG_API = "https://api.telegram.org/bot{token}/{method}"
SESSION = requests.Session()

# Admin interactive flow state management
# Format: {(bot_id, user_id): (action, timestamp)}
pending_inputs = {}


# ---------------------------
# UTILS
# ---------------------------
def utcnow() -> datetime:
    return datetime.now(timezone.utc)


def now_local_str(fmt: str) -> str:
    if LOCAL_TZ:
        return datetime.now(LOCAL_TZ).strftime(fmt)
    return datetime.now().strftime(fmt)


def _exec_ddl_multi(conn, ddl: str):
    stmts = [s.strip() for s in ddl.split(";") if s.strip()]
    for s in stmts:
        conn.execute(text(s))


def init_db():
    ddl = """
    CREATE EXTENSION IF NOT EXISTS pgcrypto;

    CREATE TABLE IF NOT EXISTS bots (
      id UUID PRIMARY KEY,
      token TEXT NOT NULL,
      bot_username TEXT,
      secret_token TEXT UNIQUE NOT NULL,
      owner_id BIGINT NOT NULL,
      admin_group_id BIGINT,
      lock_bot BOOLEAN NOT NULL DEFAULT FALSE,
      start_text TEXT,
      start_media_type TEXT,
      start_media_file_id TEXT,
      loading_text TEXT,
      loading_media_type TEXT,
      loading_media_file_id TEXT,
      created_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
    );

    CREATE TABLE IF NOT EXISTS admins (
      bot_id UUID NOT NULL REFERENCES bots(id) ON DELETE CASCADE,
      admin_user_id BIGINT NOT NULL,
      expiry_at TIMESTAMPTZ,
      added_by BIGINT NOT NULL,
      PRIMARY KEY (bot_id, admin_user_id)
    );

    CREATE TABLE IF NOT EXISTS users (
      bot_id UUID NOT NULL REFERENCES bots(id) ON DELETE CASCADE,
      user_id BIGINT NOT NULL,
      username TEXT,
      first_name TEXT,
      phone TEXT,
      member_id TEXT,
      is_verified BOOLEAN NOT NULL DEFAULT FALSE,
      balance NUMERIC NOT NULL DEFAULT 0,
      shared_count BIGINT NOT NULL DEFAULT 0,
      joined_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
      upline_user_id BIGINT,
      credited_upline BOOLEAN NOT NULL DEFAULT FALSE,
      PRIMARY KEY (bot_id, user_id)
    );

    CREATE UNIQUE INDEX IF NOT EXISTS users_phone_unique
      ON users(bot_id, phone) WHERE phone IS NOT NULL;

    CREATE TABLE IF NOT EXISTS actions (
      bot_id UUID NOT NULL REFERENCES bots(id) ON DELETE CASCADE,
      key TEXT NOT NULL,
      type TEXT NOT NULL,
      text TEXT,
      media_file_id TEXT,
      delay_seconds INT NOT NULL DEFAULT 0,
      created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
      PRIMARY KEY (bot_id, key)
    );

    CREATE TABLE IF NOT EXISTS user_states (
      bot_id UUID NOT NULL REFERENCES bots(id) ON DELETE CASCADE,
      user_id BIGINT NOT NULL,
      state TEXT NOT NULL,
      payload JSONB,
      updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
      PRIMARY KEY (bot_id, user_id)
    );

    CREATE TABLE IF NOT EXISTS withdrawals (
      id UUID PRIMARY KEY,
      bot_id UUID NOT NULL REFERENCES bots(id) ON DELETE CASCADE,
      user_id BIGINT NOT NULL,
      request_text TEXT NOT NULL,
      status TEXT NOT NULL DEFAULT 'PENDING',
      created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
      processed_at TIMESTAMPTZ,
      processed_by BIGINT,
      approved_amount NUMERIC,
      request_amount NUMERIC
    );
    ALTER TABLE withdrawals ADD COLUMN IF NOT EXISTS request_amount NUMERIC;

    -- ----------------------------
    -- SAFE MIGRATIONS (NEW FIELDS)
    -- ----------------------------
    ALTER TABLE users ADD COLUMN IF NOT EXISTS phone_updated_at TIMESTAMPTZ;
    ALTER TABLE users ADD COLUMN IF NOT EXISTS is_premium BOOLEAN NOT NULL DEFAULT FALSE;
    ALTER TABLE users ADD COLUMN IF NOT EXISTS premium_until TIMESTAMPTZ;

    ALTER TABLE bots ADD COLUMN IF NOT EXISTS join_lock BOOLEAN NOT NULL DEFAULT FALSE;
    ALTER TABLE bots ADD COLUMN IF NOT EXISTS join_targets TEXT;         -- lines: @channel or -100123...
    ALTER TABLE bots ADD COLUMN IF NOT EXISTS join_message TEXT;         -- custom prompt join
    ALTER TABLE bots ADD COLUMN IF NOT EXISTS join_message_media_type TEXT;
    ALTER TABLE bots ADD COLUMN IF NOT EXISTS join_message_media_file_id TEXT;
    ALTER TABLE bots ADD COLUMN IF NOT EXISTS joined_message TEXT;       -- message after user passes join gate
    ALTER TABLE bots ADD COLUMN IF NOT EXISTS joined_message_media_type TEXT;
    ALTER TABLE bots ADD COLUMN IF NOT EXISTS joined_message_media_file_id TEXT;
    ALTER TABLE bots ADD COLUMN IF NOT EXISTS contact_message TEXT;      -- custom prompt share contact
    ALTER TABLE bots ADD COLUMN IF NOT EXISTS contact_message_media_type TEXT;
    ALTER TABLE bots ADD COLUMN IF NOT EXISTS contact_message_media_file_id TEXT;
    ALTER TABLE bots ADD COLUMN IF NOT EXISTS pending_message TEXT;      -- after contact, pending approval
    ALTER TABLE bots ADD COLUMN IF NOT EXISTS pending_message_media_type TEXT;
    ALTER TABLE bots ADD COLUMN IF NOT EXISTS pending_message_media_file_id TEXT;
    ALTER TABLE bots ADD COLUMN IF NOT EXISTS verified_message TEXT;     -- premium approved message
    ALTER TABLE bots ADD COLUMN IF NOT EXISTS verified_message_media_type TEXT;
    ALTER TABLE bots ADD COLUMN IF NOT EXISTS verified_message_media_file_id TEXT;
    ALTER TABLE bots ADD COLUMN IF NOT EXISTS rejected_message TEXT;     -- premium rejected message
    ALTER TABLE bots ADD COLUMN IF NOT EXISTS rejected_message_media_type TEXT;
    ALTER TABLE bots ADD COLUMN IF NOT EXISTS rejected_message_media_file_id TEXT;
    ALTER TABLE bots ADD COLUMN IF NOT EXISTS group_contact_message TEXT;-- message to admin group about new contact
    ALTER TABLE bots ADD COLUMN IF NOT EXISTS group_contact_message_media_type TEXT;
    ALTER TABLE bots ADD COLUMN IF NOT EXISTS group_contact_message_media_file_id TEXT;
    ALTER TABLE bots ADD COLUMN IF NOT EXISTS withdrawal_prompt TEXT;    -- user prompt for withdraw
    ALTER TABLE bots ADD COLUMN IF NOT EXISTS withdrawal_request_message TEXT;
    ALTER TABLE bots ADD COLUMN IF NOT EXISTS withdrawal_request_media_type TEXT;
    ALTER TABLE bots ADD COLUMN IF NOT EXISTS withdrawal_request_media_file_id TEXT;
    ALTER TABLE bots ADD COLUMN IF NOT EXISTS manual_approval BOOLEAN NOT NULL DEFAULT FALSE;
    ALTER TABLE bots ADD COLUMN IF NOT EXISTS inplace_callbacks BOOLEAN NOT NULL DEFAULT FALSE;

    ALTER TABLE bots ADD COLUMN IF NOT EXISTS affiliate_amount NUMERIC;
    ALTER TABLE bots ADD COLUMN IF NOT EXISTS min_withdraw_amount NUMERIC;
    ALTER TABLE bots ADD COLUMN IF NOT EXISTS withdrawal_approve_message TEXT;
    ALTER TABLE bots ADD COLUMN IF NOT EXISTS withdrawal_approve_media_type TEXT;
    ALTER TABLE bots ADD COLUMN IF NOT EXISTS withdrawal_approve_media_file_id TEXT;
    ALTER TABLE bots ADD COLUMN IF NOT EXISTS withdrawal_reject_message TEXT;
    ALTER TABLE bots ADD COLUMN IF NOT EXISTS withdrawal_reject_media_type TEXT;
    ALTER TABLE bots ADD COLUMN IF NOT EXISTS withdrawal_reject_media_file_id TEXT;
    ALTER TABLE bots ADD COLUMN IF NOT EXISTS withdrawal_failed_message TEXT;
    ALTER TABLE bots ADD COLUMN IF NOT EXISTS withdrawal_failed_media_type TEXT;
    ALTER TABLE bots ADD COLUMN IF NOT EXISTS withdrawal_failed_media_file_id TEXT;
    ALTER TABLE bots ADD COLUMN IF NOT EXISTS withdrawal_submitted_message TEXT;
    ALTER TABLE bots ADD COLUMN IF NOT EXISTS withdrawal_submitted_media_type TEXT;
    ALTER TABLE bots ADD COLUMN IF NOT EXISTS withdrawal_submitted_media_file_id TEXT;
    
    -- Scanner media (per provider)
    CREATE TABLE IF NOT EXISTS scanner_media (
      bot_id UUID NOT NULL,
      provider TEXT NOT NULL,
      media_type TEXT NOT NULL,
      file_id TEXT NOT NULL,
      updated_at TIMESTAMPTZ DEFAULT NOW(),
      PRIMARY KEY (bot_id, provider)
    );

    -- Scanner games list (txt one-per-line)
    CREATE TABLE IF NOT EXISTS scanner_games (
      bot_id UUID NOT NULL,
      provider TEXT NOT NULL,
      game TEXT NOT NULL,
      PRIMARY KEY (bot_id, provider, game)
    );

    -- Scanner cooldown per user+provider
    CREATE TABLE IF NOT EXISTS scanner_cooldowns (
      bot_id UUID NOT NULL,
      user_id BIGINT NOT NULL,
      provider TEXT NOT NULL,
      last_at TIMESTAMPTZ NOT NULL,
      PRIMARY KEY (bot_id, user_id, provider)
    );

    -- ----------------------------
    -- Scanner daily limit (per bot + per user override)
    -- ----------------------------
    ALTER TABLE bots ADD COLUMN IF NOT EXISTS scan_limit_per_day INT;  -- null/<=0 = unlimited
    ALTER TABLE bots ADD COLUMN IF NOT EXISTS scan_limit_message TEXT;
    ALTER TABLE bots ADD COLUMN IF NOT EXISTS scan_limit_message_media_type TEXT;
    ALTER TABLE bots ADD COLUMN IF NOT EXISTS scan_limit_message_media_file_id TEXT;
    ALTER TABLE bots ADD COLUMN IF NOT EXISTS scanner_duration_seconds INT NOT NULL DEFAULT 25;

    CREATE TABLE IF NOT EXISTS scan_daily_usage (
      bot_id UUID NOT NULL,
      user_id BIGINT NOT NULL,
      day DATE NOT NULL,
      count INT NOT NULL DEFAULT 0,
      updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
      PRIMARY KEY (bot_id, user_id, day)
    );

    CREATE TABLE IF NOT EXISTS scan_limit_overrides (
      bot_id UUID NOT NULL,
      user_id BIGINT NOT NULL,
      limit_per_day INT NOT NULL,
      updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
      PRIMARY KEY (bot_id, user_id)
    );

    -- ----------------------------
    -- BIGINT MIGRATION for scan limits (fix integer overflow)
    -- Safe migration: converts existing INT values to BIGINT
    -- ----------------------------
    ALTER TABLE scan_limit_overrides 
      ALTER COLUMN limit_per_day TYPE BIGINT USING limit_per_day::BIGINT;

    ALTER TABLE scan_daily_usage 
      ALTER COLUMN count TYPE BIGINT USING count::BIGINT;

    ALTER TABLE bots 
      ALTER COLUMN scan_limit_per_day TYPE BIGINT USING scan_limit_per_day::BIGINT;

    -- ----------------------------
    -- LIVEGRAM (forward user msgs to admin group, admin replies go back)
    -- ----------------------------
    ALTER TABLE bots ADD COLUMN IF NOT EXISTS livegram BOOLEAN NOT NULL DEFAULT FALSE;
    ALTER TABLE bots ADD COLUMN IF NOT EXISTS livegram_scope TEXT DEFAULT 'private'; -- 'private' or 'all'

    CREATE TABLE IF NOT EXISTS livegram_messages (
      bot_id UUID NOT NULL,
      fwd_message_id BIGINT NOT NULL,
      source_chat_id BIGINT NOT NULL,
      source_user_id BIGINT NOT NULL,
      source_message_id BIGINT,
      created_at TIMESTAMPTZ DEFAULT NOW(),
      PRIMARY KEY (bot_id, fwd_message_id)
    );

"""
    with engine.begin() as conn:
        _exec_ddl_multi(conn, ddl)
    logger.info("✅ DB Init OK")


init_db()


# ---------------------------
# TELEGRAM SAFE HTML
# ---------------------------
_ALLOWED_TAGS = {
    "b", "strong", "i", "em", "u", "ins", "s", "strike", "del",
    "code", "pre", "a", "tg-spoiler", "tg-emoji"
}


def _trim(s: str, limit: int) -> str:
    if not s:
        return ""
    if len(s) <= limit:
        return s
    return s[:limit] + "…"


def sanitize_telegram_html(text_: str, max_len: int = None) -> str:
    """
    Robust Telegram HTML sanitizer using a proper parser.
    Handles: overlapping tags, orphaned closing tags, unclosed tags,
    unknown tags (silently stripped), and bare '<' characters.
    """
    if not text_:
        return ""

    from html.parser import HTMLParser

    # Trim raw text first so auto-close tags appended later aren't cut
    limit = max_len if max_len is not None else TG_MAX_TEXT
    text_ = _trim(text_, limit)
    # Remove any partial tag left by trim (e.g. "<a href=...")
    text_ = re.sub(r'<[^>]*$', '', text_)

    class _San(HTMLParser):
        def __init__(self):
            super().__init__(convert_charrefs=False)
            self.parts = []
            self.stack = []  # [(tag_lower, attrs)]

        def handle_starttag(self, tag, attrs):
            tl = tag.lower()
            if tl not in _ALLOWED_TAGS:
                return  # silently drop unknown tags
            self.stack.append((tl, attrs))
            a = "".join(f' {k}="{v}"' for k, v in attrs if v is not None)
            self.parts.append(f"<{tl}{a}>")

        def handle_endtag(self, tag):
            tl = tag.lower()
            if tl not in _ALLOWED_TAGS:
                return
            # Find matching opener in stack
            idx = None
            for i in range(len(self.stack) - 1, -1, -1):
                if self.stack[i][0] == tl:
                    idx = i
                    break
            if idx is None:
                return  # orphan closer, skip

            # Close intervening tags (fix overlapping)
            reopen = []
            while len(self.stack) > idx + 1:
                t, a = self.stack.pop()
                self.parts.append(f"</{t}>")
                reopen.append((t, a))

            # Close the matched tag
            self.stack.pop()
            self.parts.append(f"</{tl}>")

            # Reopen intervening tags
            for t, a in reversed(reopen):
                astr = "".join(f' {k}="{v}"' for k, v in a if v is not None)
                self.parts.append(f"<{t}{astr}>")
                self.stack.append((t, a))

        def handle_data(self, data):
            self.parts.append(data)

        def handle_entityref(self, name):
            self.parts.append(f"&{name};")

        def handle_charref(self, name):
            self.parts.append(f"&#{name};")

        def finish(self):
            # Auto-close remaining unclosed tags
            for t, _ in reversed(self.stack):
                self.parts.append(f"</{t}>")
            return "".join(self.parts)

    san = _San()
    try:
        san.feed(text_)
        return san.finish()
    except Exception:
        # If parser fails entirely, escape everything
        return html.escape(text_)



def _u16_index_map(s: str):
    """Build UTF-16 code-unit boundaries for Telegram entity offsets."""
    u16 = 0
    boundaries = [(0, 0)]
    for i, ch in enumerate(s):
        u16 += 1 if ord(ch) <= 0xFFFF else 2
        boundaries.append((u16, i + 1))
    return boundaries


def _u16_to_py(boundaries, u16_pos: int) -> int:
    for u16, py in boundaries:
        if u16 >= u16_pos:
            return py
    return boundaries[-1][1]


def entities_to_html(text_: str, entities: Optional[list]) -> str:
    """Convert Telegram entities (bold/italic/link/etc) into Telegram HTML markup."""
    if not text_ or not entities:
        return text_ or ""
    s = text_
    b = _u16_index_map(s)
    inserts = []
    for ent in entities:
        try:
            typ = ent.get("type")
            off_u16 = int(ent.get("offset", 0))
            ln_u16 = int(ent.get("length", 0))
            if ln_u16 <= 0:
                continue
            start = _u16_to_py(b, off_u16)
            end = _u16_to_py(b, off_u16 + ln_u16)
            if start >= end:
                continue

            open_tag = close_tag = None
            if typ == "bold":
                open_tag, close_tag = "<b>", "</b>"
            elif typ == "italic":
                open_tag, close_tag = "<i>", "</i>"
            elif typ == "underline":
                open_tag, close_tag = "<u>", "</u>"
            elif typ == "strikethrough":
                open_tag, close_tag = "<s>", "</s>"
            elif typ == "code":
                open_tag, close_tag = "<code>", "</code>"
            elif typ == "pre":
                open_tag, close_tag = "<pre>", "</pre>"
            elif typ == "spoiler":
                open_tag, close_tag = "<tg-spoiler>", "</tg-spoiler>"
            elif typ == "text_link":
                url = _normalize_url(ent.get("url") or "")
                if url:
                    open_tag, close_tag = f'<a href="{_h(url)}">', "</a>"
            else:
                continue

            if open_tag and close_tag:
                inserts.append((start, open_tag))
                inserts.append((end, close_tag))
        except Exception:
            continue

    if not inserts:
        return s

    inserts.sort(key=lambda x: x[0], reverse=True)
    for idx, tag in inserts:
        s = s[:idx] + tag + s[idx:]
    return s
# ---------------------------
# TELEGRAM API
# ---------------------------
def tg_call(token: str, method: str, params=None, data=None, files=None):
    try:
        r = SESSION.post(
            TG_API.format(token=token, method=method),
            params=params,
            data=data,
            files=files,
            timeout=25,
        )
        try:
            js = r.json()
        except Exception:
            logger.error(f"TG non-JSON {method}: status={r.status_code} body={r.text[:250]}")
            return None

        if not js.get("ok"):
            desc = (js.get("description") or "").lower()
            code = js.get("error_code")
            if method in ("editMessageText", "editMessageCaption", "editMessageMedia") and "message is not modified" in desc:
                return None
            # Suppress noisy but harmless errors to DEBUG
            if code == 403 and any(x in desc for x in [
                "bot was blocked", "user is deactivated",
                "bot can't initiate", "bots can't send",
            ]):
                logger.debug(f"TG Skip {method}: {desc}")
                return None
            if code == 400 and any(x in desc for x in [
                "chat not found",
                "there is no text in the message",
                "query is too old",
            ]):
                logger.debug(f"TG Skip {method}: {desc}")
                return None
            logger.error(f"TG Error {method}: {js}")
            return None
        return js.get("result")
    except Exception as e:
        logger.error(f"TG Exception {method}: {e}")
        return None


def send_message(token, chat_id, text_, reply_markup=None, parse_mode="HTML", reply_to_message_id=None):
    if not text_:
        return None
    text_ = sanitize_telegram_html(text_) if parse_mode == "HTML" else text_
    data = {
        "chat_id": chat_id,
        "text": text_,
        "parse_mode": parse_mode,
        "disable_web_page_preview": True,
    }
    if reply_to_message_id:
        data["reply_to_message_id"] = reply_to_message_id
    if reply_markup:
        data["reply_markup"] = json.dumps(reply_markup)
    return tg_call(token, "sendMessage", data=data)


# -------------------------------------------------------------------
# Backward-compatible Telegram helper aliases
# Some older parts of the code call tg_send_message / tg_send_photo etc.
# Keep these wrappers so the webhook won't crash with NameError.
# -------------------------------------------------------------------
def tg_send_message(token, chat_id, text_, reply_markup=None, parse_mode="HTML", reply_to_message_id=None):
    return send_message(token, chat_id, text_, reply_markup=reply_markup, parse_mode=parse_mode, reply_to_message_id=reply_to_message_id)

def tg_send_photo(token, chat_id, file_id, caption=None, reply_markup=None, parse_mode="HTML"):
    return send_media(token, chat_id, "photo", file_id, caption=caption, reply_markup=reply_markup, parse_mode=parse_mode)

def tg_send_video(token, chat_id, file_id, caption=None, reply_markup=None, parse_mode="HTML"):
    return send_media(token, chat_id, "video", file_id, caption=caption, reply_markup=reply_markup, parse_mode=parse_mode)

def tg_send_animation(token, chat_id, file_id, caption=None, reply_markup=None, parse_mode="HTML"):
    return send_media(token, chat_id, "animation", file_id, caption=caption, reply_markup=reply_markup, parse_mode=parse_mode)

def tg_send_document(token, chat_id, file_id, caption=None, reply_markup=None, parse_mode="HTML"):
    return send_media(token, chat_id, "document", file_id, caption=caption, reply_markup=reply_markup, parse_mode=parse_mode)



def delete_message(token, chat_id, message_id):
    data = {"chat_id": chat_id, "message_id": message_id}
    return tg_call(token, "deleteMessage", data=data)


def edit_message(token, chat_id, message_id, text_, reply_markup=None, parse_mode="HTML"):
    text_ = sanitize_telegram_html(text_) if parse_mode == "HTML" else text_
    data = {
        "chat_id": chat_id,
        "message_id": message_id,
        "text": text_,
        "parse_mode": parse_mode,
        "disable_web_page_preview": True,
    }
    if reply_markup:
        data["reply_markup"] = json.dumps(reply_markup)
    return tg_call(token, "editMessageText", data=data)


def edit_caption(token, chat_id, message_id, caption_, reply_markup=None, parse_mode="HTML"):
    caption_ = sanitize_telegram_html(caption_) if parse_mode == "HTML" else caption_
    caption_ = _trim(caption_, TG_MAX_CAPTION)
    data = {
        "chat_id": chat_id,
        "message_id": message_id,
        "caption": caption_,
        "parse_mode": parse_mode,
    }
    if reply_markup:
        data["reply_markup"] = json.dumps(reply_markup)
    return tg_call(token, "editMessageCaption", data=data)


def send_media(token, chat_id, media_type, file_id_or_url, caption=None, reply_markup=None, parse_mode="HTML"):
    method_map = {"photo": "sendPhoto", "video": "sendVideo", "animation": "sendAnimation", "document": "sendDocument"}
    field_map = {"photo": "photo", "video": "video", "animation": "animation", "document": "document"}

    if media_type not in method_map:
        return send_message(token, chat_id, caption or "", reply_markup=reply_markup, parse_mode=parse_mode)

    cap = sanitize_telegram_html(caption, max_len=TG_MAX_CAPTION) if (caption and parse_mode == "HTML") else caption
    if cap is not None and parse_mode != "HTML":
        cap = _trim(cap, TG_MAX_CAPTION)

    data = {"chat_id": chat_id, field_map[media_type]: file_id_or_url, "parse_mode": parse_mode}
    if cap:
        data["caption"] = cap
    if reply_markup:
        data["reply_markup"] = json.dumps(reply_markup)
    return tg_call(token, method_map[media_type], data=data)


def _input_media(media_type: str, file_id_or_url: str, caption: str, parse_mode: str) -> dict:
    # Telegram InputMedia types: photo, video, animation, document
    cap = sanitize_telegram_html(caption, max_len=TG_MAX_CAPTION) if (caption and parse_mode == "HTML") else caption
    cap = _trim(cap or "", TG_MAX_CAPTION) if parse_mode != "HTML" else (cap or "")

    m = {"type": media_type, "media": file_id_or_url}
    if cap:
        m["caption"] = cap
        m["parse_mode"] = parse_mode
    return m


def edit_media(token, chat_id, message_id, media_type, file_id_or_url, caption=None, reply_markup=None, parse_mode="HTML"):
    """
    Edit media in-place (only works if current message is a media message sent by bot).
    """
    if media_type not in ("photo", "video", "animation", "document"):
        return None

    media_obj = _input_media(media_type, file_id_or_url, caption or "", parse_mode)
    data = {
        "chat_id": chat_id,
        "message_id": message_id,
        "media": json.dumps(media_obj),
    }
    if reply_markup:
        data["reply_markup"] = json.dumps(reply_markup)
    return tg_call(token, "editMessageMedia", data=data)


def answer_callback(token, callback_query_id, text_=None, show_alert=False):
    data = {"callback_query_id": callback_query_id, "show_alert": show_alert}
    if text_:
        data["text"] = text_
    return tg_call(token, "answerCallbackQuery", data=data)


# ---------------------------
# LIVEGRAM HELPERS
# ---------------------------
def livegram_forward_to_admin(bot_row, msg):
    """Forward a user message to the admin group and save the mapping."""
    token = bot_row["token"]
    bot_id = str(bot_row["id"])
    admin_chat = bot_row.get("admin_group_id")
    if not admin_chat:
        admin_chat = bot_row.get("owner_id")
    if not admin_chat:
        return
    admin_chat = int(admin_chat)

    from_user = msg.get("from") or {}
    uid = from_user.get("id")
    fname = from_user.get("first_name") or "?"
    uname = from_user.get("username")
    source_chat_id = msg["chat"]["id"]
    source_msg_id = msg.get("message_id")
    chat_type = msg.get("chat", {}).get("type", "private")

    # Forward the original message first
    fwd_result = tg_call(token, "forwardMessage", data={
        "chat_id": admin_chat,
        "from_chat_id": source_chat_id,
        "message_id": source_msg_id,
    })
    logger.info("LIVEGRAM-FWD: result=%s", fwd_result)

    fwd_msg_id = None
    if fwd_result and fwd_result.get("message_id"):
        fwd_msg_id = fwd_result["message_id"]

    # Save mapping
    logger.info("LIVEGRAM-FWD: fwd_msg_id=%s bot_id=%s source_chat=%s source_user=%s",
                fwd_msg_id, bot_id, source_chat_id, uid)
    if fwd_msg_id:
        try:
            with engine.begin() as conn:
                conn.execute(text(
                    "INSERT INTO livegram_messages (bot_id, fwd_message_id, source_chat_id, source_user_id, source_message_id) "
                    "VALUES (:b, :fwd, :sc, :su, :sm) ON CONFLICT DO NOTHING"
                ), {"b": bot_id, "fwd": fwd_msg_id, "sc": source_chat_id, "su": uid, "sm": source_msg_id})
            logger.info("LIVEGRAM-FWD: saved mapping fwd_msg_id=%s -> source=%s", fwd_msg_id, source_chat_id)
        except Exception as e:
            logger.error("Livegram save mapping error: %s", e)


def livegram_handle_admin_reply(bot_row, msg):
    """If admin replies to a forwarded livegram message in admin group, send reply back to source."""
    token = bot_row["token"]
    bot_id = str(bot_row["id"])
    reply = msg.get("reply_to_message")
    if not reply:
        return False

    replied_msg_id = reply.get("message_id")
    if not replied_msg_id:
        return False


    # Also check if the reply is to a header message (which replies to the forwarded msg)
    # Try the replied message ID first, then the message it was replying to
    ids_to_check = [replied_msg_id]
    if reply.get("reply_to_message"):
        ids_to_check.append(reply["reply_to_message"].get("message_id"))

    row = None
    try:
        with engine.connect() as conn:
            for check_id in ids_to_check:
                if check_id:
                    row = conn.execute(text(
                        "SELECT source_chat_id, source_user_id, source_message_id "
                        "FROM livegram_messages WHERE bot_id=:b AND fwd_message_id=:fwd"
                    ), {"b": bot_id, "fwd": check_id}).mappings().first()
                    if row:
                        break
    except Exception as e:
        logger.error("Livegram lookup error: %s", e)
        return False

    if not row:
        return False

    dest_chat = int(row["source_chat_id"])
    dest_msg = row.get("source_message_id")


    # Copy admin reply to the source chat
    copy_data = {
        "chat_id": dest_chat,
        "from_chat_id": msg["chat"]["id"],
        "message_id": msg["message_id"],
    }
    if dest_msg:
        copy_data["reply_to_message_id"] = int(dest_msg)
    result = tg_call(token, "copyMessage", data=copy_data)

    # Confirm to admin
    if result:
        tg_call(token, "sendMessage", data={
            "chat_id": msg["chat"]["id"],
            "text": "✅ Reply terhantar!",
            "reply_to_message_id": msg["message_id"],
        })

    return True


# TEXT/BTN HELPERS
# ---------------------------
DATE_FMT = "%d/%m/%Y"

HELP_PLACEHOLDERS_FULL = (
    "📌 <b>PLACEHOLDER LIST (FULL)</b>\n"
    "━━━━━━━━━━━━━━━━━━\n"
    "• {firstname} : Nama user\n"
    "• {username}  : @username (kalau tiada, guna firstname)\n"
    "• {member_id} : Member ID (auto)\n"
    "• {date}      : Tarikh local\n"
    "• {rand:1-100}: Random number ikut range\n"
    "• [balance]   : Baki user (RM)\n"
    "• [share]     : Jumlah share\n"
    "• [link]      : Link affiliate (/start upline_id)\n"
    "• [web](https://example.com) : Auto jadi link clickable\n"
    "━━━━━━━━━━━━━━━━━━\n"
    "📌 <b>BUTTON SYNTAX</b>\n"
    "• !1link Nama|https://example.com\n"
    "• !1web Nama|https://example.com\n"
    "• !1callback Nama|key\n"
    "• !1share NamaButton\n"
    "• !1withdrawal Withdraw\n"
    "Row ikut nombor !1 !2 !3\n"
    "━━━━━━━━━━━━━━━━━━\n"
    "📌 <b>SETCOMMAND</b>\n"
    "• Reply content + /setcommand hello\n"
    "User type /hello → bot reply content\n"
)

HELP_PLACEHOLDERS_SHORT = (
    "🧠 <b>Quick placeholders</b>\n"
    "• {firstname} • {username} • {member_id}\n"
    "• [balance] • [share] • [link] • {date}\n"
    "• {rand:1-100}\n"
    "• {count} • {limit} • {remaining} • {reset}\n"
    "• [web](https://example.com)\n"
)


def parse_buttons(text_: str, share_inline_query: Optional[str] = None) -> Tuple[str, Optional[dict]]:
    if not text_:
        return "", None

    lines = text_.split("\n")
    visible: List[str] = []
    rows: Dict[int, List[dict]] = {}

    for line in lines:
        if line.startswith("!"):
            m = re.match(r"!(\d+)(link|callback|share|withdrawal|web)\s+(.+)$", line.strip())
            if m:
                row, typ, content = int(m.group(1)), m.group(2), m.group(3).strip()
                rows.setdefault(row, [])

                if typ == "link":
                    if "|" in content:
                        name, url = content.split("|", 1)
                    else:
                        name, url = "Link", content
                    url = url.strip()
                    if not url.startswith("http"):
                        url = "https://" + url
                    rows[row].append({"text": name.strip(), "url": url})

                elif typ == "callback":
                    # Supports:
                    #   !1callback Name|key
                    #   !1callback Name|key delay=5
                    if "|" in content:
                        name, key_raw = content.split("|", 1)
                    else:
                        name, key_raw = content, "error"

                    key_raw = (key_raw or "").strip()
                    delay_override = None
                    mdel = re.search(r"\bdelay\s*=\s*(\d+)\b", key_raw, flags=re.I)
                    if mdel:
                        try:
                            delay_override = int(mdel.group(1))
                        except Exception:
                            delay_override = None
                        key_raw = re.sub(r"\s*\bdelay\s*=\s*\d+\b", "", key_raw, flags=re.I).strip()

                    key_clean = key_raw.strip() or "error"
                    cb = f"cb:{key_clean}"
                    if delay_override is not None:
                        cb = f"{cb};d={delay_override}"
                    rows[row].append({"text": name.strip(), "callback_data": cb})

                elif typ == "share":
                    q = share_inline_query or "Jom join!"
                    # Use Telegram share deep link to avoid @BotUsername prefix
                    import urllib.parse as _up
                    _share_url = _up.quote(q, safe="")
                    rows[row].append({"text": content, "url": f"https://t.me/share/url?url=&text={_share_url}"})

                elif typ == "withdrawal":
                    rows[row].append({"text": content, "callback_data": "req_withdraw"})

                elif typ == "web":
                    if "|" in content:
                        name, url = content.split("|", 1)
                    else:
                        name, url = "Open", content
                    url = url.strip()
                    if not url.startswith("http"):
                        url = "https://" + url
                    rows[row].append({"text": name.strip(), "web_app": {"url": url}})

                continue

        visible.append(line)

    kb = [rows[r] for r in sorted(rows.keys())]
    return "\n".join(visible).strip(), ({"inline_keyboard": kb} if kb else None)


def _h(val) -> str:
    # escape placeholder values sahaja (template text admin masih boleh guna HTML)
    return html.escape("" if val is None else str(val), quote=True)


def _normalize_url(url: str) -> str:
    u = (url or "").strip()
    if not u:
        return ""
    if not re.match(r"^https?://", u, flags=re.I):
        u = "https://" + u
    return u


def _convert_md_links_to_html(text_: str) -> str:
    """
    Convert [label](url) to <a href="url">label</a> for Telegram HTML parse_mode.
    """
    if not text_:
        return ""

    def repl(m):
        label = m.group(1)
        url = m.group(2)
        url = _normalize_url(url)
        if not url:
            return m.group(0)
        return f'<a href="{_h(url)}">{_h(label)}</a>'

    return re.sub(r"\[([^\]\n]{1,120})\]\(([^)\s]+)\)", repl, text_)



def _convert_basic_md_to_html(text_: str) -> str:
    """Convert a tiny subset of Markdown to Telegram-HTML.

    Supported:
    - **bold**  -> <b>bold</b>
    - __italic__ -> <i>italic</i>

    Notes:
    - This is intentionally minimal to avoid breaking normal text that uses '*'.
    - Admin text is treated as trusted input; sanitizer will still escape unknown tags.
    """
    if not text_:
        return ""

    # Bold: **text**
    text_ = re.sub(r"\*\*([^\*\n]{1,300})\*\*", r"<b>\1</b>", text_)
    # Italic: __text__
    text_ = re.sub(r"__([^_\n]{1,300})__", r"<i>\1</i>", text_)
    return text_

def render_placeholders(text_: str, bot_username: str, user_row: dict) -> str:
    if not text_:
        return ""

    u = user_row or {}
    fname = _h(u.get("first_name") or "")
    uname = f"@{_h(u['username'])}" if u.get("username") else fname
    bal = float(u.get("balance") or 0)
    share = int(u.get("shared_count") or 0)
    link = f"https://t.me/{bot_username}?start={u.get('user_id')}" if bot_username else ""

    out = text_.replace("{firstname}", str(fname)).replace("{username}", str(uname))
    out = out.replace("{member_id}", _h(u.get("member_id") or "000000"))
    out = out.replace("[balance]", f"RM{bal:.2f}").replace("[share]", str(share))
    out = out.replace("[link]", _h(link)).replace("{date}", now_local_str(DATE_FMT))
    out = re.sub(r"\{rand:(\d+)-(\d+)\}", lambda m: str(random.randint(int(m.group(1)), int(m.group(2)))), out)

    # NEW: support [web](https://...) style link
    out = _convert_md_links_to_html(out)
    out = _convert_basic_md_to_html(out)
    return out


def make_share_query(bot_username: str, user_row: dict) -> str:
    return render_placeholders("🤫 Rahsia pemain pro — scan RTP real-time + FREE credit untuk kau! Cuba sekarang 🔥\n[link]", bot_username, user_row)


# ---------------------------
# DB HELPERS
# ---------------------------
def is_owner(uid: int, bot_row: dict) -> bool:
    return uid == int(bot_row["owner_id"])


def is_admin(uid: int, bot_id: str) -> bool:
    with engine.connect() as conn:
        res = conn.execute(
            text("SELECT expiry_at FROM admins WHERE bot_id=:b AND admin_user_id=:u"),
            {"b": bot_id, "u": uid},
        ).mappings().first()
    return True if res and (res["expiry_at"] is None or res["expiry_at"] > utcnow()) else False

# ---------------------------
# Scanner (provider media + games)
# ---------------------------

def norm_provider(p: str) -> str:
    p = (p or "").strip().lower()
    # allow letters, numbers, underscore only
    p = re.sub(r"[^a-z0-9_]+", "", p)
    return p


def upsert_scanner_media(conn: sa.engine.Connection, bot_id: str, provider: str, media_type: str, file_id: str) -> None:
    provider = norm_provider(provider)
    conn.execute(
        text(
            """
            INSERT INTO scanner_media (bot_id, provider, media_type, file_id, updated_at)
            VALUES (:bot_id, :provider, :media_type, :file_id, NOW())
            ON CONFLICT (bot_id, provider)
            DO UPDATE SET media_type = EXCLUDED.media_type,
                          file_id = EXCLUDED.file_id,
                          updated_at = NOW()
            """
        ),
        {"bot_id": bot_id, "provider": provider, "media_type": media_type, "file_id": file_id},
    )


def get_scanner_media(conn: sa.engine.Connection, bot_id: str, provider: str) -> Optional[Dict]:
    provider = norm_provider(provider)
    row = conn.execute(
        text(
            """
            SELECT media_type, file_id
            FROM scanner_media
            WHERE bot_id = :bot_id AND provider = :provider
            """
        ),
        {"bot_id": bot_id, "provider": provider},
    ).mappings().first()
    return dict(row) if row else None


def parse_games_text(raw: str) -> List[str]:
    if not raw:
        return []
    games = []
    seen = set()
    for line in raw.splitlines():
        g = line.strip()
        if not g:
            continue
        # keep original casing, but dedupe case-insensitively
        k = g.lower()
        if k in seen:
            continue
        seen.add(k)
        games.append(g)
    return games


def replace_scanner_games(conn: sa.engine.Connection, bot_id: str, provider: str, games: List[str]) -> int:
    provider = norm_provider(provider)
    conn.execute(
        text("DELETE FROM scanner_games WHERE bot_id = :bot_id AND provider = :provider"),
        {"bot_id": bot_id, "provider": provider},
    )
    if not games:
        return 0
    conn.execute(
        text(
            """
            INSERT INTO scanner_games (bot_id, provider, game)
            VALUES (:bot_id, :provider, :game)
            ON CONFLICT DO NOTHING
            """
        ),
        [{"bot_id": bot_id, "provider": provider, "game": g} for g in games],
    )
    return len(games)


def get_scanner_games(conn: sa.engine.Connection, bot_id: str, provider: str) -> List[str]:
    provider = norm_provider(provider)
    rows = conn.execute(
        text(
            """
            SELECT game FROM scanner_games
            WHERE bot_id = :bot_id AND provider = :provider
            ORDER BY game ASC
            """
        ),
        {"bot_id": bot_id, "provider": provider},
    ).scalars().all()
    return list(rows or [])


def scanner_check_and_touch_cooldown(
    conn: sa.engine.Connection, bot_id: str, user_id: int, provider: str, cooldown_seconds: int = 5
) -> int:
    """Returns remaining seconds if still cooling down; otherwise 0 and updates last_at."""
    provider = norm_provider(provider)
    now = datetime.now(timezone.utc)
    row = conn.execute(
        text(
            """
            SELECT last_at FROM scanner_cooldowns
            WHERE bot_id = :bot_id AND user_id = :user_id AND provider = :provider
            """
        ),
        {"bot_id": bot_id, "user_id": user_id, "provider": provider},
    ).mappings().first()

    if row and row.get("last_at"):
        last_at = row["last_at"]
        # ensure tz-aware
        if last_at.tzinfo is None:
            last_at = last_at.replace(tzinfo=timezone.utc)
        elapsed = (now - last_at).total_seconds()
        if elapsed < cooldown_seconds:
            return int(cooldown_seconds - elapsed + 0.999)

    conn.execute(
        text(
            """
            INSERT INTO scanner_cooldowns (bot_id, user_id, provider, last_at)
            VALUES (:bot_id, :user_id, :provider, :last_at)
            ON CONFLICT (bot_id, user_id, provider)
            DO UPDATE SET last_at = EXCLUDED.last_at
            """
        ),
        {"bot_id": bot_id, "user_id": user_id, "provider": provider, "last_at": now},
    )
    return 0


# ---------------------------
# Scanner daily limit (per day)
# ---------------------------
def _today_local_date() -> datetime.date:
    try:
        if LOCAL_TZ:
            return datetime.now(LOCAL_TZ).date()
    except Exception:
        pass
    return datetime.now().date()


def _find_user_id_by_username(conn, bot_id: str, username_no_at: str) -> Optional[int]:
    u = (username_no_at or "").strip().lstrip("@")
    if not u:
        return None
    row = conn.execute(
        text('''
            SELECT user_id FROM users
            WHERE bot_id=:b AND lower(username)=lower(:u)
            ORDER BY joined_at DESC
            LIMIT 1
        '''),
        {"b": bot_id, "u": u},
    ).mappings().first()
    return int(row["user_id"]) if row and row.get("user_id") is not None else None


def get_scan_limit_for_user(conn, bot_row: dict, bot_id: str, user_id: int) -> Optional[int]:
    """Return limit per day for user. Priority: override table, then bot setting."""
    try:
        r = conn.execute(
            text("SELECT limit_per_day FROM scan_limit_overrides WHERE bot_id=:b AND user_id=:u"),
            {"b": bot_id, "u": int(user_id)},
        ).mappings().first()
        if r and r.get("limit_per_day") is not None:
            return int(r["limit_per_day"])
    except Exception:
        pass
    lim = bot_row.get("scan_limit_per_day")
    if lim is None:
        return None
    try:
        return int(lim)
    except Exception:
        return None


def scan_daily_touch_or_block(conn, bot_row: dict, bot_id: str, user_id: int) -> Tuple[bool, int, Optional[int]]:
    """Atomically increment daily scan count if under limit. Returns (allowed, used_after, limit)."""
    lim = get_scan_limit_for_user(conn, bot_row, bot_id, user_id)
    if lim is None:
        return True, 0, None
    try:
        lim_i = int(lim)
    except Exception:
        lim_i = 0
    if lim_i <= 0:
        return True, 0, lim_i

    day = _today_local_date()

    # Atomic upsert only when count < limit
    r = conn.execute(
        text('''
            INSERT INTO scan_daily_usage (bot_id, user_id, day, count)
            VALUES (:b, :u, :d, 1)
            ON CONFLICT (bot_id, user_id, day) DO UPDATE
              SET count = scan_daily_usage.count + 1,
                  updated_at = NOW()
              WHERE scan_daily_usage.count < :lim
            RETURNING count
        '''),
        {"b": bot_id, "u": int(user_id), "d": day, "lim": lim_i},
    ).mappings().first()

    if r and r.get("count") is not None:
        return True, int(r["count"]), lim_i

    # exceeded: read current count
    cur = conn.execute(
        text("SELECT count FROM scan_daily_usage WHERE bot_id=:b AND user_id=:u AND day=:d"),
        {"b": bot_id, "u": int(user_id), "d": day},
    ).mappings().first()
    used = int((cur or {}).get("count") or lim_i)
    return False, used, lim_i


def scan_daily_get_stats(conn, bot_row: dict, bot_id: str, user_id: int) -> Tuple[int, Optional[int], str, str]:
    """Return (used_today, limit_int_or_None, remaining_str, reset_str).

    - used_today: integer count for today (0 if none).
    - limit_int_or_None: None if unlimited, else int (can be <=0 meaning unlimited/off).
    - remaining_str: '∞' if unlimited else remaining count as string.
    - reset_str: next reset timestamp string in local time.
    """
    # limit
    lim = get_scan_limit_for_user(conn, bot_row, bot_id, user_id)
    lim_i: Optional[int] = None
    if lim is not None:
        try:
            lim_i = int(lim)
        except Exception:
            lim_i = 0

    day = _today_local_date()
    cur = conn.execute(
        text("SELECT count FROM scan_daily_usage WHERE bot_id=:b AND user_id=:u AND day=:d"),
        {"b": bot_id, "u": int(user_id), "d": day},
    ).mappings().first()
    used = int((cur or {}).get("count") or 0)

    # remaining
    if lim is None:
        remaining = "∞"
    else:
        try:
            _li = int(lim_i or 0)
        except Exception:
            _li = 0
        if _li <= 0:
            remaining = "∞"
        else:
            remaining = str(max(_li - used, 0))

    # reset time (next local midnight)
    try:
        now = datetime.now(LOCAL_TZ) if LOCAL_TZ else datetime.now()
        next_mid = (now + timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
        reset_str = next_mid.strftime("%d/%m/%Y %H:%M")
    except Exception:
        reset_str = "00:00"

    return used, lim_i, remaining, reset_str


def apply_scan_placeholders(conn, text_: str, bot_row: dict, bot_id: str, user_id: int) -> str:
    """Replace scan placeholders in any template text.

    Supported:
    - {count} or {used}: used scans today
    - {limit}: daily limit or 'UNLIMITED'
    - {remaining}: remaining scans today or '∞'
    - {reset}: next reset local timestamp
    """
    if not text_:
        return ""
    if not any(p in text_ for p in ("{count}", "{used}", "{limit}", "{remaining}", "{reset}")):
        return text_

    used, lim_i, remaining, reset_str = scan_daily_get_stats(conn, bot_row, bot_id, user_id)

    # limit display
    if lim_i is None:
        lim_disp = "UNLIMITED"
    else:
        try:
            li = int(lim_i)
        except Exception:
            li = 0
        lim_disp = "UNLIMITED" if li <= 0 else str(li)

    out = text_
    out = out.replace("{count}", str(used))
    out = out.replace("{used}", str(used))
    out = out.replace("{limit}", lim_disp)
    out = out.replace("{remaining}", remaining)
    out = out.replace("{reset}", reset_str)
    return out


# Cache: last shown games per (bot_id, user_id, provider) to avoid repeats
_scan_last_games: Dict[tuple, set] = {}

def build_scanner_caption(firstname: str, provider_label: str, games: List[str], member_id: str = "", cache_key: tuple = None) -> str:
    """HTML caption — premium scan result. Uses cache_key to avoid repeating games."""
    firstname = firstname or "Boss"
    pool = list(games)
    random.shuffle(pool)

    # Avoid repeating games from last scan
    last_shown = _scan_last_games.get(cache_key, set()) if cache_key else set()
    if last_shown:
        # Put previously shown games at the end so new games come first
        fresh = [g for g in pool if g not in last_shown]
        stale = [g for g in pool if g in last_shown]
        random.shuffle(stale)
        pool = fresh + stale
        # If all games already shown, reset (full shuffle)
        if not fresh:
            random.shuffle(pool)
            last_shown = set()

    # Generate scan ID
    scan_id = f"SC-{random.randint(10000, 99999)}"

    # Timestamp
    try:
        now_local = datetime.now(LOCAL_TZ) if LOCAL_TZ else datetime.now()
        stamp = now_local.strftime("%d %b %Y %H:%M")
    except Exception:
        stamp = datetime.now().strftime("%d %b %Y %H:%M")

    sep = "━━━━━━━━━━━━━━━━━━"
    mid_str = f" | 🆔 {html.escape(member_id)}" if member_id else ""

    # Build header
    header = (
        f"🔍 <b>SCAN RESULT — {html.escape(provider_label)}</b>\n"
        f"{sep}\n"
        f"📋 Scan ID: <code>#{scan_id}</code>\n"
        f"👤 <b>{html.escape(firstname)}</b>{mid_str}"
    )

    # Pre-generate percentages with a more realistic spread
    game_pcts = []
    for g in pool:
        roll = random.random()
        if roll < 0.15:
            pct = random.randint(80, 95)
        elif roll < 0.65:
            pct = random.randint(60, 79)
        else:
            pct = random.randint(34, 59)
        game_pcts.append((g, pct))

    greens = sorted([item for item in game_pcts if item[1] >= 80], key=lambda item: item[1], reverse=True)
    yellows = sorted([item for item in game_pcts if 60 <= item[1] < 80], key=lambda item: item[1], reverse=True)
    reds = sorted([item for item in game_pcts if item[1] < 60], key=lambda item: item[1], reverse=True)

    # Guarantee at least some red results for the visible list when game pool is big enough
    min_reds = 2 if len(pool) >= 10 else (1 if len(pool) >= 5 else 0)
    while len(reds) < min_reds and yellows:
        g, _old_pct = yellows.pop()
        new_pct = random.randint(40, 59)
        reds.append((g, new_pct))
    reds.sort(key=lambda item: item[1], reverse=True)

    # Visible ordering: greens first, then yellows, then reds so color pattern remains obvious in caption
    visible_green_cap = max(2, min(4, max(1, len(pool) // 7)))
    visible_yellow_cap = max(3, min(6, max(2, len(pool) // 4)))
    game_pcts = (
        greens[:visible_green_cap]
        + yellows[:visible_yellow_cap]
        + reds
        + yellows[visible_yellow_cap:]
        + greens[visible_green_cap:]
    )
    total_scanned = len(pool)

    # Build a max-sized footer first to know exact reserved space
    footer_sample = (
        f"{sep}\n"
        f"📊 Scanned: <b>{total_scanned}</b> | 🔥 Hot: <b>99</b> | ⚡ Best: <b>95%</b>\n"
        f"🕒 <i>{html.escape(stamp)}</i>\n"
        f"⚠️ <i>Valid 15 minit sahaja</i>"
    )

    # Exact budget: header + sep-divider + footer + newlines between sections
    reserved = len(header) + 1 + len(sep) + 1 + len(footer_sample) + 1
    budget = TG_MAX_CAPTION - reserved

    # Pick games within budget
    chosen = []
    used = 0
    for g, pct in game_pcts:
        g_esc = html.escape(g)
        if pct >= 80:
            line = f"🟢 <b>{g_esc}</b> — <b>{pct}%</b>"
        elif pct >= 60:
            line = f"🟡 {g_esc} — {pct}%"
        else:
            line = f"🔴 {g_esc} — {pct}%"
        if used + len(line) + 1 > budget:
            break
        chosen.append((g, line, pct))
        used += len(line) + 1

    # Save shown games to cache for next scan
    if cache_key:
        _scan_last_games[cache_key] = {g for g, _, _ in chosen}

    # Stats from chosen games
    hot_count = sum(1 for _, _, p in chosen if p >= 80)
    best_pct = max((p for _, _, p in chosen), default=0)

    # Build actual footer
    footer = (
        f"{sep}\n"
        f"📊 Scanned: <b>{total_scanned}</b> | 🔥 Hot: <b>{hot_count}</b> | ⚡ Best: <b>{best_pct}%</b>\n"
        f"🕒 <i>{html.escape(stamp)}</i>\n"
        f"⚠️ <i>Valid 15 minit sahaja</i>"
    )

    lines_out = [header, sep] + [line for _, line, _ in chosen] + [footer]
    return "\n".join(lines_out)


def build_scanner_text_result(firstname: str, provider_label: str, games: List[str], member_id: str = "", cache_key: tuple = None) -> str:
    """Build final result as plain TEXT (up to 4096 chars). Uses same logic as build_scanner_caption but without the 1024 caption limit."""
    firstname = firstname or "Boss"
    pool = list(games)
    random.shuffle(pool)

    # Avoid repeating games from last scan
    last_shown = _scan_last_games.get(cache_key, set()) if cache_key else set()
    if last_shown:
        fresh = [g for g in pool if g not in last_shown]
        stale = [g for g in pool if g in last_shown]
        random.shuffle(stale)
        pool = fresh + stale
        if not fresh:
            random.shuffle(pool)
            last_shown = set()

    scan_id = f"SC-{random.randint(10000, 99999)}"

    try:
        now_local = datetime.now(LOCAL_TZ) if LOCAL_TZ else datetime.now()
        stamp = now_local.strftime("%d %b %Y %H:%M")
    except Exception:
        stamp = datetime.now().strftime("%d %b %Y %H:%M")

    sep = "━━━━━━━━━━━━━━━━━━"
    mid_str = f" | 🆔 {html.escape(member_id)}" if member_id else ""

    header = (
        f"🔍 <b>SCAN RESULT — {html.escape(provider_label)}</b>\n"
        f"{sep}\n"
        f"📋 Scan ID: <code>#{scan_id}</code>\n"
        f"👤 <b>{html.escape(firstname)}</b>{mid_str}"
    )

    # Pre-generate percentages with a more realistic spread
    game_pcts = []
    for g in pool:
        roll = random.random()
        if roll < 0.15:
            pct = random.randint(80, 95)
        elif roll < 0.65:
            pct = random.randint(60, 79)
        else:
            pct = random.randint(34, 59)
        game_pcts.append((g, pct))

    greens = sorted([item for item in game_pcts if item[1] >= 80], key=lambda item: item[1], reverse=True)
    yellows = sorted([item for item in game_pcts if 60 <= item[1] < 80], key=lambda item: item[1], reverse=True)
    reds = sorted([item for item in game_pcts if item[1] < 60], key=lambda item: item[1], reverse=True)

    min_reds = 2 if len(pool) >= 10 else (1 if len(pool) >= 5 else 0)
    while len(reds) < min_reds and yellows:
        g, _old_pct = yellows.pop()
        new_pct = random.randint(40, 59)
        reds.append((g, new_pct))
    reds.sort(key=lambda item: item[1], reverse=True)

    game_pcts = (
        greens[:max(2, min(4, max(1, len(pool) // 7)))]
        + yellows[:max(3, min(6, max(2, len(pool) // 4)))]
        + reds
        + yellows[max(3, min(6, max(2, len(pool) // 4))):]
        + greens[max(2, min(4, max(1, len(pool) // 7))):]
    )
    total_scanned = len(pool)

    # Budget = 4096 (TG_MAX_TEXT)
    footer_sample = (
        f"{sep}\n"
        f"📊 Scanned: <b>{total_scanned}</b> | 🔥 Hot: <b>99</b> | ⚡ Best: <b>95%</b>\n"
        f"🕒 <i>{html.escape(stamp)}</i>\n"
        f"⚠️ <i>Valid 15 minit sahaja</i>"
    )

    reserved = len(header) + 1 + len(sep) + 1 + len(footer_sample) + 1
    budget = TG_MAX_TEXT - reserved

    chosen = []
    used = 0
    for g, pct in game_pcts:
        g_esc = html.escape(g)
        if pct >= 80:
            line = f"🟢 <b>{g_esc}</b> — <b>{pct}%</b>"
        elif pct >= 60:
            line = f"🟡 {g_esc} — {pct}%"
        else:
            line = f"🔴 {g_esc} — {pct}%"
        if used + len(line) + 1 > budget:
            break
        chosen.append((g, line, pct))
        used += len(line) + 1

    if cache_key:
        _scan_last_games[cache_key] = {g for g, _, _ in chosen}

    hot_count = sum(1 for _, _, p in chosen if p >= 80)
    best_pct = max((p for _, _, p in chosen), default=0)

    footer = (
        f"{sep}\n"
        f"📊 Scanned: <b>{total_scanned}</b> | 🔥 Hot: <b>{hot_count}</b> | ⚡ Best: <b>{best_pct}%</b>\n"
        f"🕒 <i>{html.escape(stamp)}</i>\n"
        f"⚠️ <i>Valid 15 minit sahaja</i>"
    )

    lines_out = [header, sep] + [line for _, line, _ in chosen] + [footer]
    return "\n".join(lines_out)


def build_scanner_result_keyboard(provider: str) -> dict:
    """Inline keyboard untuk result scanner: Scan Kembali + Kembali ke Menu Scanner."""
    provider_clean = norm_provider(provider)
    key = provider_clean or provider
    return {
        "inline_keyboard": [[
            {"text": "🟢 Scan Kembali", "callback_data": f"cb:scan_{key}"},
            {"text": "⬅️ Kembali", "callback_data": "cb:menuscanner"},
        ]]
    }


def send_scanner_result(token: str, chat_id: int, firstname: str, provider: str, media: Dict, games: List[str], member_id: str = "", cache_key: tuple = None) -> None:
    provider_clean = norm_provider(provider)
    provider_label = provider_clean.upper() if provider_clean else provider
    caption = build_scanner_caption(firstname, provider_label, games, member_id=member_id, cache_key=cache_key)

    kb = build_scanner_result_keyboard(provider)

    media_type = media.get("media_type")
    file_id = media.get("file_id")

    # give some "animation"
    try:
        requests.post(
            f"https://api.telegram.org/bot{token}/sendChatAction",
            json={"chat_id": chat_id, "action": "typing"},
            timeout=10,
        )
    except Exception:
        pass

    if media_type == "photo":
        tg_send_photo(token, chat_id, file_id, caption=caption, reply_markup=kb, parse_mode="HTML")
    elif media_type == "video":
        tg_send_video(token, chat_id, file_id, caption=caption, reply_markup=kb, parse_mode="HTML")
    elif media_type == "animation":
        tg_send_animation(token, chat_id, file_id, caption=caption, reply_markup=kb, parse_mode="HTML")
    else:
        # fallback document
        tg_send_document(token, chat_id, file_id, caption=caption, reply_markup=kb, parse_mode="HTML")



# ---------------------------
# Scanner animation (BM rotation + progress bar) + edit in-place
# ---------------------------
SCAN_BM_FRAMES = [
    # Server & Connection (20)
    "📊 <b>sambung ke server TELETHON...</b>",
    "🔌 <b>Menghubungi server utama...</b>",
    "🌐 <b>Menyambung ke database pusat...</b>",
    "📡 <b>Mengakses API {prov}...</b>",
    "🛰️ <b>Sinkronisasi dengan cloud server...</b>",
    "⚡ <b>Menjalin sambungan SSL...</b>",
    "🔐 <b>Verifikasi koneksi selamat...</b>",
    "🌍 <b>Routing ke server {prov}...</b>",
    "📶 <b>Stabilkan isyarat...</b>",
    "🔗 <b>Membina jalur data...</b>",
    "💫 <b>Optimizing bandwidth...</b>",
    "🚀 <b>Boosting connection speed...</b>",
    "🔄 <b>Reconnecting to main hub...</b>",
    "📨 <b>Fetching live data feed...</b>",
    "🎯 <b>Locking target server...</b>",
    "⚙️ <b>Konfigurasi protokol...</b>",
    "🌟 <b>Establishing secure tunnel...</b>",
    "💻 <b>Handshake dengan {prov}...</b>",
    "🔋 <b>Powering up connection...</b>",
    "📍 <b>Pinpointing data source...</b>",
    
    # Data Processing (20)
    "🧾 <b>Sedang bypass data {prov}</b>...",
    "🔍 <b>Menyusun Data hacking...</b>",
    "📊 <b>Menganalisa game patterns...</b>",
    "🎲 <b>Scanning RTP database...</b>",
    "💎 <b>Extracting premium data...</b>",
    "🧮 <b>Calculating win rates...</b>",
    "📈 <b>Parsing statistical models...</b>",
    "🔬 <b>Deep scan algoritma {prov}...</b>",
    "🎰 <b>Decrypting game matrices...</b>",
    "💡 <b>Processing AI predictions...</b>",
    "🧩 <b>Assembling data fragments...</b>",
    "📋 <b>Indexing game catalog...</b>",
    "🔎 <b>Validating data integrity...</b>",
    "⚗️ <b>Filtering noise data...</b>",
    "🎯 <b>Targeting high RTP games...</b>",
    "📦 <b>Packaging scan results...</b>",
    "🧪 <b>Testing data accuracy...</b>",
    "🗂️ <b>Organizing game list...</b>",
    "💾 <b>Caching frequent queries...</b>",
    "🔐 <b>Encrypting sensitive info...</b>",
    
    # System Operations (20)
    "⏳ <b>Sila tunggu sebentar...</b>",
    "⌛ <b>Processing request...</b>",
    "🕐 <b>Hampir siap...</b>",
    "⚡ <b>Speeding up process...</b>",
    "🔄 <b>Refreshing cache...</b>",
    "💫 <b>Optimizing results...</b>",
    "🎨 <b>Formatting display...</b>",
    "🧹 <b>Cleaning temp data...</b>",
    "🔧 <b>Fine-tuning parameters...</b>",
    "⚙️ <b>Calibrating sensors...</b>",
    "🎛️ <b>Adjusting settings...</b>",
    "📊 <b>Generating graphs...</b>",
    "🖥️ <b>Rendering UI...</b>",
    "🎬 <b>Preparing animation...</b>",
    "🎪 <b>Setting up display...</b>",
    "🔮 <b>Finalizing prediction...</b>",
    "✨ <b>Polishing output...</b>",
    "🎁 <b>Wrapping results...</b>",
    "🏁 <b>Almost done...</b>",
    "⏱️ <b>Final countdown...</b>",
    
    # Hacking/Bypass Theme (20)
    "🔓 <b>Bypassing firewall {prov}...</b>",
    "🛡️ <b>Cracking security layer...</b>",
    "🎭 <b>Masking IP address...</b>",
    "👾 <b>Injecting bypass script...</b>",
    "🕵️ <b>Stealth mode activated...</b>",
    "🔑 <b>Unlocking restricted data...</b>",
    "🚪 <b>Opening backdoor access...</b>",
    "🎪 <b>Circumventing detector...</b>",
    "🌀 <b>Tunneling through proxy...</b>",
    "🔒 <b>Breaking encryption...</b>",
    "⚔️ <b>Penetrating defense...</b>",
    "🎯 <b>Exploiting vulnerability...</b>",
    "🧨 <b>Defusing anti-bot...</b>",
    "🎲 <b>Randomizing signature...</b>",
    "👻 <b>Going ghost mode...</b>",
    "🦾 <b>Brute-forcing gateway...</b>",
    "🔨 <b>Hammering weak point...</b>",
    "⚡ <b>Flash attack initiated...</b>",
    "🎪 <b>Spoofing credentials...</b>",
    "🔮 <b>Magic bypass enabled...</b>",
    
    # Provider-Specific (20)
    "🎰 <b>Scanning {prov} slots...</b>",
    "🎮 <b>Loading {prov} games...</b>",
    "🃏 <b>Shuffling {prov} deck...</b>",
    "🎲 <b>Rolling {prov} dice...</b>",
    "💰 <b>Checking {prov} jackpot...</b>",
    "🏆 <b>Hunting {prov} bonuses...</b>",
    "🎁 <b>Unwrapping {prov} prizes...</b>",
    "⭐ <b>Rating {prov} games...</b>",
    "🔥 <b>Finding {prov} hot games...</b>",
    "❄️ <b>Avoiding {prov} cold slots...</b>",
    "📈 <b>Tracking {prov} trends...</b>",
    "🎯 <b>Targeting {prov} features...</b>",
    "💎 <b>Mining {prov} gems...</b>",
    "🌟 <b>Discovering {prov} secrets...</b>",
    "🔍 <b>Investigating {prov} RTP...</b>",
    "📊 <b>Benchmarking {prov} rates...</b>",
    "🎪 <b>Unveiling {prov} mystery...</b>",
    "🎨 <b>Styling {prov} results...</b>",
    "🏅 <b>Ranking {prov} winners...</b>",
    "🎉 <b>Celebrating {prov} finds...</b>",
    "🧠 <b>Menganalisis corak kemenangan {prov}...</b>",
    "🛰️ <b>Mengimbas signal RTP tersembunyi {prov}...</b>",
    "🧬 <b>Menyusun DNA pattern slot {prov}...</b>",
    "📡 <b>Mengesan momentum spin aktif {prov}...</b>",
    "🎯 <b>Menapis result paling padu untuk {prov}...</b>",
]

SCAN_DURATION_PRESETS = (15, 25, 40, 60)
DEFAULT_SCANNER_DURATION_SECONDS = 25
SCANNER_LOADING_TEXT_DELAY_SECONDS = 5


def get_scanner_duration_seconds(bot_row: Optional[dict]) -> int:
    raw = (bot_row or {}).get("scanner_duration_seconds")
    try:
        val = int(raw)
    except Exception:
        val = DEFAULT_SCANNER_DURATION_SECONDS
    if val not in SCAN_DURATION_PRESETS:
        return DEFAULT_SCANNER_DURATION_SECONDS
    return val


def scanner_duration_label(seconds: int) -> str:
    sec = int(seconds or DEFAULT_SCANNER_DURATION_SECONDS)
    return f"{sec}s"


def scanner_loading_text_count(seconds: int) -> int:
    try:
        sec = int(seconds or DEFAULT_SCANNER_DURATION_SECONDS)
    except Exception:
        sec = DEFAULT_SCANNER_DURATION_SECONDS
    if sec <= 0:
        sec = DEFAULT_SCANNER_DURATION_SECONDS
    return max(1, int(sec / SCANNER_LOADING_TEXT_DELAY_SECONDS))

def animate_scanning_progress(token: str, chat_id: int, message_id: int, provider: str = "", total_seconds: int = DEFAULT_SCANNER_DURATION_SECONDS) -> None:
    """Edit mesej scanner ikut tempoh preset dengan 1 text = 5 saat fixed."""
    if not chat_id or not message_id:
        return
    prov = (provider or "").strip().upper()
    try:
        total_seconds = int(total_seconds or DEFAULT_SCANNER_DURATION_SECONDS)
    except Exception:
        total_seconds = DEFAULT_SCANNER_DURATION_SECONDS
    if total_seconds <= 0:
        total_seconds = DEFAULT_SCANNER_DURATION_SECONDS

    total_steps = scanner_loading_text_count(total_seconds)
    delay = float(SCANNER_LOADING_TEXT_DELAY_SECONDS)
    icons = ["🕐", "🕑", "🕒", "🕓", "🕔", "🕕", "🕖", "🕗", "🕘", "🕙", "🕚", "🕛"]

    for i in range(total_steps):
        ico = icons[i % len(icons)]
        line = random.choice(SCAN_BM_FRAMES).format(prov=prov)
        pct = int(((i + 1) / total_steps) * 100)
        filled = max(0, min(10, pct // 10))
        bar = "▓" * filled + "░" * (10 - filled)
        html_text = f"{ico} {line}\n<code>[{bar}] {pct}%</code>"

        ok = edit_message(token, chat_id, message_id, html_text, parse_mode="HTML")
        if not ok:
            edit_caption(token, chat_id, message_id, html_text, parse_mode="HTML")

        time.sleep(delay)

def _coerce_media_dict(media):
    """Accept dict or file_id string. Return normalized dict."""
    if isinstance(media, dict):
        return media
    if isinstance(media, str) and media.strip():
        return {"media_type": "photo", "file_id": media.strip()}
    return {}

def send_scanner_result_edit(token: str, chat_id: int, message_id: int, firstname: str, provider: str, media, games: List[str], member_id: str = "", cache_key: tuple = None) -> bool:
    """Try to edit current message into scanner result (media+caption). Return True if edited."""
    media = _coerce_media_dict(media)
    provider_clean = norm_provider(provider)
    provider_label = provider_clean.upper() if provider_clean else provider
    caption = build_scanner_caption(firstname, provider_label, games, member_id=member_id, cache_key=cache_key)

    kb = build_scanner_result_keyboard(provider)

    media_type = (media.get("media_type") or "").strip().lower()
    file_id = (media.get("file_id") or "").strip()

    # Prefer media edit if possible
    if media_type and file_id:
        try:
            edit_media(token, chat_id, message_id, media_type, file_id, caption=caption, reply_markup=kb, parse_mode="HTML")
            return True
        except Exception:
            pass

    # Fallback: caption edit (if current is already media)
    try:
        ok = edit_caption(token, chat_id, message_id, caption, reply_markup=kb, parse_mode="HTML")
        if ok:
            return True
    except Exception:
        pass

    # Fallback: text edit (may fail if current message is media)
    try:
        ok = edit_message(token, chat_id, message_id, caption, reply_markup=kb, parse_mode="HTML")
        if ok:
            return True
    except Exception:
        pass

    # Last resort: delete old message and send new one
    try:
        delete_message(token, chat_id, message_id)
        send_message(token, chat_id, caption or " ", reply_markup=kb, parse_mode="HTML")
        return True
    except Exception:
        return False

def run_scanner_flow(token: str, chat_id: int, message_id: int, firstname: str, provider: str, media, games: List[str], member_id: str = "", cache_key: tuple = None, bot_row: Optional[dict] = None) -> None:
    """Single scanner flow entrypoint: duration-aware loading animation, then final media+caption result."""
    scan_duration = get_scanner_duration_seconds(bot_row)

    try:
        animate_scanning_progress(
            token,
            chat_id,
            message_id,
            provider=provider,
            total_seconds=scan_duration,
        )
    except Exception:
        pass

    if not send_scanner_result_edit(token, chat_id, message_id, firstname, provider, media, games, member_id=member_id, cache_key=cache_key):
        send_scanner_result(token, chat_id, firstname, provider, _coerce_media_dict(media), games, member_id=member_id, cache_key=cache_key)


def require_admin(bot_row: dict, uid: int) -> bool:
    return is_owner(uid, bot_row) or is_admin(uid, str(bot_row["id"]))


def get_bot_by_secret(secret: str):
    with engine.connect() as conn:
        return conn.execute(text("SELECT * FROM bots WHERE secret_token=:s"), {"s": secret}).mappings().first()


def get_bot_by_id(bot_id: str):
    with engine.connect() as conn:
        return conn.execute(text("SELECT * FROM bots WHERE id=:i"), {"i": bot_id}).mappings().first()


def get_bot_by_token(token_: str):
    with engine.connect() as conn:
        return conn.execute(text("SELECT * FROM bots WHERE token=:t"), {"t": token_}).mappings().first()


def get_bot_by_username(username: str):
    """Lookup bot by @username (case-insensitive, strips @)."""
    u = (username or "").strip().lstrip("@").lower()
    if not u:
        return None
    with engine.connect() as conn:
        return conn.execute(
            text("SELECT * FROM bots WHERE LOWER(bot_username)=:u"),
            {"u": u},
        ).mappings().first()


def list_bots_by_owner(owner_id: int):
    with engine.connect() as conn:
        return conn.execute(
            text("""
                SELECT id, bot_username, owner_id, lock_bot, admin_group_id, created_at
                FROM bots
                WHERE owner_id=:o
                ORDER BY created_at DESC
                LIMIT 50
            """),
            {"o": owner_id},
        ).mappings().all()


def get_user_row(bot_id: str, uid: int):
    with engine.connect() as conn:
        return conn.execute(
            text("SELECT * FROM users WHERE bot_id=:b AND user_id=:u"),
            {"b": bot_id, "u": uid},
        ).mappings().first()


def upsert_user(bot_id: str, user: dict, upline_id: Optional[int], affiliate_amount: Optional[float] = None):
    uid = int(user["id"])
    if upline_id == uid:
        upline_id = None
    new_mid = str(random.randint(100000, 999999))

    with engine.begin() as conn:
        res = conn.execute(text("""
            INSERT INTO users (bot_id, user_id, username, first_name, member_id, joined_at, upline_user_id)
            VALUES (:b, :u, :un, :fn, :mid, NOW(), :up)
            ON CONFLICT (bot_id, user_id) DO NOTHING
        """), {
            "b": bot_id,
            "u": uid,
            "un": user.get("username"),
            "fn": user.get("first_name", ""),
            "mid": new_mid,
            "up": upline_id
        })

        is_new = (res.rowcount == 1)
        if not is_new:
            conn.execute(text("""
                UPDATE users SET username=:un, first_name=:fn
                WHERE bot_id=:b AND user_id=:u
            """), {"un": user.get("username"), "fn": user.get("first_name", ""), "b": bot_id, "u": uid})

            conn.execute(text("""
                UPDATE users SET member_id=:m
                WHERE bot_id=:b AND user_id=:u AND (member_id IS NULL OR member_id='')
            """), {"m": new_mid, "b": bot_id, "u": uid})

        if is_new and upline_id:
            upd = conn.execute(text("""
                UPDATE users
                SET balance=balance+:a, shared_count=shared_count+1
                WHERE bot_id=:b AND user_id=:up
            """), {"a": float(affiliate_amount if affiliate_amount is not None else AFFILIATE_AMOUNT), "b": bot_id, "up": upline_id})
            if upd.rowcount == 1:
                conn.execute(
                    text("UPDATE users SET credited_upline=TRUE WHERE bot_id=:b AND user_id=:u"),
                    {"b": bot_id, "u": uid},
                )

        row = conn.execute(
            text("SELECT * FROM users WHERE bot_id=:b AND user_id=:u"),
            {"b": bot_id, "u": uid},
        ).mappings().first()

    return row, is_new


def set_user_state(bot_id, uid, state, payload=None):
    with engine.begin() as conn:
        conn.execute(text("""
            INSERT INTO user_states (bot_id, user_id, state, payload, updated_at)
            VALUES (:b, :u, :s, CAST(:p AS jsonb), NOW())
            ON CONFLICT (bot_id, user_id) DO UPDATE
            SET state=excluded.state, payload=excluded.payload, updated_at=NOW()
        """), {"b": bot_id, "u": uid, "s": state, "p": json.dumps(payload or {})})


def get_user_state(bot_id, uid):
    with engine.connect() as conn:
        return conn.execute(
            text("SELECT * FROM user_states WHERE bot_id=:b AND user_id=:u"),
            {"b": bot_id, "u": uid},
        ).mappings().first()


def clear_user_state(bot_id, uid):
    with engine.begin() as conn:
        conn.execute(text("DELETE FROM user_states WHERE bot_id=:b AND user_id=:u"), {"b": bot_id, "u": uid})


def save_content_from_reply(reply_msg):
    # Preserve Telegram formatting (bold/italic/link/etc) when admin replies
    txt_raw = reply_msg.get("text") or reply_msg.get("caption") or ""
    ents = reply_msg.get("entities") if reply_msg.get("text") else reply_msg.get("caption_entities")
    txt = entities_to_html(txt_raw, ents)

    mt, mid = None, None
    if reply_msg.get("photo"):
        mt, mid = "photo", reply_msg["photo"][-1]["file_id"]
    elif reply_msg.get("video"):
        mt, mid = "video", reply_msg["video"]["file_id"]
    elif reply_msg.get("animation"):
        mt, mid = "animation", reply_msg["animation"]["file_id"]
    elif reply_msg.get("document"):
        mt, mid = "document", reply_msg["document"]["file_id"]
    return mt, mid, txt


def merge_command_extra(base_text: str, command_text: str, command_name: str) -> str:
    """Merge extra lines written under a command message into the replied content."""
    raw = (command_text or "").replace("\r\n", "\n")
    if not raw:
        return base_text or ""

    lines = raw.split("\n")
    extra_lines = lines[1:] if lines and lines[0].strip().startswith(command_name) else lines
    extra = "\n".join(extra_lines).strip()
    if not extra:
        return base_text or ""
    if not (base_text or "").strip():
        return extra
    return (base_text.rstrip() + "\n\n" + extra).strip()


def _message_has_media(msg: dict) -> bool:
    if not msg:
        return False
    return bool(msg.get("photo") or msg.get("video") or msg.get("animation") or msg.get("document"))


# ---------------------------
# JOIN LOCK + VERIFY + PREMIUM GATE
# ---------------------------
def tg_get_chat_member(bot_token: str, chat_ref: str, user_id: int):
    data = {"chat_id": chat_ref, "user_id": user_id}
    return tg_call(bot_token, "getChatMember", data=data)


def tg_get_chat(bot_token: str, chat_ref: str):
    data = {"chat_id": chat_ref}
    return tg_call(bot_token, "getChat", data=data)


def parse_join_targets(raw: Optional[str]) -> List[str]:
    if not raw:
        return []
    items = []
    for part in re.split(r"[\n,]+", raw):
        s = (part or "").strip()
        if s:
            items.append(s)
    return items


def build_join_buttons(bot_token: str, targets: List[str]) -> Tuple[List[List[dict]], List[str]]:
    rows: List[List[dict]] = []
    unresolved: List[str] = []
    for t in targets[:8]:
        target = (t or "").strip()
        if not target:
            continue

        btn = None
        if target.startswith("@"):
            btn = {"text": f"✅ Join {target}", "url": f"https://t.me/{target.lstrip('@')}"}
        else:
            try:
                chat = tg_get_chat(bot_token, target) or {}
                username = (chat.get("username") or "").strip()
                invite_link = (chat.get("invite_link") or "").strip()
                title = (chat.get("title") or chat.get("username") or target).strip()
                if username:
                    btn = {"text": f"✅ Join {title}", "url": f"https://t.me/{username}"}
                elif invite_link:
                    btn = {"text": f"✅ Join {title}", "url": invite_link}
            except Exception:
                btn = None

        if btn:
            rows.append([btn])
        else:
            unresolved.append(target)

    return rows, unresolved


def merge_inline_keyboards(*markups) -> Optional[dict]:
    rows: List[list] = []
    for mk in markups:
        if not mk:
            continue
        for row in (mk.get("inline_keyboard") or []):
            if row:
                rows.append(row)
    return {"inline_keyboard": rows} if rows else None


def build_join_keyboard(bot_token: str, targets: List[str], extra_markup: Optional[dict] = None) -> Tuple[Optional[dict], List[str]]:
    join_rows, unresolved = build_join_buttons(bot_token, targets)
    join_markup = {"inline_keyboard": join_rows} if join_rows else None
    recheck_markup = {"inline_keyboard": [[{"text": "🔁 Saya Dah Join", "callback_data": "gate:recheck"}]]}
    merged = merge_inline_keyboards(extra_markup, join_markup, recheck_markup)
    return merged, unresolved


def ensure_joined(bot_row: dict, chat_id: int, uid: int, user_row: Optional[dict] = None) -> bool:
    if not bot_row.get("join_lock"):
        return True

    targets = parse_join_targets(bot_row.get("join_targets"))
    if not targets:
        return True

    token = bot_row["token"]
    missing = []
    for t in targets:
        try:
            res = tg_get_chat_member(token, t, uid)
            if not res:
                missing.append(t)
                continue
            status = (res.get("status") or "").lower()
            if status in ("left", "kicked"):
                missing.append(t)
        except Exception:
            missing.append(t)

    if not missing:
        return True

    msg_template = bot_row.get("join_message") or (
        "🧲 <b>AKSES TERKUNCI</b>\n"
        "Bossku kena join channel/group dulu baru boleh guna bot 😘\n\n"
        "Sila join:\n"
        + "\n".join([f"• <code>{html.escape(x)}</code>" for x in missing])
        + "\n\nLepas join, tekan <b>🔁 Saya Dah Join</b>."
    )

    join_user = user_row or {
        "user_id": uid,
        "first_name": "Boss",
        "username": None,
        "balance": 0,
        "shared_count": 0,
        "member_id": "000000",
    }
    msg_rendered = render_placeholders(msg_template, bot_row.get("bot_username") or "", join_user)
    share_q = make_share_query(bot_row.get("bot_username") or "", join_user)
    msg_text, custom_markup = parse_buttons(msg_rendered, share_inline_query=share_q)
    kb, unresolved = build_join_keyboard(token, missing, extra_markup=custom_markup)

    if bot_row.get("join_message") and unresolved:
        unresolved_txt = "\n".join([f"• <code>{html.escape(x)}</code>" for x in unresolved])
        if unresolved_txt:
            msg_text = (msg_text + "\n\nSila join:\n" + unresolved_txt + "\n\nLepas join, tekan <b>🔁 Saya Dah Join</b>.").strip()

    mt = bot_row.get("join_message_media_type")
    mf = bot_row.get("join_message_media_file_id")
    if mt and mf:
        send_media(token, chat_id, mt, mf, caption=msg_text, reply_markup=kb, parse_mode="HTML")
    else:
        send_message(token, chat_id, msg_text, reply_markup=kb, parse_mode="HTML")
    return False


def ensure_contact_verified(bot_row: dict, chat_id: int, user_row: dict) -> bool:
    if not bot_row.get("lock_bot"):
        return True
    if user_row and user_row.get("is_verified"):
        return True

    kb = {
        "keyboard": [[{"text": "📲 SHARE CONTACT", "request_contact": True}]],
        "resize_keyboard": True,
        "one_time_keyboard": True
    }
    msg = bot_row.get("contact_message") or "🔒 <b>BOT DIKUNCI</b>\nSila sahkan nombor telefon anda."
    mt = bot_row.get("contact_message_media_type")
    mf = bot_row.get("contact_message_media_file_id")
    if mt and mf:
        send_media(bot_row["token"], chat_id, mt, mf, caption=msg, reply_markup=kb, parse_mode="HTML")
    else:
        send_message(bot_row["token"], chat_id, msg, reply_markup=kb, parse_mode="HTML")
    return False


def ensure_premium_if_needed(bot_row: dict, chat_id: int, uid: int, user_row: dict) -> bool:
    # Manual approval = premium gating
    if not bot_row.get("manual_approval"):
        return True

    # Admin/Owner bypass: jangan sangkut manual approve untuk admin
    try:
        if require_admin(bot_row, int(uid)):
            return True
    except Exception:
        pass

    if user_row and user_row.get("is_premium"):
        return True

    # 1) Tell user they're pending
    msg = bot_row.get("pending_message") or (
        "⏳ <b>SEMAKAN PREMIUM</b>\n"
        "Bossku, request kau dah masuk. Tunggu admin approve dulu ya 😘\n"
        "Lepas approve, bot akan bagi akses premium terus."
    )
    mt = bot_row.get("pending_message_media_type")
    mf = bot_row.get("pending_message_media_file_id")
    if mt and mf:
        send_media(bot_row["token"], chat_id, mt, mf, caption=msg, parse_mode="HTML")
    else:
        send_message(bot_row["token"], chat_id, msg, parse_mode="HTML")

    # 2) Notify admin/owner with Approve/Reject buttons
    try:
        send_premium_request_to_admin(bot_row, int(uid), user_row or {})
    except Exception:
        pass

    return False


def ensure_access(bot_row: dict, chat_id: int, uid: int, user_row: dict) -> bool:
    if not ensure_joined(bot_row, chat_id, uid, user_row):
        return False
    if not ensure_contact_verified(bot_row, chat_id, user_row):
        return False
    if not ensure_premium_if_needed(bot_row, chat_id, uid, user_row):
        return False
    return True




def get_bot_affiliate_amount(bot_row: dict) -> float:
    try:
        v = bot_row.get("affiliate_amount")
        if v is None:
            return float(AFFILIATE_AMOUNT)
        return float(v)
    except Exception:
        return float(AFFILIATE_AMOUNT)


def get_bot_min_withdraw(bot_row: dict) -> float:
    try:
        v = bot_row.get("min_withdraw_amount")
        if v is None:
            return float(MIN_WITHDRAW_DEFAULT)
        return float(v)
    except Exception:
        return float(MIN_WITHDRAW_DEFAULT)



def render_withdrawal_template(tpl: str, amount: float, bal_before: float, bal_after: float) -> str:
    """Simple template rendering for withdrawal approve/reject messages."""
    if not tpl:
        return ""
    out = tpl
    out = out.replace("{amount}", f"RM{float(amount):.2f}")
    out = out.replace("{balance}", f"RM{float(bal_before):.2f}")
    out = out.replace("{balance_after}", f"RM{float(bal_after):.2f}")
    return out

def build_withdraw_insufficient_msg(min_wd: float, bal: float, bot_row: dict = None) -> str:
    """Build withdrawal insufficient balance message. Supports custom template from bot config."""
    if bot_row:
        tpl = (bot_row.get("withdrawal_failed_message") or "").strip()
        if tpl:
            return (tpl
                    .replace("{min_withdraw}", f"RM{min_wd:.2f}")
                    .replace("{balance}", f"RM{bal:.2f}")
                    .replace("{minimum}", f"RM{min_wd:.2f}")
                    )
    return (
        f"❌ Baki tidak mencukupi untuk withdraw.\n"
        f"Minimum Withdraw: RM{min_wd:.2f}\n"
        f"Baki Semasa: RM{bal:.2f}"
    )

# ---------------------------
# CLONE BOT DATA
# ---------------------------
_clone_in_progress: dict = {}  # bot_id -> timestamp, auto-expires after 180s
def _reupload_file_id(src_token: str, dst_token: str, dst_chat_id: int, file_id: str, media_type: str) -> str:
    """
    Download file from source bot and re-upload to target bot.
    Returns new file_id valid for target bot, or original file_id on failure.
    Handles Telegram 429 rate-limit with retry.
    """
    import time as _time
    if not file_id or not media_type:
        return file_id
    try:
        # 1) getFile from source bot
        r = SESSION.post(
            f"https://api.telegram.org/bot{src_token}/getFile",
            data={"file_id": file_id}, timeout=15,
        )
        js = r.json()
        if not js.get("ok"):
            logger.warning("clone getFile failed: %s", js.get("description"))
            return file_id
        file_path = js["result"]["file_path"]

        # 2) Download file bytes
        dl = SESSION.get(
            f"https://api.telegram.org/file/bot{src_token}/{file_path}",
            timeout=30,
        )
        if dl.status_code != 200:
            return file_id
        file_bytes = dl.content
        fname = file_path.split("/")[-1] or "file"

        # 3) Re-upload to target bot via owner's chat (with retry on 429)
        method_map = {"photo": "sendPhoto", "video": "sendVideo",
                      "animation": "sendAnimation", "document": "sendDocument"}
        field_map = {"photo": "photo", "video": "video",
                     "animation": "animation", "document": "document"}
        method = method_map.get(media_type, "sendDocument")
        field = field_map.get(media_type, "document")

        up_js = None
        for attempt in range(4):  # max 4 attempts
            fdata = BytesIO(file_bytes)
            fdata.name = fname
            up = SESSION.post(
                f"https://api.telegram.org/bot{dst_token}/{method}",
                data={"chat_id": dst_chat_id, "disable_notification": True},
                files={field: (fdata.name, fdata)},
                timeout=60,
            )
            up_js = up.json()
            if up_js.get("ok"):
                break
            # Handle 429 Too Many Requests
            desc = (up_js.get("description") or "").lower()
            retry_after = up_js.get("parameters", {}).get("retry_after")
            if up_js.get("error_code") == 429 or "too many requests" in desc:
                wait = int(retry_after or 5) + 1
                logger.info("clone re-upload rate-limited, waiting %ds (attempt %d)", wait, attempt + 1)
                _time.sleep(wait)
                continue
            else:
                logger.warning("clone re-upload failed: %s", up_js.get("description"))
                return file_id

        if not up_js or not up_js.get("ok"):
            return file_id

        # 4) Extract new file_id + delete temp message
        result = up_js["result"]
        try:
            SESSION.post(
                f"https://api.telegram.org/bot{dst_token}/deleteMessage",
                data={"chat_id": dst_chat_id, "message_id": result["message_id"]},
                timeout=5,
            )
        except Exception:
            pass

        if media_type == "photo":
            photos = result.get("photo") or []
            return photos[-1]["file_id"] if photos else file_id
        elif media_type == "video":
            return (result.get("video") or {}).get("file_id") or file_id
        elif media_type == "animation":
            return (result.get("animation") or {}).get("file_id") or file_id
        else:
            return (result.get("document") or {}).get("file_id") or file_id
    except Exception as e:
        logger.error("clone _reupload_file_id error: %s", e)
        return file_id


def clone_bot_data(source_bot_id: str, target_bot_id: str, chat_id: int = 0) -> dict:
    """
    Clone setup/content from source bot to target bot.
    Re-uploads all media so file_ids are valid for the target bot.
    Copies: bot config, scanner_media, scanner_games, actions.
    Does NOT copy: users, referral data, withdrawals, scan usage.
    Returns summary dict with counts.
    """
    # Columns to clone from bots table (config/content only, not identity)
    CLONE_COLS = [
        "start_text", "start_media_type", "start_media_file_id",
        "loading_text", "loading_media_type", "loading_media_file_id",
        "join_message", "join_message_media_type", "join_message_media_file_id",
        "joined_message", "joined_message_media_type", "joined_message_media_file_id",
        "contact_message", "pending_message", "verified_message", "rejected_message",
        "group_contact_message", "withdrawal_prompt",
        "manual_approval", "inplace_callbacks",
        "affiliate_amount", "min_withdraw_amount",
        "scan_limit_per_day", "scan_limit_message",
        "scan_limit_message_media_type", "scan_limit_message_media_file_id",
        "scanner_duration_seconds",
        "withdrawal_approve_message", "withdrawal_approve_media_type", "withdrawal_approve_media_file_id",
        "withdrawal_reject_message", "withdrawal_reject_media_type", "withdrawal_reject_media_file_id",
    ]
    # Pairs of (media_type_col, file_id_col) that need re-upload
    MEDIA_PAIRS = [
        ("start_media_type", "start_media_file_id"),
        ("loading_media_type", "loading_media_file_id"),
        ("join_message_media_type", "join_message_media_file_id"),
        ("joined_message_media_type", "joined_message_media_file_id"),
        ("scan_limit_message_media_type", "scan_limit_message_media_file_id"),
        ("withdrawal_approve_media_type", "withdrawal_approve_media_file_id"),
        ("withdrawal_reject_media_type", "withdrawal_reject_media_file_id"),
    ]

    with engine.begin() as conn:
        # 1) Fetch source & target bot rows
        src = conn.execute(text("SELECT * FROM bots WHERE id=:i"), {"i": source_bot_id}).mappings().first()
        if not src:
            return {"error": "Source bot not found"}
        tgt = conn.execute(text("SELECT * FROM bots WHERE id=:i"), {"i": target_bot_id}).mappings().first()
        if not tgt:
            return {"error": "Target bot not found"}

        src_token = src["token"]
        dst_token = tgt["token"]
        reup_count = 0
        logger.info("CLONE: source=%s (%s) -> target=%s (%s)",
                     source_bot_id, src.get("bot_username"), target_bot_id, tgt.get("bot_username"))
        # Debug: dump source bot key fields
        _dbg_fields = ["start_text", "start_message", "start_media_type", "start_media_file_id",
                        "loading_text", "loading_media_type", "loading_media_file_id"]
        for _f in _dbg_fields:
            _v = src.get(_f)
            logger.info("CLONE SRC %s = %s", _f, repr(_v)[:120] if _v else "NULL")
        # Log all column names from source to spot unexpected fields
        logger.info("CLONE SRC columns: %s", list(src.keys()))
        # Check for duplicate bot rows with same username
        _src_uname = (src.get("bot_username") or "").lower()
        if _src_uname:
            _dup_rows = conn.execute(
                text("SELECT id, start_text, start_media_type FROM bots WHERE LOWER(bot_username)=:u"),
                {"u": _src_uname},
            ).mappings().all()
            logger.info("CLONE: found %d rows for username '%s'", len(_dup_rows), _src_uname)
            for _dr in _dup_rows:
                logger.info("  row id=%s start_text=%s media=%s",
                            _dr["id"], repr(_dr.get("start_text"))[:80] if _dr.get("start_text") else "NULL",
                            _dr.get("start_media_type") or "NULL")
            # If our source has no start_text but another row does, use that
            if not src.get("start_text"):
                for _dr in _dup_rows:
                    if _dr.get("start_text") and str(_dr["id"]) != str(source_bot_id):
                        _better_id = str(_dr["id"])
                        logger.info("CLONE: switching source to id=%s (has start_text!)", _better_id)
                        src = conn.execute(text("SELECT * FROM bots WHERE id=:i"), {"i": _better_id}).mappings().first()
                        source_bot_id = _better_id
                        src_token = src["token"]
                        break

        # Re-upload bot-level media file_ids
        params = {f"_{c}": src.get(c) for c in CLONE_COLS}
        # Fallback: editstart bug saved to start_message instead of start_text
        if not params.get("_start_text") and src.get("start_message"):
            params["_start_text"] = src.get("start_message")
            logger.info("CLONE: using start_message fallback (start_text was NULL)")
        for mt_col, fid_col in MEDIA_PAIRS:
            mt_val = src.get(mt_col)
            fid_val = src.get(fid_col)
            if mt_val and fid_val:
                new_fid = _reupload_file_id(src_token, dst_token, chat_id, fid_val, mt_val)
                params[f"_{fid_col}"] = new_fid
                reup_count += 1

        sets = ", ".join(f"{c}=:_{c}" for c in CLONE_COLS if src.get(c) is not None or c in ("manual_approval", "inplace_callbacks"))
        params["_tid"] = target_bot_id
        if sets:
            conn.execute(text(f"UPDATE bots SET {sets} WHERE id=:_tid"), params)

        # 2) Clone scanner_media (re-upload each)
        conn.execute(text("DELETE FROM scanner_media WHERE bot_id=:b"), {"b": target_bot_id})
        media_rows = conn.execute(
            text("SELECT provider, media_type, file_id FROM scanner_media WHERE bot_id=:b"),
            {"b": source_bot_id},
        ).mappings().all()
        logger.info("CLONE: scanner_media rows found: %d (source=%s)", len(media_rows), source_bot_id)
        for m in media_rows:
            new_fid = _reupload_file_id(src_token, dst_token, chat_id, m["file_id"], m["media_type"])
            conn.execute(
                text("INSERT INTO scanner_media (bot_id, provider, media_type, file_id, updated_at) VALUES (:b, :p, :mt, :fid, NOW())"),
                {"b": target_bot_id, "p": m["provider"], "mt": m["media_type"], "fid": new_fid},
            )
            reup_count += 1

        # 3) Clone scanner_games
        conn.execute(text("DELETE FROM scanner_games WHERE bot_id=:b"), {"b": target_bot_id})
        game_rows = conn.execute(
            text("SELECT provider, game FROM scanner_games WHERE bot_id=:b"),
            {"b": source_bot_id},
        ).mappings().all()
        for g in game_rows:
            conn.execute(
                text("INSERT INTO scanner_games (bot_id, provider, game) VALUES (:b, :p, :g)"),
                {"b": target_bot_id, "p": g["provider"], "g": g["game"]},
            )

        logger.info("CLONE: scanner_games rows found: %d", len(game_rows))

        # 4) Clone actions (re-upload media)
        conn.execute(text("DELETE FROM actions WHERE bot_id=:b"), {"b": target_bot_id})
        action_rows = conn.execute(
            text("SELECT key, type, text, media_file_id, delay_seconds FROM actions WHERE bot_id=:b"),
            {"b": source_bot_id},
        ).mappings().all()
        logger.info("CLONE: actions rows found: %d", len(action_rows))
        for a in action_rows:
            fid = a["media_file_id"]
            if fid and a["type"] != "text":
                fid = _reupload_file_id(src_token, dst_token, chat_id, fid, a["type"])
                reup_count += 1
            conn.execute(
                text("INSERT INTO actions (bot_id, key, type, text, media_file_id, delay_seconds) VALUES (:b, :k, :t, :tx, :mf, :d)"),
                {"b": target_bot_id, "k": a["key"], "t": a["type"], "tx": a["text"], "mf": fid, "d": a["delay_seconds"]},
            )

    return {
        "media": len(media_rows),
        "games": len(game_rows),
        "actions": len(action_rows),
        "reuploaded": reup_count,
        "source_id": str(source_bot_id)[:8],
        "source_user": src.get("bot_username") or "?",
        "start_text_len": len(src.get("start_text") or ""),
    }


# ---------------------------
# ADMIN MANAGEMENT
# ---------------------------
def list_admins(bot_id: str):
    with engine.connect() as conn:
        return conn.execute(
            text("""
                SELECT admin_user_id, expiry_at, added_by
                FROM admins
                WHERE bot_id=:b
                ORDER BY admin_user_id ASC
            """),
            {"b": bot_id},
        ).mappings().all()


def add_admin(bot_id: str, admin_user_id: int, added_by: int, days: Optional[int]):
    expiry_at = None
    if days is not None:
        days = min(int(days), 3650)  # cap to ~10 years to prevent OverflowError
        expiry_at = utcnow() + timedelta(days=days)
    with engine.begin() as conn:
        conn.execute(
            text("""
                INSERT INTO admins (bot_id, admin_user_id, expiry_at, added_by)
                VALUES (:b, :u, :e, :by)
                ON CONFLICT (bot_id, admin_user_id) DO UPDATE
                SET expiry_at=excluded.expiry_at, added_by=excluded.added_by
            """),
            {"b": bot_id, "u": int(admin_user_id), "e": expiry_at, "by": int(added_by)},
        )


def del_admin(bot_id: str, admin_user_id: int) -> bool:
    with engine.begin() as conn:
        res = conn.execute(
            text("DELETE FROM admins WHERE bot_id=:b AND admin_user_id=:u"),
            {"b": bot_id, "u": int(admin_user_id)},
        )
    return res.rowcount > 0


# ---------------------------
# CLOUD TASKS BROADCAST
# ---------------------------
def can_use_tasks() -> bool:
    return _HAS_CLOUD_TASKS and all([GCP_PROJECT, TASKS_LOCATION, TASKS_QUEUE, TASKS_HANDLER_URL, TASKS_SECRET])


def can_use_tasks_action() -> bool:
    # For delayed callback actions (/task/action). Uses PUBLIC_BASE_URL as handler URL.
    return _HAS_CLOUD_TASKS and all([GCP_PROJECT, TASKS_LOCATION, TASKS_QUEUE, TASKS_SECRET, PUBLIC_BASE_URL])


def enqueue_broadcast_task(payload: dict) -> None:
    client = tasks_v2.CloudTasksClient()
    parent = client.queue_path(GCP_PROJECT, TASKS_LOCATION, TASKS_QUEUE)
    task = {
        "http_request": {
            "http_method": tasks_v2.HttpMethod.POST,
            "url": TASKS_HANDLER_URL,
            "headers": {"Content-Type": "application/json", "X-Tasks-Secret": TASKS_SECRET},
            "body": json.dumps(payload).encode("utf-8"),
        }
    }
    client.create_task(request={"parent": parent, "task": task})




def enqueue_action_task(payload: dict, delay_seconds: int) -> None:
    """
    Cloud Tasks scheduled action for delayed callbacks.
    """
    client = tasks_v2.CloudTasksClient()
    parent = client.queue_path(GCP_PROJECT, TASKS_LOCATION, TASKS_QUEUE)

    # schedule_time uses UTC Timestamp
    from google.protobuf import timestamp_pb2
    run_at = utcnow() + timedelta(seconds=max(0, int(delay_seconds or 0)))
    ts = timestamp_pb2.Timestamp()
    ts.FromDatetime(run_at)

    task = {
        "schedule_time": ts,
        "http_request": {
            "http_method": tasks_v2.HttpMethod.POST,
            "url": f"{PUBLIC_BASE_URL}/task/action",
            "headers": {"Content-Type": "application/json", "X-Tasks-Secret": TASKS_SECRET},
            "body": json.dumps(payload).encode("utf-8"),
        }
    }
    client.create_task(request={"parent": parent, "task": task})
# ---------------------------
# SETTINGS UI HELPERS
# ---------------------------
def get_bot_stats(bot_id: str):
    with engine.connect() as conn:
        total = conn.execute(text("SELECT COUNT(*) c FROM users WHERE bot_id=:b"), {"b": bot_id}).mappings().first()["c"]
        verified = conn.execute(text("SELECT COUNT(*) c FROM users WHERE bot_id=:b AND is_verified=TRUE"), {"b": bot_id}).mappings().first()["c"]
        premium = conn.execute(text("SELECT COUNT(*) c FROM users WHERE bot_id=:b AND is_premium=TRUE"), {"b": bot_id}).mappings().first()["c"]
        pending_wd = conn.execute(text("SELECT COUNT(*) c FROM withdrawals WHERE bot_id=:b AND status='PENDING'"), {"b": bot_id}).mappings().first()["c"]
    return {"total_users": int(total), "verified_users": int(verified), "premium_users": int(premium), "pending_withdrawals": int(pending_wd)}


def get_callbacks_page(bot_id: str, page: int, page_size: int):
    if page < 1:
        page = 1
    off = (page - 1) * page_size
    with engine.connect() as conn:
        total = conn.execute(
            text("SELECT COUNT(*) c FROM actions WHERE bot_id=:b"),
            {"b": bot_id},
        ).mappings().first()["c"]
        rows = conn.execute(
            text("""
                SELECT key, type, delay_seconds
                FROM actions
                WHERE bot_id=:b
                ORDER BY key ASC
                LIMIT :lim OFFSET :off
            """),
            {"b": bot_id, "lim": page_size, "off": off},
        ).mappings().all()
    return int(total), rows


def delete_callback(bot_id: str, key: str) -> bool:
    with engine.begin() as conn:
        res = conn.execute(
            text("DELETE FROM actions WHERE bot_id=:b AND key=:k"),
            {"b": bot_id, "k": key},
        )
    return res.rowcount > 0


def settings_help_all() -> str:
    """Return comprehensive help text with all admin commands."""
    return (
        "📚 <b>ADMIN COMMANDS GUIDE</b>\n"
        "━━━━━━━━━━━━━━━━━━\n\n"
        
        "🎯 <b>CONTENT SETUP</b>\n"
        "• <code>/setstart</code> - Set START message (reply to content)\n"
        "• <code>/setloading</code> - Set LOADING message (reply to content)\n"
        "• <code>/setcallback key</code> - Set callback action\n"
        "• <code>/setcallback key delay=5</code> - With delay\n"
        "• <code>/setcommand name</code> - Set custom command\n"
        "• <code>/delcallback key</code> - Delete callback\n\n"
        
        "🔐 <b>SECURITY & ACCESS</b>\n"
        "• <code>/setlockbot on|off</code> - Lock/unlock bot\n"
        "• <code>/setjoin @ch1,@ch2</code> - Set join targets\n"
        "• <code>/setjoinmsg</code> - Custom join lock message\n"
        "• <code>/setjoinedmsg</code> - Message after user passes JoinLock\n"
        "• <code>/setadmingroup</code> - Set admin group (in group)\n"
        "• JoinLock, PhoneLock, Manual → via /settings buttons\n\n"
        
        "👥 <b>ADMIN MANAGEMENT</b>\n"
        "• <code>/addadmin user_id</code> - Add admin (30 days default)\n"
        "• <code>/addadmin user_id 7</code> - Add admin (7 days)\n"
        "• <code>/deladmin user_id</code> - Remove admin\n"
        "• <code>/admins</code> - List all admins\n"
        "• <code>/approve user_id</code> - Manually approve user\n"
        "• <code>/myid</code> - Get your user ID\n\n"
        
        "💰 <b>FINANCIAL</b>\n"
        "• <code>/setshareamt 1.00</code> - Set share commission (RM)\n"
        "• <code>/setminwithdraw 30.00</code> - Set min withdraw (RM)\n"
        "• <code>/getrates</code> - View current rates\n\n"
        
        "✉️ <b>MESSAGES</b>\n"
        "• <code>/setwithdrawalmsg</code> - Withdrawal prompt message\n"
        "• <code>/setwithdrawalreject</code> - Rejection message\n"
        "• <code>/setcontactmsg</code> - Contact request message\n"
        "• <code>/setpendingmsg</code> - Pending approval message\n"
        "• <code>/setverifiedmsg</code> - Verified success message\n"
        "• <code>/setrejectedmsg</code> - Rejected message\n"
        "• <code>/setgroupcontactmsg</code> - Group contact message\n\n"
        
        "🎰 <b>SCANNER</b>\n"
        "• <code>/addscanner provider</code> - Add scanner media (reply)\n"
        "• <code>/addgames provider</code> - Add games (reply JSON)\n"
        "• <code>/updategames provider</code> - Update existing games\n"
        "• <code>/setscanlimit 20</code> - Set global daily limit\n"
        "• <code>/setscanlimit off</code> - Remove daily limit\n"
        "• <code>/setscanlimit 5 @user</code> - User override\n"
        "• <code>/setscanlimit reset</code> - Reset today's usage\n"
        "• <code>/setscanlimitmsg</code> - Custom limit message (reply)\n"
        "• <code>/clearscan</code> - Clear scan data\n\n"
        
        "📣 <b>BROADCAST</b>\n"
        "• <code>/broadcast</code> - Send to ALL users (reply)\n"
        "• <code>/broadcast verified</code> - Send to verified only\n\n"
        
        "➕ <b>MULTI-BOT</b>\n"
        "• <code>/addbot</code> - Add new bot (paste token)\n"
        "• <code>/mybots</code> - List your bots\n"
        "• <code>/clone @botusername</code> - Clone data from another bot\n\n"
        
        "📊 <b>DATA</b>\n"
        "• Export ALL / Export VERIFIED → via /settings buttons\n\n"
        
        "🧩 <b>BUTTON SYNTAX</b>\n"
        "• <code>!1link Name|https://url.com</code>\n"
        "• <code>!1callback Name|key</code>\n"
        "• <code>!1callback Name|key delay=5</code>\n"
        "• <code>!1share Button Text</code>\n"
        "• <code>!1withdrawal Withdraw</code>\n"
        "Row numbers: !1 !2 !3 etc\n\n"
        
        "📝 <b>PLACEHOLDERS</b>\n"
        "• {firstname} {username} {member_id} {date}\n"
        "• [balance] [share] [link]\n"
        "• {rand:1-100} {count} {limit} {remaining}\n"
        "• [web](https://url.com) - clickable link\n"
        "• **bold** __italic__ - text formatting\n"
    )

def settings_how(topic: str) -> str:
    if topic == "setstart":
        return (
            "📌 <b>Cara set START</b>\n"
            "1) Reply mesej (text/gambar/video) yang kau nak jadi START\n"
            "2) Tulis:\n"
            "<code>/setstart</code>\n\n"
            "Tambah button (optional):\n"
            "<code>!1link Join|https://example.com</code>\n"
            "<code>!2callback Claim|bonus</code>\n"
            "<code>!3withdrawal Withdraw</code>\n"
            "Link dalam text pun boleh guna: <code>[web](https://example.com)</code>\n"
        )
    if topic == "setloading":
        return (
            "⏳ <b>Cara set LOADING</b>\n"
            "1) Reply content (text/gambar/video)\n"
            "2) Tulis:\n"
            "<code>/setloading</code>\n"
        )
    if topic == "callback":
        return (
            "🧩 <b>Cara buat CALLBACK</b>\n"
            "1) Dalam text START/LOADING letak button callback:\n"
            "<code>!1callback Claim Bonus|bonus</code>\n"
            "2) Lepas tu reply satu mesej (text/gambar/video) dan tulis:\n"
            "<code>/setcallback bonus</code>\n"
            "Optional delay:\n"
            "<code>/setcallback bonus delay=5</code>\n"
        )
    if topic == "setcommand":
        return (
            "🧷 <b>Cara buat SETCOMMAND</b>\n"
            "1) Reply content (text/gambar/video)\n"
            "2) Tulis:\n"
            "<code>/setcommand hello</code>\n"
            "User akan trigger dengan:\n"
            "<code>/hello</code>\n"
            "Optional delay:\n"
            "<code>/setcommand hello delay=3</code>\n"
        )

        return (
            "🎬 <b>Cara set MEDIA SCANNER (ikut provider)</b>\n"
            "━━━━━━━━━━━━━━━━━━\n"
            "1) Send GIF / video / gambar ke bot\n"
            "2) Reply media tu, kemudian taip:\n"
            "<code>/<disabled> jili</code>\n"
            "atau\n"
            "<code>/<disabled> mega888</code>\n\n"
            "✅ Bila user tekan button scanner (callback key: <code>scanner jili</code>), bot akan auto keluarkan media tu dulu.\n\n"
            "Buang media provider:\n"
            "<code>/<disabled> jili</code>\n"
        )
    if topic == "broadcast":
        return (
            "📣 <b>Broadcast</b>\n"
            "1) Reply content yang nak dihantar\n"
            "2) Tulis:\n"
            "<code>/broadcast</code>\n"
            "atau verified sahaja:\n"
            "<code>/broadcast verified</code>\n"
            "Buttons pun support (letak line !1link / !1callback etc dalam command text bawah).\n"
        )
    if topic == "addbot":
        return (
            "➕ <b>Multi-bot: /addbot</b>\n"
            "1) PM bot utama (yang ada /settings)\n"
            "2) Taip <code>/addbot</code>\n"
            "3) Paste token bot baru dari BotFather\n"
            "\n"
            "✅ Kalau env <code>PUBLIC_BASE_URL</code> dah set, dia auto setWebhook.\n"
        )
    if topic == "joinlock":
        return (
            "🧲 <b>Join Lock (Wajib join)</b>\n"
            "• ON/OFF: guna button di /settings\n"
            "• Set target:\n"
            "<code>/setjoin @channel1,@channel2</code>\n"
            "atau newline:\n"
            "<code>/setjoin</code> (reply text list)\n\n"
            "Custom ayat join:\n"
            "<code>/setjoinmsg</code> (reply text)\n"
        )
    if topic == "manualapprove":
        return (
            "✅ <b>Manual Approve Premium</b>\n"
            "Bila ON:\n"
            "1) User share contact\n"
            "2) Bot post ke Admin Group (ada button Approve/Reject)\n"
            "3) Admin approve → user jadi Premium & dapat mesej custom\n\n"
            "Custom ayat:\n"
            "• <code>/setpendingmsg</code> (reply text)\n"
            "• <code>/setverifiedmsg</code> (reply text)\n"
            "• <code>/setrejectedmsg</code> (reply text)\n"
            "• <code>/setgroupcontactmsg</code> (reply text)\n"
        )
    if topic == "addscanner":
        return (
            "🎰 <b>How to Add Scanner Provider</b>\n"
            "━━━━━━━━━━━━━━━━━━\n\n"
            "Scanner functionality requires TWO things:\n"
            "1️⃣ <b>Scanner Media</b> (intro GIF/video)\n"
            "2️⃣ <b>Games Database</b> (game list with RTP)\n\n"
            "📌 <b>Step 1: Add Scanner Media</b>\n"
            "• Send GIF/video/photo to bot\n"
            "• Reply to that media:\n"
            "<code>/setscannermedia jili</code> atau <code>/addscanner jili</code>\n"
            "<code>/setscannermedia mega888</code> atau <code>/addscanner mega888</code>\n\n"
            "📌 <b>Step 2: Add Games</b>\n"
            "• See \"📖 How: Add/Update Games\" button\n\n"
            "✅ <b>When Ready</b>\n"
            "Create callback button:\n"
            "<code>!1callback Mega888 Scanner|mega888</code>\n"
            "User tekan → bot auto keluarkan media + run scanner!\n\n"
            "💡 <b>Note:</b> Provider key must match (e.g., 'jili', 'mega888')\n"
        )
    if topic == "addgames":
        return (
            "🎮 <b>How to Add/Update Games</b>\n"
            "━━━━━━━━━━━━━━━━━━\n\n"
            "📌 <b>Using Admin Panel (Web)</b>\n"
            "1) Login ke admin panel\n"
            "2) Go to Games Management\n"
            "3) Upload CSV or use JSON import\n\n"
            "📌 <b>Using Bot Commands</b>\n"
            "Format JSON list, reply ke bot:\n"
            "<code>/addgames jili</code>\n\n"
            "<b>JSON Example:</b>\n"
            "<code>[\n"
            '  {"name": "Game 1", "rtp": 96.5},\n'
            '  {"name": "Game 2", "rtp": 97.2}\n'
            "]</code>\n\n"
            "📊 <b>CSV Format (Admin Panel):</b>\n"
            "<code>name,rtp\n"
            "Fortune Gems,96.52\n"
            "Lucky Treasure,97.18</code>\n\n"
            "✅ Games auto-linked to provider key\n"
        )
    if topic == "setshareamt":
        return (
            "💰 <b>Set RM per Share</b>\n"
            "1) Reply apa-apa mesej (optional)\n"
            "2) Tulis:\n"
            "<code>/setshareamt 1</code>  (contoh RM1 per share)\n"
            "Boleh decimal:\n"
            "<code>/setshareamt 0.50</code>\n"
        )
    if topic == "setminwithdraw":
        return (
            "🏧 <b>Set Minimum Withdraw</b>\n"
            "Tulis:\n"
            "<code>/setminwithdraw 50</code>\n"
            "Boleh decimal:\n"
            "<code>/setminwithdraw 10.50</code>\n"
        )
    if topic == "setwithdrawalmsg":
        return (
            "📝 <b>Set mesej bila user tekan Withdraw (cukup balance)</b>\n"
            "Reply content (text/gambar/video) yang kau nak, kemudian tulis:\n"
            "<code>/setwithdrawalmsg</code>\n"
            "Tips: support placeholder macam {firstname}, [balance], [link].\n"
        )
    if topic == "setwithdrawalreject":
        return (
            "📝 <b>Set mesej bila Withdraw ditolak (reject)</b>\n"
            "Reply content, kemudian tulis:\n"
            "<code>/setwithdrawalreject</code>\n"
        )
    if topic == "setpendingmsg":
        return (
            "📝 <b>Set mesej bila Withdraw masuk status PENDING</b>\n"
            "Reply content, kemudian tulis:\n"
            "<code>/setpendingmsg</code>\n"
        )
    if topic == "setrejectedmsg":
        return (
            "📝 <b>Set mesej bila Admin tekan Reject</b>\n"
            "Reply content, kemudian tulis:\n"
            "<code>/setrejectedmsg</code>\n"
        )
    if topic == "setverifiedmsg":
        return (
            "📝 <b>Set mesej bila user VERIFIED (lepas share contact)</b>\n"
            "Reply content, kemudian tulis:\n"
            "<code>/setverifiedmsg</code>\n"
        )
    if topic == "setcontactmsg":
        return (
            "📝 <b>Set mesej minta user share contact</b>\n"
            "Reply content, kemudian tulis:\n"
            "<code>/setcontactmsg</code>\n"
        )
    if topic == "setjoinedmsg":
        return (
            "📝 <b>Set mesej lepas user dah join channel/group</b>\n"
            "Reply content (text/gambar/video) yang kau nak, kemudian tulis:\n"
            "<code>/setjoinedmsg</code>\n\n"
            "Tip: support placeholder macam {firstname}, {username}, {member_id}, [balance], [share], [link] dan button syntax."
        )
    if topic == "setgroupcontactmsg":
        return (
            "📝 <b>Set mesej minta user share contact (dalam group)</b>\n"
            "Reply content, kemudian tulis:\n"
            "<code>/setgroupcontactmsg</code>\n"
        )
    if topic == "setlockbot":
        return (
            "🔐 <b>Lock Bot</b>\n"
            "ON (hanya admin boleh guna):\n"
            "<code>/setlockbot on</code>\n"
            "OFF:\n"
            "<code>/setlockbot off</code>\n"
        )



    if topic == "setscanlimit":
        return (
            "🧮 <b>Set Scan Limit Harian</b>\\n"
            "Global limit (semua user):\\n"
            "<code>/setscanlimit 20</code>\\n"
            "OFF (unlimited):\\n"
            "<code>/setscanlimit off</code>\\n\\n"
            "Override ikut user (username / user_id):\\n"
            "<code>/setscanlimit 5 @username</code>\\n"
            "<code>/setscanlimit 5 123456789</code>\\n\\n"
            "Reset usage hari ini:\\n"
            "<code>/setscanlimit reset</code>\\n"
            "<code>/setscanlimit reset @username</code>\\n\\n"
            "Buang override user:\\n"
            "<code>/setscanlimit del @username</code>\\n"
        )

    if topic == "setscanlimitmsg":
        return (
            "📝 <b>Set Mesej Bila User Capai Limit Scan</b>\\n"
            "Reply content (text/gambar/video) yang kau nak, kemudian tulis:\\n"
            "<code>/setscanlimitmsg</code>\\n\\n"
            "Tip: boleh letak placeholder biasa macam {firstname}, [balance], [link]."
        )

    return "OK"

def build_settings_text(bot_row: dict, stats: dict, cb_total: int, cb_rows: list, page: int, page_size: int):
    bot_id = str(bot_row["id"])
    username = bot_row.get("bot_username") or "-"
    lock_phone = "🟢 ON" if bot_row.get("lock_bot") else "🔴 OFF"
    lock_join = "🟢 ON" if bot_row.get("join_lock") else "🔴 OFF"
    manual = "🟢 ON" if bot_row.get("manual_approval") else "🔴 OFF"
    inplace = "🟢 ON" if bot_row.get("inplace_callbacks") else "🔴 OFF"

    adming = bot_row.get("admin_group_id")
    adming_txt = f"<code>{adming}</code>" if adming else "<i>Not set</i>"
    base_url_txt = PUBLIC_BASE_URL if PUBLIC_BASE_URL else "<i>Not set</i>"

    # Calculate verified percentage
    verified_pct = 0
    if stats['total_users'] > 0:
        verified_pct = int((stats['verified_users'] / stats['total_users']) * 100)

    pages = max(1, (cb_total + page_size - 1) // page_size)

    # Financial settings
    share_amt = bot_row.get("affiliate_amount") or AFFILIATE_AMOUNT
    min_wd = bot_row.get("min_withdraw_amount") or MIN_WITHDRAW_DEFAULT
    
    # Scan configuration
    scan_limit = bot_row.get("scan_limit_per_day")
    scan_status = "♾️ UNLIMITED" if not scan_limit else f"🔢 {int(scan_limit)}/day"
    scan_duration = get_scanner_duration_seconds(bot_row)
    scan_loading_count = scanner_loading_text_count(scan_duration)
    joined_msg_status = "✅ CUSTOM" if (bot_row.get("joined_message") or bot_row.get("joined_message_media_type")) else "🟡 DEFAULT"

    txt = (
        "⚙️ <b>ADMIN CONTROL PANEL</b>\n"
        "━━━━━━━━━━━━━━━━━━\n"
        f"🤖 <b>@{username}</b>\n"
        "\n"
        "📊 <b>QUICK STATS</b>\n"
        f"👥 Users: <b>{stats['total_users']}</b> | "
        f"✅ Verified: <b>{stats['verified_users']}</b> ({verified_pct}%)\n"
        f"💎 Premium: <b>{stats['premium_users']}</b> | "
        f"💰 Pending WD: <b>{stats['pending_withdrawals']}</b>\n"
        "\n"
        "━━━━━━━━━━━━━━━━━━\n"
        "⚡ <b>CURRENT STATUS</b>\n"
        f"🧲 JoinLock: {lock_join} | 🔒 PhoneLock: {lock_phone}\n"
        f"✅ Manual: {manual} | 🧩 Inplace: {inplace}\n"
        f"💵 Share: <b>RM{share_amt:.2f}</b> | 🏧 Min WD: <b>RM{min_wd:.2f}</b>\n"
        f"🎰 Scan Limit: {scan_status}\n"
        f"⏱️ Scan Loading: <b>{scanner_duration_label(scan_duration)}</b> ({scan_loading_count} text × 5s)\n"
        f"🎉 Joined Msg: <b>{joined_msg_status}</b>\n"
        f"🧩 Callbacks: <b>{cb_total}</b>\n"
        "\n"
        "━━━━━━━━━━━━━━━━━━\n"
        "📁 <i>Pilih kategori di bawah untuk manage</i>\n"
    )
    
    return txt, pages



def build_settings_keyboard_full(page: int, pages: int):
    """Improved settings keyboard with grouped categories."""
    prev_page = page - 1 if page > 1 else 1
    next_page = page + 1 if page < pages else pages

    kb = {
        "inline_keyboard": [
            # SECURITY & ACCESS TOGGLES
            [
                {"text": "🔐 SECURITY & ACCESS", "callback_data": "st:noop"},
            ],
            [
                {"text": "🧲 JoinLock ON", "callback_data": "st:join:on"},
                {"text": "🧲 JoinLock OFF", "callback_data": "st:join:off"},
            ],
            [
                {"text": "🔒 PhoneLock ON", "callback_data": "st:lock:on"},
                {"text": "🔓 PhoneLock OFF", "callback_data": "st:lock:off"},
            ],
            [
                {"text": "✅ Manual Approve ON", "callback_data": "st:manual:on"},
                {"text": "❌ Manual Approve OFF", "callback_data": "st:manual:off"},
            ],
            [
                {"text": "👥 Set Admin Group (in GROUP)", "callback_data": "st:admingroup:set"},
            ],
            
            # FINANCIAL SETTINGS
            [
                {"text": "💰 FINANCIAL SETTINGS", "callback_data": "st:noop"},
            ],
            [
                {"text": "💵 View Rates", "callback_data": "st:financial:viewrates"},
            ],
            [
                {"text": "✏️ Edit Share Amount", "callback_data": "st:financial:setshare"},
                {"text": "✏️ Edit Min Withdraw", "callback_data": "st:financial:setminwd"},
            ],
            
            # CONTENT & MESSAGES
            [
                {"text": "📝 CONTENT MANAGEMENT", "callback_data": "st:noop"},
            ],
            [
                {"text": "✏️ Edit START Message", "callback_data": "st:content:editstart"},
                {"text": "✏️ Edit LOADING Message", "callback_data": "st:content:editloading"},
            ],
            [
                {"text": "📌 Preview START", "callback_data": "st:preview:start"},
                {"text": "⏳ Preview LOADING", "callback_data": "st:preview:loading"},
            ],
            [
                {"text": "📌 How: START", "callback_data": "st:how:setstart"},
                {"text": "⏳ How: LOADING", "callback_data": "st:how:setloading"},
            ],
            [
                {"text": "🧩 How: CALLBACK", "callback_data": "st:how:callback"},
                {"text": "🧷 How: COMMAND", "callback_data": "st:how:setcommand"},
            ],
            
            # FEATURES
            [
                {"text": "⚙️ BOT FEATURES", "callback_data": "st:noop"},
            ],
            [
                {"text": "🧩 Inplace ON", "callback_data": "st:inplace:on"},
                {"text": "🧩 Inplace OFF", "callback_data": "st:inplace:off"},
            ],
            [
                {"text": "🧲 How: JoinLock", "callback_data": "st:how:joinlock"},
                {"text": "✅ How: Manual Approve", "callback_data": "st:how:manualapprove"},
            ],
            [
                {"text": "📣 How: Broadcast", "callback_data": "st:how:broadcast"},
                {"text": "➕ How: AddBot", "callback_data": "st:how:addbot"},
            ],
            
            # SCAN SETTINGS
            [
                {"text": "🎰 SCAN SETTINGS", "callback_data": "st:noop"},
            ],
            [
                {"text": "⚙️ Set Global Limit", "callback_data": "st:scan:setglobal"},
            ],
            [
                {"text": "📊 View Usage", "callback_data": "st:scan:viewusage"},
                {"text": "🔄 Reset Usage", "callback_data": "st:scan:reset"},
            ],
            [
                {"text": "✏️ Custom Limit Message", "callback_data": "st:scan:editmsg"},
            ],
            
            # WITHDRAWAL MESSAGES
            [
                {"text": "✉️ WITHDRAWAL MESSAGES", "callback_data": "st:noop"},
            ],
            [
                {"text": "✏️ Edit Request Message", "callback_data": "st:withdrawal:editrequest"},
            ],
            [
                {"text": "✏️ Edit Approve Message", "callback_data": "st:withdrawal:editapprove"},
                {"text": "✏️ Edit Reject Message", "callback_data": "st:withdrawal:editreject"},
            ],
            
            # VERIFICATION MESSAGES
            [
                {"text": "📢 VERIFICATION MESSAGES", "callback_data": "st:noop"},
            ],
            [
                {"text": "✏️ JoinLock Message", "callback_data": "st:verify:editjoin"},
                {"text": "✏️ Contact Request", "callback_data": "st:verify:editcontact"},
            ],
            [
                {"text": "✏️ Pending Message", "callback_data": "st:verify:editpending"},
                {"text": "✏️ Verified Message", "callback_data": "st:verify:editverified"},
            ],
            [
                {"text": "✏️ Rejected Message", "callback_data": "st:verify:editrejected"},
                {"text": "✏️ Group Contact Msg", "callback_data": "st:verify:editgroupcontact"},
            ],
            
            # SCANNER MANAGEMENT
            [
                {"text": "🎮 SCANNER MANAGEMENT", "callback_data": "st:noop"},
            ],
            [
                {"text": "📖 How: Add Scanner", "callback_data": "st:how:addscanner"},
                {"text": "📖 How: Add/Update Games", "callback_data": "st:how:addgames"},
            ],
            
            # DATA MANAGEMENT
            [
                {"text": "📊 DATA & EXPORT", "callback_data": "st:noop"},
            ],
            [
                {"text": "📤 Export ALL Users", "callback_data": "st:export:all"},
                {"text": "✅ Export VERIFIED", "callback_data": "st:export:verified"},
            ],
            [
                {"text": "📃 My Bots", "callback_data": "st:mybots:0"},
                {"text": "🗑 Delete Callback", "callback_data": "st:cbdelmenu:1"},
            ],
            
            # UTILITIES
            [
                {"text": "🔧 UTILITIES", "callback_data": "st:noop"},
            ],
            [
                {"text": "📚 All Commands Help", "callback_data": "st:help:all"},
            ],
            [
                {"text": "🧠 Full Placeholders", "callback_data": "st:placeholders:full"},
                {"text": "🔄 Refresh Panel", "callback_data": f"st:refresh:{page}"},
            ],
            
            # PAGINATION
            [
                {"text": "⬅️ Prev", "callback_data": f"st:cbpage:{prev_page}"},
                {"text": f"📄 {page}/{pages}", "callback_data": "st:noop"},
                {"text": "Next ➡️", "callback_data": f"st:cbpage:{next_page}"},
            ],
        ]
    }
    return kb


def build_settings_category_nav(active: str = "home"):
    def lab(key, label):
        return f"👉 {label}" if key == active else label

    return {
        "inline_keyboard": [
            [
                {"text": lab("content", "📝 Content"), "callback_data": "st:cat:content"},
                {"text": lab("security", "🔐 Security"), "callback_data": "st:cat:security"},
            ],
            [
                {"text": lab("economy", "💰 Financial"), "callback_data": "st:cat:economy"},
                {"text": lab("scanner", "🎰 Scanner"), "callback_data": "st:cat:scanner"},
            ],
            [
                {"text": lab("withdraw", "✉️ Withdrawal"), "callback_data": "st:cat:withdraw"},
                {"text": lab("verify", "📢 Verification"), "callback_data": "st:cat:verify"},
            ],
            [
                {"text": lab("callback", "🧩 Callbacks"), "callback_data": "st:cat:callback"},
                {"text": lab("data", "📊 Data & Export"), "callback_data": "st:cat:data"},
            ],
            [
                {"text": "🔧 Utilities", "callback_data": "st:cat:utils"},
                {"text": "🔄 Refresh", "callback_data": "st:refresh:1"},
            ],
        ]
    }


def build_settings_keyboard_by_category(bot_row: dict, cat: str, page: int, pages: int):
    """Category sub-menu keyboard with focused actions and Back button."""
    cat = (cat or "").lower().strip()
    kb = {"inline_keyboard": []}

    if cat == "content":
        kb["inline_keyboard"].extend([
            [{"text": "📝 CONTENT MANAGEMENT", "callback_data": "st:noop"}],
            [{"text": "✏️ Edit START Message", "callback_data": "st:content:editstart"},
             {"text": "✏️ Edit LOADING Message", "callback_data": "st:content:editloading"}],
            [{"text": "📌 Preview START", "callback_data": "st:preview:start"},
             {"text": "⏳ Preview LOADING", "callback_data": "st:preview:loading"}],
            [{"text": "📖 How: /setstart", "callback_data": "st:how:setstart"},
             {"text": "📖 How: /setloading", "callback_data": "st:how:setloading"}],
        ])

    elif cat == "security":
        kb["inline_keyboard"].extend([
            [{"text": "🔐 SECURITY & ACCESS", "callback_data": "st:noop"}],
            [{"text": "🧲 JoinLock ON", "callback_data": "st:join:on"},
             {"text": "🧲 JoinLock OFF", "callback_data": "st:join:off"}],
            [{"text": "🔒 PhoneLock ON", "callback_data": "st:lock:on"},
             {"text": "🔓 PhoneLock OFF", "callback_data": "st:lock:off"}],
            [{"text": "✅ Manual Approve ON", "callback_data": "st:manual:on"},
             {"text": "❌ Manual Approve OFF", "callback_data": "st:manual:off"}],
            [{"text": "👥 Set Admin Group", "callback_data": "st:admingroup:set"}],
            [{"text": "📖 How: JoinLock", "callback_data": "st:how:joinlock"},
             {"text": "📖 How: Manual", "callback_data": "st:how:manualapprove"}],
        ])

    elif cat == "economy":
        kb["inline_keyboard"].extend([
            [{"text": "💰 FINANCIAL SETTINGS", "callback_data": "st:noop"}],
            [{"text": "💵 View Rates", "callback_data": "st:financial:viewrates"}],
            [{"text": "✏️ Edit Share Amount", "callback_data": "st:financial:setshare"},
             {"text": "✏️ Edit Min Withdraw", "callback_data": "st:financial:setminwd"}],
        ])

    elif cat == "scanner":
        cur_scan_duration = get_scanner_duration_seconds(bot_row)
        cur_scan_loading_count = scanner_loading_text_count(cur_scan_duration)
        joined_msg_set = bool(bot_row.get("joined_message") or bot_row.get("joined_message_media_type"))
        kb["inline_keyboard"].extend([
            [{"text": "🎰 SCANNER MANAGEMENT", "callback_data": "st:noop"}],
            [{"text": "⚙️ Set Global Limit", "callback_data": "st:scan:setglobal"}],
            [{"text": f"⏳ Scanner Loading: {scanner_duration_label(cur_scan_duration)} ({cur_scan_loading_count}×5s)", "callback_data": "st:noop"}],
            [
                {"text": ("✅ 15s" if cur_scan_duration == 15 else "15s"), "callback_data": "st:scan:dur:15"},
                {"text": ("✅ 25s" if cur_scan_duration == 25 else "25s"), "callback_data": "st:scan:dur:25"},
            ],
            [
                {"text": ("✅ 40s" if cur_scan_duration == 40 else "40s"), "callback_data": "st:scan:dur:40"},
                {"text": ("✅ 60s" if cur_scan_duration == 60 else "60s"), "callback_data": "st:scan:dur:60"},
            ],
            [{"text": "📊 View Usage", "callback_data": "st:scan:viewusage"},
             {"text": "🔄 Reset Usage", "callback_data": "st:scan:reset"}],
            [{"text": "✏️ Custom Limit Msg", "callback_data": "st:scan:editmsg"}],
            [{"text": ("✏️ Edit Joined Msg ✅" if joined_msg_set else "✏️ Edit Joined Msg"), "callback_data": "st:scan:editjoined"},
             {"text": "👁️ Preview Joined Msg", "callback_data": "st:preview:joined"}],
            [{"text": "📖 How: Add Scanner", "callback_data": "st:how:addscanner"},
             {"text": "📖 How: Add Games", "callback_data": "st:how:addgames"}],
            [{"text": "📖 How: Joined Msg", "callback_data": "st:how:setjoinedmsg"}],
        ])

    elif cat == "withdraw":
        kb["inline_keyboard"].extend([
            [{"text": "✉️ WITHDRAWAL MESSAGES", "callback_data": "st:noop"}],
            [{"text": "✏️ Edit Request Msg", "callback_data": "st:withdrawal:editrequest"}],
            [{"text": "✏️ Edit Approve Msg", "callback_data": "st:withdrawal:editapprove"},
             {"text": "✏️ Edit Reject Msg", "callback_data": "st:withdrawal:editreject"}],
            [{"text": "✏️ Edit Failed Msg", "callback_data": "st:withdrawal:editfailed"},
             {"text": "✏️ Edit Submitted Msg", "callback_data": "st:withdrawal:editsubmitted"}],
        ])

    elif cat == "verify":
        kb["inline_keyboard"].extend([
            [{"text": "📢 VERIFICATION MESSAGES", "callback_data": "st:noop"}],
            [{"text": "✏️ JoinLock Msg", "callback_data": "st:verify:editjoin"},
             {"text": "✏️ Contact Request", "callback_data": "st:verify:editcontact"}],
            [{"text": "✏️ Pending Msg", "callback_data": "st:verify:editpending"},
             {"text": "✏️ Verified Msg", "callback_data": "st:verify:editverified"}],
            [{"text": "✏️ Rejected Msg", "callback_data": "st:verify:editrejected"},
             {"text": "✏️ Group Contact", "callback_data": "st:verify:editgroupcontact"}],
        ])

    elif cat == "callback":
        prev_page = page - 1 if page > 1 else 1
        next_page = page + 1 if page < pages else pages
        kb["inline_keyboard"].extend([
            [{"text": "🧩 CALLBACKS / ACTIONS", "callback_data": "st:noop"}],
            [{"text": "🧩 Inplace ON", "callback_data": "st:inplace:on"},
             {"text": "🧩 Inplace OFF", "callback_data": "st:inplace:off"}],
            [{"text": "📖 How: Callback", "callback_data": "st:how:callback"},
             {"text": "📖 How: Command", "callback_data": "st:how:setcommand"}],
            [{"text": "🗑 Delete Callback", "callback_data": "st:cbdelmenu:1"}],
            [{"text": "⬅️ Prev", "callback_data": f"st:cbpage:{prev_page}"},
             {"text": f"📄 {page}/{pages}", "callback_data": "st:noop"},
             {"text": "Next ➡️", "callback_data": f"st:cbpage:{next_page}"}],
        ])

    elif cat == "data":
        kb["inline_keyboard"].extend([
            [{"text": "📊 DATA & EXPORT", "callback_data": "st:noop"}],
            [{"text": "📤 Export ALL Users", "callback_data": "st:export:all"},
             {"text": "✅ Export VERIFIED", "callback_data": "st:export:verified"}],
            [{"text": "📃 My Bots", "callback_data": "st:mybots:0"}],
            [{"text": "📋 Clone Bot", "callback_data": "st:how:clone"}],
        ])

    elif cat == "utils":
        lg_on = bot_row.get("livegram", False)
        lg_scope = bot_row.get("livegram_scope") or "private"
        lg_status = "🟢 ON" if lg_on else "🔴 OFF"
        lg_scope_label = "📨 All (Private+Group)" if lg_scope == "all" else "📨 Private Only"

        kb["inline_keyboard"].extend([
            [{"text": "🔧 UTILITIES", "callback_data": "st:noop"}],
            [{"text": f"💬 Livegram: {lg_status}", "callback_data": f"st:livegram:{'off' if lg_on else 'on'}"}],
            [{"text": lg_scope_label, "callback_data": f"st:livegram:scope:{'private' if lg_scope == 'all' else 'all'}"}],
            [{"text": "📚 All Commands Help", "callback_data": "st:help:all"}],
            [{"text": "🧠 Full Placeholders", "callback_data": "st:placeholders:full"}],
            [{"text": "📖 How: Broadcast", "callback_data": "st:how:broadcast"},
             {"text": "📖 How: AddBot", "callback_data": "st:how:addbot"}],
        ])

    # Always add Back button
    kb["inline_keyboard"].append([
        {"text": "⬅️ Kembali ke Menu", "callback_data": "st:refresh:1"},
    ])

    return kb


def build_settings_keyboard(bot_row: dict, page: int, pages: int, cat: Optional[str] = None):
    """Default keyboard: show category nav menu. If cat given, show sub-menu."""
    if cat:
        return build_settings_keyboard_by_category(bot_row, cat, page, pages)
    return build_settings_category_nav()
def send_or_edit_settings_panel(bot_row: dict, chat_id: int, uid: int, page: int = 1, edit_ctx: Optional[dict] = None, cat: Optional[str] = None):
    bot_id = str(bot_row["id"])
    stats = get_bot_stats(bot_id)
    cb_total, cb_rows = get_callbacks_page(bot_id, page, SETTINGS_CB_PAGE_SIZE)
    text_panel, pages = build_settings_text(bot_row, stats, cb_total, cb_rows, page, SETTINGS_CB_PAGE_SIZE)
    if cat:
        text_panel = f"🗂️ <b>Settings Category:</b> <code>{html.escape(str(cat))}</code>\n━━━━━━━━━━━━━━━━━━\n" + text_panel
    kb = build_settings_keyboard(bot_row, page, pages, cat=cat)

    token = bot_row["token"]
    if edit_ctx and edit_ctx.get("message_id"):
        edit_message(token, chat_id, edit_ctx["message_id"], text_panel, reply_markup=kb, parse_mode="HTML")
    else:
        send_message(token, chat_id, text_panel, reply_markup=kb, parse_mode="HTML")


def preview_start(bot_row: dict, chat_id: int, uid: int):
    bot_id = str(bot_row["id"])
    token = bot_row["token"]
    user_row = get_user_row(bot_id, uid) or {"user_id": uid, "first_name": "Admin", "username": None, "balance": 0, "shared_count": 0, "member_id": "000000"}

    start_text = bot_row.get("start_text") or bot_row.get("start_message") or "Selamat datang {firstname}!\n\n!1share Share Link"
    final_text = render_placeholders(start_text, bot_row.get("bot_username") or "", user_row)
    share_q = make_share_query(bot_row.get("bot_username") or "", user_row)
    final_text, markup = parse_buttons(final_text, share_inline_query=share_q)

    mt, mid = bot_row.get("start_media_type"), bot_row.get("start_media_file_id")
    if mt and mid:
        send_media(token, chat_id, mt, mid, caption="📌 <b>PREVIEW START</b>\n\n" + final_text, reply_markup=markup)
    else:
        send_message(token, chat_id, "📌 <b>PREVIEW START</b>\n\n" + final_text, reply_markup=markup)


def preview_loading(bot_row: dict, chat_id: int, uid: int):
    """Preview LOADING message in /settings.

    Fallback to a default loading text if loading_text not set.
    """
    bot_id = str(bot_row["id"])
    token = bot_row["token"]
    user_row = get_user_row(bot_id, uid) or {
        "user_id": uid,
        "first_name": "Preview",
        "username": None,
        "balance": 0,
        "shared_count": 0,
        "member_id": "000000",
    }

    load_text = bot_row.get("loading_text") or "⏳ Loading... sila tunggu"
    final_text = render_placeholders(load_text, bot_row.get("bot_username") or "", user_row)

    # Allow buttons inside loading text too (optional)
    share_q = make_share_query(bot_row.get("bot_username") or "", user_row)
    final_text, markup = parse_buttons(final_text, share_inline_query=share_q)

    mt, mid = bot_row.get("loading_media_type"), bot_row.get("loading_media_file_id")
    if mt and mid:
        send_media(token, chat_id, mt, mid, caption="📌 <b>PREVIEW LOADING</b>\n\n" + final_text, reply_markup=markup)
    else:
        send_message(token, chat_id, "📌 <b>PREVIEW LOADING</b>\n\n" + final_text, reply_markup=markup)


def preview_joined(bot_row: dict, chat_id: int, uid: int):
    """Preview JOINED message after user passes JoinLock."""
    bot_id = str(bot_row["id"])
    token = bot_row["token"]
    user_row = get_user_row(bot_id, uid) or {
        "user_id": uid,
        "first_name": "Preview",
        "username": None,
        "balance": 0,
        "shared_count": 0,
        "member_id": "000000",
    }

    joined_text = bot_row.get("joined_message") or "✅ Akses dah dibuka. Sila tekan /start"
    final_text = render_placeholders(joined_text, bot_row.get("bot_username") or "", user_row)
    share_q = make_share_query(bot_row.get("bot_username") or "", user_row)
    final_text, markup = parse_buttons(final_text, share_inline_query=share_q)

    mt, mid = bot_row.get("joined_message_media_type"), bot_row.get("joined_message_media_file_id")
    if mt and mid:
        send_media(token, chat_id, mt, mid, caption="📌 <b>PREVIEW JOINED MESSAGE</b>\n\n" + final_text, reply_markup=markup)
    else:
        send_message(token, chat_id, "📌 <b>PREVIEW JOINED MESSAGE</b>\n\n" + final_text, reply_markup=markup)


def send_joined_message(bot_row: dict, chat_id: int, user_row: Optional[dict] = None) -> None:
    """Send message after user successfully passes JoinLock."""
    token = bot_row["token"]
    uid = int((user_row or {}).get("user_id") or 0)
    joined_user = user_row or {
        "user_id": uid,
        "first_name": "Boss",
        "username": None,
        "balance": 0,
        "shared_count": 0,
        "member_id": "000000",
    }

    joined_text = bot_row.get("joined_message") or "✅ Akses dah dibuka. Sila tekan /start"
    final_text = render_placeholders(joined_text, bot_row.get("bot_username") or "", joined_user)
    share_q = make_share_query(bot_row.get("bot_username") or "", joined_user)
    final_text, markup = parse_buttons(final_text, share_inline_query=share_q)

    mt, mid = bot_row.get("joined_message_media_type"), bot_row.get("joined_message_media_file_id")
    if mt and mid:
        send_media(token, chat_id, mt, mid, caption=final_text, reply_markup=markup)
    else:
        send_message(token, chat_id, final_text, reply_markup=markup, parse_mode="HTML")


def edit_loading_message(bot_row: dict, chat_id: int, message_id: int, user_row: dict):
    """
    Premium UI: edit the SAME message into LOADING content (text or media).
    """
    token = bot_row["token"]
    load_text = bot_row.get("loading_text") or "⏳ Loading... sila tunggu"

    final_text = render_placeholders(load_text, bot_row.get("bot_username") or "", user_row)
    share_q = make_share_query(bot_row.get("bot_username") or "", user_row)
    final_text, markup = parse_buttons(final_text, share_inline_query=share_q)

    mt, mid = bot_row.get("loading_media_type"), bot_row.get("loading_media_file_id")

    # If loading has no media -> edit text
    if not mt or not mid:
        return edit_message(token, chat_id, message_id, final_text or " ", reply_markup=markup, parse_mode="HTML")

    # Try edit media (best). If fails, fallback to edit text.
    try:
        return edit_media(token, chat_id, message_id, mt, mid, caption=final_text or "", reply_markup=markup, parse_mode="HTML")
    except Exception:
        return edit_message(token, chat_id, message_id, final_text or " ", reply_markup=markup, parse_mode="HTML")

def export_users_excel(bot_row: dict, chat_id: int, target: str = "all"):
    bot_id = str(bot_row["id"])
    token = bot_row["token"]

    q = "SELECT user_id, username, first_name, phone, member_id, is_verified, is_premium, balance, shared_count, joined_at, upline_user_id FROM users WHERE bot_id=:b"
    if target == "verified":
        q += " AND is_verified=TRUE"
    q += " ORDER BY joined_at DESC"

    with engine.connect() as conn:
        rows = conn.execute(text(q), {"b": bot_id}).mappings().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "users"
    ws.append(["user_id", "username", "first_name", "phone", "member_id", "is_verified", "is_premium", "balance", "shared_count", "joined_at", "upline_user_id"])
    for r in rows:
        ws.append([
            r.get("user_id"),
            r.get("username"),
            r.get("first_name"),
            r.get("phone"),
            r.get("member_id"),
            bool(r.get("is_verified")),
            bool(r.get("is_premium")),
            float(r.get("balance") or 0),
            int(r.get("shared_count") or 0),
            str(r.get("joined_at") or ""),
            r.get("upline_user_id"),
        ])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    fname = f"users_{target}_{now_local_str('%Y%m%d_%H%M%S')}.xlsx"
    files = {"document": (fname, bio)}
    data = {"chat_id": chat_id, "caption": f"📤 <b>EXPORT USERS</b>\nTarget: <b>{target}</b>\nTotal: <b>{len(rows)}</b>", "parse_mode": "HTML"}
    tg_call(token, "sendDocument", data=data, files=files)


def send_mybots(bot_row: dict, chat_id: int, owner_id: int, page: int = 0):
    token = bot_row["token"]
    bots = list_bots_by_owner(owner_id)
    if not bots:
        send_message(token, chat_id, "📃 <b>MyBots</b>\nTiada bot dalam DB untuk owner ni.", parse_mode="HTML")
        return

    per = 8
    pages = max(1, (len(bots) + per - 1) // per)
    page = max(0, min(page, pages - 1))

    start = page * per
    chunk = bots[start:start + per]
    lines = []
    for b in chunk:
        lock = "ON" if b.get("lock_bot") else "OFF"
        lines.append(f"• <b>@{b.get('bot_username') or '-'}</b> | lock={lock} | id=<code>{b['id']}</code>")

    text_ = (
        "📃 <b>MyBots</b>\n"
        f"Owner: <code>{owner_id}</code>\n"
        f"Page: <b>{page+1}/{pages}</b>\n"
        "━━━━━━━━━━━━━━━━━━\n"
        + "\n".join(lines)
        + "\n━━━━━━━━━━━━━━━━━━\n"
        "Tip: buka bot tu & guna /settings dekat bot tersebut.\n"
    )

    kb = {"inline_keyboard": [[
        {"text": "⬅️ Prev", "callback_data": f"st:mybots:{max(0, page-1)}"},
        {"text": "Next ➡️", "callback_data": f"st:mybots:{min(pages-1, page+1)}"},
    ]]}
    send_message(token, chat_id, text_, reply_markup=kb, parse_mode="HTML")


# ---------------------------
# MULTI-BOT: /addbot
# ---------------------------
def gen_secret_token() -> str:
    alphabet = "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789-_"
    return "".join(random.choice(alphabet) for _ in range(43))


def validate_bot_token(token_: str) -> Optional[dict]:
    return tg_call(token_, "getMe")


def set_webhook_for_token(token_: str, secret_token: str) -> bool:
    if not PUBLIC_BASE_URL:
        return False
    url = f"{PUBLIC_BASE_URL}/telegram"
    data = {
        "url": url,
        "secret_token": secret_token,
        "drop_pending_updates": "true",
    }
    res = tg_call(token_, "setWebhook", data=data)
    return True if res is not None else False


def handle_addbot_start(bot_row: dict, chat_id: int, uid: int):
    if not require_admin(bot_row, uid):
        return
    bot_id = str(bot_row["id"])
    set_user_state(bot_id, uid, "await_addbot_token", payload={})
    send_message(
        bot_row["token"],
        chat_id,
        "➕ <b>ADD BOT</b>\n"
        "Hantar <b>TOKEN</b> bot baru (dari BotFather).\n\n"
        "Format token biasanya: <code>123456:ABCDEF...</code>\n\n"
        "⚠️ Token tu rahsia — jangan share dekat orang lain.",
        parse_mode="HTML",
    )


def handle_addbot_receive_token(bot_row: dict, chat_id: int, uid: int, token_text: str):
    bot_id = str(bot_row["id"])
    token = bot_row["token"]

    tok = (token_text or "").strip()
    if not re.match(r"^\d+:[A-Za-z0-9_-]{20,}$", tok):
        send_message(token, chat_id, "❌ Token format tak betul. Try lagi.")
        return

    if get_bot_by_token(tok):
        send_message(token, chat_id, "⚠️ Token ni dah ada dalam DB (bot sudah ditambah).")
        clear_user_state(bot_id, uid)
        return

    info = validate_bot_token(tok)
    if not info or not info.get("username"):
        send_message(token, chat_id, "❌ Token tak valid / getMe fail. Pastikan token betul.")
        return

    new_bot_username = info["username"]
    new_bot_tg_id = int(info["id"])
    new_id = str(uuid.uuid4())

    secret = None
    ok_insert = False
    for _ in range(5):
        s = gen_secret_token()
        try:
            with engine.begin() as conn:
                conn.execute(text("""
                    INSERT INTO bots (id, token, bot_username, secret_token, owner_id, created_at)
                    VALUES (:i, :t, :u, :s, :o, NOW())
                """), {"i": new_id, "t": tok, "u": new_bot_username, "s": s, "o": uid})
            secret = s
            ok_insert = True
            break
        except IntegrityError:
            continue

    if not ok_insert or not secret:
        send_message(token, chat_id, "❌ Gagal create secret_token (collision). Cuba lagi.")
        return

    ok_wh = set_webhook_for_token(tok, secret)

    msg = (
        "✅ <b>BOT BERJAYA DITAMBAH</b>\n"
        "━━━━━━━━━━━━━━━━━━\n"
        f"🤖 Username: <b>@{new_bot_username}</b>\n"
        f"🆔 Telegram bot id: <code>{new_bot_tg_id}</code>\n"
        f"🔐 secret_token: <code>{secret}</code>\n"
        f"🌐 webhook auto-set: <b>{'YES' if ok_wh else 'NO'}</b>\n"
        "━━━━━━━━━━━━━━━━━━\n"
    )
    if not ok_wh:
        msg += (
            "⚠️ Auto setWebhook tak jalan sebab <code>PUBLIC_BASE_URL</code> belum set.\n"
            "Set env dekat Cloud Run: <code>PUBLIC_BASE_URL=https://SERVICE_URL</code>\n"
            "Lepas tu boleh set webhook manual:\n"
            f"<code>curl -sS \"https://api.telegram.org/bot{tok}/setWebhook\" "
            f"-d \"url={PUBLIC_BASE_URL or 'https://YOUR_CLOUD_RUN_URL'}/telegram\" "
            f"-d \"secret_token={secret}\" -d \"drop_pending_updates=true\"</code>\n"
        )

    send_message(token, chat_id, msg, parse_mode="HTML")
    clear_user_state(bot_id, uid)


# ---------------------------
# CONTACT REPORT (FULL LIST)
# ---------------------------
def build_contacts_excel(bot_row: dict):
    bot_id = str(bot_row["id"])
    with engine.connect() as conn:
        rows = conn.execute(text("""
            SELECT user_id, username, first_name, phone, member_id, is_verified, is_premium,
                   phone_updated_at, joined_at, upline_user_id
            FROM users
            WHERE bot_id=:b AND phone IS NOT NULL
            ORDER BY phone_updated_at DESC NULLS LAST, joined_at DESC
        """), {"b": bot_id}).mappings().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "contacts"
    ws.append(["phone_updated_at", "user_id", "username", "first_name", "phone", "member_id", "is_verified", "is_premium", "joined_at", "upline_user_id"])
    for r in rows:
        ws.append([
            str(r.get("phone_updated_at") or ""),
            r.get("user_id"),
            r.get("username"),
            r.get("first_name"),
            r.get("phone"),
            r.get("member_id"),
            bool(r.get("is_verified")),
            bool(r.get("is_premium")),
            str(r.get("joined_at") or ""),
            r.get("upline_user_id"),
        ])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio, len(rows)


def send_contact_report_to_admin(bot_row: dict, latest_user: dict):
    if not bot_row.get("admin_group_id"):
        return
    token = bot_row["token"]

    bio, total = build_contacts_excel(bot_row)
    fname = f"contacts_latest_{now_local_str('%Y%m%d_%H%M%S')}.xlsx"

    msg_tpl = bot_row.get("group_contact_message") or (
        "📥 <b>CONTACT BARU MASUK, BOSSKU 😘</b>\n"
        "━━━━━━━━━━━━━━━━━━\n"
        "👤 Nama: <b>{firstname}</b>\n"
        "📞 Phone: <code>{phone}</code>\n"
        "🆔 UID: <code>{uid}</code>\n"
        "🧾 MemberID: <code>{member_id}</code>\n"
        "🕒 Masa: <b>{date}</b>\n"
        "━━━━━━━━━━━━━━━━━━\n"
        "Status: <b>Verified(Phone)=YES</b> | Premium=<b>{premium}</b>\n\n"
        "➡️ Kalau <b>ManualApprove ON</b>, admin tekan button bawah.\n"
    )

    latest = latest_user or {}
    premium = "YES" if latest.get("is_premium") else "NO"
    caption = msg_tpl.format(
        firstname=html.escape(str(latest.get("first_name") or "-")),
        phone=html.escape(str(latest.get("phone") or "-")),
        uid=html.escape(str(latest.get("user_id") or "-")),
        member_id=html.escape(str(latest.get("member_id") or "-")),
        date=html.escape(now_local_str("%d/%m/%Y %H:%M")),
        premium=premium,
    )

    kb = None
    if bot_row.get("manual_approval"):
        kb = {
            "inline_keyboard": [[
                {"text": "✅ Approve Premium", "callback_data": f"adm:ap:{latest.get('user_id')}"},
                {"text": "❌ Reject", "callback_data": f"adm:rj:{latest.get('user_id')}"},
            ]]
        }

    files = {"document": (fname, bio)}
    data = {
        "chat_id": bot_row["admin_group_id"],
        "caption": _trim(caption + f"\n📎 Report: <b>{total}</b> contacts (latest first)", TG_MAX_CAPTION),
        "parse_mode": "HTML",
    }
    if kb:
        data["reply_markup"] = json.dumps(kb)
    tg_call(token, "sendDocument", data=data, files=files)


# ---------------------------
# APP LOGIC HANDLERS
# ---------------------------
def handle_start(bot_row, chat_id, user, text_msg):
    bot_id, token = str(bot_row["id"]), bot_row["token"]
    uid = user.get("id")
    # Clear any stuck user state (e.g. await_addbot_token) on /start
    if uid:
        clear_user_state(bot_id, uid)
    parts = (text_msg or "").split()
    upline = int(parts[1]) if len(parts) >= 2 and parts[1].isdigit() else None

    user_row, _ = upsert_user(bot_id, user, upline)

    if not ensure_access(bot_row, chat_id, int(user.get("id")), user_row):
        return

    start_text = bot_row.get("start_text") or bot_row.get("start_message") or "Selamat datang {firstname}!"
    final_text = render_placeholders(start_text, bot_row.get("bot_username") or "", user_row)
    share_q = make_share_query(bot_row.get("bot_username") or "", user_row)
    final_text, markup = parse_buttons(final_text, share_inline_query=share_q)

    mt, mid = bot_row.get("start_media_type"), bot_row.get("start_media_file_id")
    if mt and mid:
        res = send_media(token, chat_id, mt, mid, caption=final_text, reply_markup=markup)
        if res is None:
            # Media failed (e.g. invalid file_id after clone) → fallback to text
            send_message(token, chat_id, final_text, reply_markup=markup)
    else:
        send_message(token, chat_id, final_text, reply_markup=markup)


def handle_contact(bot_row, msg):
    bot_id, token = str(bot_row["id"]), bot_row["token"]
    contact = msg.get("contact")
    uid = msg["from"]["id"]
    chat_id = msg["chat"]["id"]

    # ensure user row exists (contact can be sent without prior /start)
    try:
        upsert_user(bot_id, msg.get("from") or {}, None)
    except Exception:
        pass

    if not contact or contact.get("user_id") != uid:
        send_message(token, chat_id, "❌ Guna nombor sendiri.", reply_markup={"remove_keyboard": True})
        return

    phone = (contact.get("phone_number") or "").strip()
    if not phone:
        send_message(token, chat_id, "❌ Phone kosong / tak sah.", reply_markup={"remove_keyboard": True})
        return

    with engine.connect() as conn:
        existing_samebot = conn.execute(
            text("SELECT user_id FROM users WHERE bot_id=:b AND phone=:p"),
            {"b": bot_id, "p": phone},
        ).mappings().first()

    if existing_samebot and int(existing_samebot["user_id"]) != int(uid):
        send_message(
            token,
            chat_id,
            "❌ Nombor ni dah digunakan dalam bot ini. Sila guna nombor lain.",
            reply_markup={"remove_keyboard": True},
        )
        return

    try:
        with engine.begin() as conn:
            conn.execute(
                text("""
                    UPDATE users
                    SET phone=:p, is_verified=TRUE, phone_updated_at=NOW()
                    WHERE bot_id=:b AND user_id=:u
                """),
                {"p": phone, "b": bot_id, "u": uid},
            )
    except IntegrityError:
        send_message(token, chat_id, "❌ Nombor dah guna.", reply_markup={"remove_keyboard": True})
        return

    if bot_row.get("manual_approval"):
        with engine.begin() as conn:
            conn.execute(
                text("UPDATE users SET is_premium=COALESCE(is_premium, FALSE) WHERE bot_id=:b AND user_id=:u"),
                {"b": bot_id, "u": uid},
            )

        pending_msg = bot_row.get("pending_message") or (
            "⏳ <b>SEMAKAN PREMIUM</b>\n"
            "Bossku, contact kau dah masuk ✅\n"
            "Sekarang tunggu admin approve dulu ya 😘"
        )
        send_message(token, chat_id, pending_msg, reply_markup={"remove_keyboard": True}, parse_mode="HTML")
    else:
        send_message(token, chat_id, "✅ Contact diterima. Akaun disahkan.", reply_markup={"remove_keyboard": True})

    user_row = get_user_row(bot_id, uid) or {}
    send_contact_report_to_admin(bot_row, user_row)

    user_row2 = get_user_row(bot_id, uid) or user_row
    if bot_row.get("manual_approval") and not user_row2.get("is_premium"):
        return

    # ✅ Lepas user share contact & verified, terus keluarkan mesej /start (bukan loading)
    handle_start(bot_row, chat_id, msg.get("from") or {"id": uid}, "/start")


def _classify_tg_error(token: str, method: str, params=None, data=None, files=None):
    """
    Like tg_call but returns (result, error_category).
    error_category is None on success, or one of:
      'blocked', 'deactivated', 'not_found', 'parse_error', 'other'
    """
    try:
        r = SESSION.post(
            TG_API.format(token=token, method=method),
            params=params, data=data, files=files, timeout=25,
        )
        try:
            js = r.json()
        except Exception:
            return None, "other"

        if not js.get("ok"):
            desc = (js.get("description") or "").lower()
            code = js.get("error_code")
            if code == 403:
                if "bot was blocked" in desc:
                    return None, "blocked"
                if "user is deactivated" in desc:
                    return None, "deactivated"
                if "bot can't initiate" in desc or "bots can't send" in desc:
                    return None, "no_access"
                return None, "forbidden"
            if code == 400:
                if "chat not found" in desc:
                    return None, "not_found"
                if "parse entities" in desc or "can't parse" in desc:
                    return None, "parse_error"
                return None, "bad_request"
            if code == 429:
                return None, "rate_limit"
            return None, "other"
        return js.get("result"), None
    except Exception:
        return None, "other"


def _visible_len(html_text: str) -> int:
    """Count visible text length (exclude HTML tags). Telegram counts this for limits."""
    return len(re.sub(r'<[^>]+>', '', html_text or ""))


def _broadcast_send_one(token, uid, mt, mid, ptxt, mk):
    """Send one broadcast message, return error_category or None on success."""
    if mt and mid:
        method_map = {"photo": "sendPhoto", "video": "sendVideo", "animation": "sendAnimation", "document": "sendDocument"}
        field_map = {"photo": "photo", "video": "video", "animation": "animation", "document": "document"}
        if mt in method_map:
            # Sanitize HTML without strict trimming (Telegram counts visible chars, not raw)
            cap = sanitize_telegram_html(ptxt) if ptxt else None
            vis_len = _visible_len(cap) if cap else 0

            if vis_len > TG_MAX_CAPTION:
                # Visible text exceeds 1024 — send media first, then text separately
                d = {"chat_id": uid, field_map[mt]: mid}
                _, err = _classify_tg_error(token, method_map[mt], data=d)
                if err:
                    return err
                return _broadcast_send_text(token, uid, ptxt, mk)
            else:
                # Fits as caption — send together
                d = {"chat_id": uid, field_map[mt]: mid, "parse_mode": "HTML"}
                if cap:
                    d["caption"] = cap
                if mk:
                    d["reply_markup"] = json.dumps(mk)
                _, err = _classify_tg_error(token, method_map[mt], data=d)
                if err == "parse_error":
                    logger.warning(f"Broadcast HTML parse fail, caption={cap!r:.300}")
                    d.pop("parse_mode", None)
                    _, err2 = _classify_tg_error(token, method_map[mt], data=d)
                    return err2
                return err
        else:
            return _broadcast_send_text(token, uid, ptxt, mk)
    else:
        return _broadcast_send_text(token, uid, ptxt, mk)


def _broadcast_send_text(token, uid, ptxt, mk):
    """Send text broadcast, return error_category or None on success."""
    cap = sanitize_telegram_html(ptxt) if ptxt else ""
    d = {"chat_id": uid, "text": cap or "(empty)", "parse_mode": "HTML"}
    if mk:
        d["reply_markup"] = json.dumps(mk)
    _, err = _classify_tg_error(token, "sendMessage", data=d)
    if err == "parse_error":
        logger.warning(f"Broadcast text HTML parse fail, text={cap!r:.300}")
        d.pop("parse_mode", None)
        _, err2 = _classify_tg_error(token, "sendMessage", data=d)
        return err2
    return err


def handle_broadcast_optimized(bot_row, chat_id, admin_id, text_msg, reply_msg):
    bot_id, token = str(bot_row["id"]), bot_row["token"]
    if not require_admin(bot_row, admin_id):
        return

    lines = (text_msg or "").split("\n")
    target_ver = "verified" in (lines[0].lower() if lines else "")
    btn_conf = "\n".join(lines[1:])
    mt, mid, txt = save_content_from_reply(reply_msg)
    final_txt = (txt + "\n" + btn_conf).strip()

    q = "SELECT DISTINCT user_id FROM users WHERE bot_id=:b"
    if target_ver:
        q += " AND is_verified=TRUE"

    with engine.connect() as conn:
        users = conn.execute(text(q), {"b": bot_id}).mappings().all()
    user_ids = [u["user_id"] for u in users]

    if not user_ids:
        send_message(token, chat_id, "⚠️ Tiada user untuk broadcast.")
        return

    # Direct broadcast mode (optimized for VPS)
    send_message(token, chat_id, f"📣 Broadcasting to {len(user_ids)} users...")

    # Batch fetch all users for better performance (fixes N+1 query)
    bot_username = bot_row.get("bot_username") or ""
    stmt = text("""
        SELECT user_id, first_name, username, balance, shared_count, member_id
        FROM users
        WHERE bot_id=:b AND user_id = ANY(:uids)
    """).bindparams(sa.bindparam("uids", type_=sa.ARRAY(sa.BigInteger())))

    with engine.connect() as conn:
        user_rows = conn.execute(stmt, {"b": bot_id, "uids": user_ids}).mappings().all()

    # Create lookup dict for faster access
    user_dict = {u["user_id"]: dict(u) for u in user_rows}

    # --- Delivery tracking ---
    t_start = time.time()
    sent = 0
    fail_categories = {}  # category -> [uid, uid, ...]
    first_success_msg = None  # for proof forwarding

    for uid in user_ids:
        try:
            urow = user_dict.get(uid) or {"user_id": uid}
            ptxt = render_placeholders(final_txt, bot_username, urow)
            share_q = make_share_query(bot_username, urow)
            ptxt, mk = parse_buttons(ptxt, share_inline_query=share_q)

            err = _broadcast_send_one(token, uid, mt, mid, ptxt, mk)
            if err is None:
                sent += 1
                # Capture first successful send for proof (only need chat_id)
                if first_success_msg is None:
                    first_success_msg = uid
            else:
                fail_categories.setdefault(err, []).append(uid)
            time.sleep(BROADCAST_SLEEP)
        except Exception as e:
            fail_categories.setdefault("exception", []).append(uid)
            logger.warning(f"Broadcast send failed uid={uid}: {e}")

    t_end = time.time()
    duration = t_end - t_start
    total_failed = sum(len(v) for v in fail_categories.values())

    # --- Build delivery report ---
    target_label = "Verified Users" if target_ver else "All Users"
    media_label = mt.upper() if mt else "TEXT"

    report_lines = [
        "📣 <b>BROADCAST DELIVERY REPORT</b>",
        "━━━━━━━━━━━━━━━━━━",
        f"🎯 Target: <b>{target_label}</b>",
        f"📦 Type: <b>{media_label}</b>",
        f"👥 Total: <b>{len(user_ids)}</b>",
        "",
        f"✅ Delivered: <b>{sent}</b>",
        f"❌ Failed: <b>{total_failed}</b>",
    ]

    # Category breakdown
    category_labels = {
        "blocked": "🚫 Bot Blocked",
        "deactivated": "💀 User Deactivated",
        "not_found": "🔍 Chat Not Found",
        "no_access": "🔒 No Access",
        "forbidden": "⛔ Forbidden",
        "parse_error": "⚠️ HTML Parse Error",
        "bad_request": "❓ Bad Request",
        "rate_limit": "⏳ Rate Limited",
        "other": "❗ Other Error",
        "exception": "💥 Exception",
    }

    if fail_categories:
        report_lines.append("")
        report_lines.append("<b>Failure Breakdown:</b>")
        for cat, uids in sorted(fail_categories.items(), key=lambda x: -len(x[1])):
            label = category_labels.get(cat, cat)
            report_lines.append(f"  {label}: <b>{len(uids)}</b>")

    # Duration
    if duration < 60:
        dur_str = f"{duration:.1f}s"
    else:
        dur_str = f"{int(duration // 60)}m {int(duration % 60)}s"

    report_lines.extend([
        "",
        f"🕒 Duration: <b>{dur_str}</b>",
        f"📅 {now_local_str('%d/%m/%Y %H:%M:%S')}",
        "━━━━━━━━━━━━━━━━━━",
    ])

    report_text = "\n".join(report_lines)
    send_message(token, chat_id, report_text, parse_mode="HTML")

    # --- Forward proof: send copy of broadcast to admin chat ---
    try:
        proof_txt = render_placeholders(final_txt, bot_username, {"user_id": 0, "first_name": "User", "username": "", "balance": 0, "shared_count": 0, "member_id": ""})
        proof_txt, proof_mk = parse_buttons(proof_txt, share_inline_query="")
        proof_header = "📋 <b>BROADCAST PROOF</b>\n━━━━━━━━━━━━━━━━━━\n"
        proof_full = sanitize_telegram_html(proof_header + proof_txt)

        if mt and mid:
            # Send media first (no caption), then full text separately
            send_media(token, chat_id, mt, mid, caption=None, parse_mode=None)
        send_message(token, chat_id, proof_full, reply_markup=proof_mk, parse_mode="HTML")
    except Exception as e:
        logger.warning(f"Broadcast proof forward failed: {e}")

    # --- Export failed users as Excel ---
    if fail_categories:
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "failed_users"
            ws.append(["user_id", "username", "first_name", "error_category"])
            for cat, uids in fail_categories.items():
                for fuid in uids:
                    uinfo = user_dict.get(fuid) or {}
                    ws.append([
                        fuid,
                        uinfo.get("username") or "-",
                        uinfo.get("first_name") or "-",
                        cat,
                    ])

            bio = BytesIO()
            wb.save(bio)
            bio.seek(0)

            fname = f"broadcast_failed_{now_local_str('%Y%m%d_%H%M%S')}.xlsx"
            files = {"document": (fname, bio)}
            data = {
                "chat_id": chat_id,
                "caption": f"📎 <b>FAILED LIST</b>\nTotal: <b>{total_failed}</b> users",
                "parse_mode": "HTML",
            }
            tg_call(token, "sendDocument", data=data, files=files)
        except Exception as e:
            logger.warning(f"Broadcast Excel export failed: {e}")


def handle_withdraw_request(bot_row, chat_id, user):
    bot_id = str(bot_row["id"])
    user_row, _ = upsert_user(bot_id, user, None)

    if not ensure_access(bot_row, chat_id, int(user.get("id")), user_row):
        return

    set_user_state(bot_id, user["id"], "await_withdraw")

    # Custom withdrawal request message (with media support)
    bot_latest_req = get_bot_by_id(bot_id) or bot_row
    req_mt = bot_latest_req.get("withdrawal_request_media_type")
    req_mf = bot_latest_req.get("withdrawal_request_media_file_id")
    prompt = bot_latest_req.get("withdrawal_request_message") or bot_row.get("withdrawal_prompt") or (
        "💸 <b>WITHDRAWAL REQUEST</b>\n"
        "Bossku, sila taip detail lengkap:\n"
        "✅ <b>JUMLAH</b> + <b>BANK</b> + <b>NO AKAUN</b>\n\n"
        "Contoh:\n"
        "<i>RM50 Maybank 12345678</i>"
    )
    if req_mt and req_mf:
        send_media(bot_row["token"], chat_id, req_mt, req_mf, caption=prompt, parse_mode="HTML")
    else:
        send_message(bot_row["token"], chat_id, prompt, parse_mode="HTML")


def process_withdraw(bot_row, chat_id, user, text_msg):
    bot_id, token = str(bot_row["id"]), bot_row["token"]
    uid = int(user.get("id") or 0)
    clear_user_state(bot_id, uid)

    # Check minimum balance before allowing withdrawal request
    min_wd = get_bot_min_withdraw(bot_row)

    with engine.connect() as conn:
        urow0 = conn.execute(
            text("SELECT balance FROM users WHERE bot_id=:b AND user_id=:u"),
            {"b": bot_id, "u": uid},
        ).mappings().first()

    bal0 = float((urow0 or {}).get("balance") or 0)

    if bal0 < float(min_wd):
        bot_latest = get_bot_by_id(bot_id) or bot_row
        failed_msg = build_withdraw_insufficient_msg(float(min_wd), float(bal0), bot_latest)
        mt_f = bot_latest.get("withdrawal_failed_media_type")
        mf_f = bot_latest.get("withdrawal_failed_media_file_id")
        if mt_f and mf_f:
            send_media(token, chat_id, mt_f, mf_f, caption=failed_msg, parse_mode="HTML")
        else:
            send_message(token, chat_id, failed_msg, parse_mode="HTML")
        return

    wid = str(uuid.uuid4())
    with engine.begin() as conn:
        conn.execute(
            text("INSERT INTO withdrawals (id, bot_id, user_id, request_text, request_amount) VALUES (:id, :b, :u, :r, :a)"),
            {"id": wid, "b": bot_id, "u": uid, "r": text_msg, "a": bal0},
        )

    # Custom submitted message
    bot_latest_sub = get_bot_by_id(bot_id) or bot_row
    sub_tpl = (bot_latest_sub.get("withdrawal_submitted_message") or "").strip()
    if sub_tpl:
        sub_msg = sub_tpl.replace("{balance}", f"RM{bal0:.2f}")
        mt_s = bot_latest_sub.get("withdrawal_submitted_media_type")
        mf_s = bot_latest_sub.get("withdrawal_submitted_media_file_id")
        if mt_s and mf_s:
            send_media(token, chat_id, mt_s, mf_s, caption=sub_msg, parse_mode="HTML")
        else:
            send_message(token, chat_id, sub_msg, parse_mode="HTML")
    else:
        send_message(token, chat_id, "✅ Request withdraw dihantar. Tunggu admin process ya Bossku 😘", parse_mode="HTML")

    # Admin notification
    if bot_row.get("admin_group_id"):
        uname = user.get("username")
        uname_txt = f" (@{html.escape(uname)})" if uname else ""
        rpt = (
            f"💰 <b>WITHDRAWAL REQUEST</b>\n"
            f"━━━━━━━━━━━━━━━━━━\n"
            f"👤 <b>{html.escape(str(user.get('first_name') or '-'))}</b>{uname_txt}\n"
            f"🆔 UID: <code>{uid}</code>\n"
            f"💵 <b>Baki: RM{bal0:.2f}</b>\n"
            f"📝 Detail: {html.escape(text_msg or '-')}\n"
            f"🔖 ID: <code>{wid[:8]}</code>"
        )
        kb = {
            "inline_keyboard": [[
                {"text": f"✅ Approve RM{bal0:.2f}", "callback_data": f"wd:ap:{wid}"},
                {"text": "❌ Reject", "callback_data": f"wd:rj:{wid}"},
            ]]
        }
        send_message(token, bot_row["admin_group_id"], rpt, reply_markup=kb, parse_mode="HTML")


# ---------------------------
# COMMAND PARSING + ACTIONS
# ---------------------------
def parse_command_name(text_msg: str) -> str:
    if not text_msg or not text_msg.startswith("/"):
        return ""
    first = text_msg.split()[0].strip()
    cmd = first[1:]
    if "@" in cmd:
        cmd = cmd.split("@", 1)[0]
    cmd = re.sub(r"[^a-zA-Z0-9_]", "", cmd)
    return cmd.lower()


def actions_get(bot_id: str, key: str):
    with engine.connect() as conn:
        return conn.execute(
            text("SELECT * FROM actions WHERE bot_id=:b AND key=:k"),
            {"b": bot_id, "k": key},
        ).mappings().first()


def actions_upsert(bot_id: str, key: str, ty: str, tx: str, media_id: Optional[str], delay: int):
    with engine.begin() as conn:
        conn.execute(text("""
            INSERT INTO actions (bot_id, key, type, text, media_file_id, delay_seconds)
            VALUES (:b, :k, :ty, :tx, :m, :d)
            ON CONFLICT (bot_id, key) DO UPDATE SET
              type=excluded.type, text=excluded.text, media_file_id=excluded.media_file_id, delay_seconds=excluded.delay_seconds
        """), {"b": bot_id, "k": key, "ty": ty, "tx": tx, "m": media_id, "d": delay})


# ---------------------------
# ROUTES
# ---------------------------
@app.get("/")
def index():
    paths = sorted({str(r.rule) for r in app.url_map.iter_rules()})
    return jsonify({"ok": True, "service": SERVICE_NAME, "paths": paths}), 200


@app.route("/healthz", methods=["GET", "HEAD"])
@app.route("/healthz/", methods=["GET", "HEAD"])
def healthz():
    return jsonify({"ok": True, "service": SERVICE_NAME, "ts": utcnow().isoformat()}), 200


@app.route("/health", methods=["GET", "HEAD"])
@app.route("/health/", methods=["GET", "HEAD"])
def health():
    return jsonify({"ok": True, "service": SERVICE_NAME, "ts": utcnow().isoformat()}), 200


@app.errorhandler(404)
def not_found(e):
    return jsonify({"ok": False, "err": "not_found", "path": request.path, "method": request.method}), 404



# Telegram webhook routes
@app.post("/telegram")
def telegram_webhook():
    secret = request.headers.get("X-Telegram-Bot-Api-Secret-Token", "")
    bot_row = get_bot_by_secret(secret)
    if not bot_row:
        return jsonify({"ok": True}), 200

    update = request.get_json(silent=True) or {}
    token = bot_row["token"]
    bot_id = str(bot_row["id"])

    # message
    if "message" in update:
        msg = update["message"]
        chat_id = msg["chat"]["id"]
        from_user = msg.get("from") or {}
        uid = from_user.get("id")
        text_msg = msg.get("text") or ""
        # Also check caption for commands sent with media
        if not text_msg and msg.get("caption"):
            text_msg = msg["caption"]
        cmd = (text_msg.split()[0] if text_msg else "").split("@")[0]

        # Debug: log admin commands
        if text_msg.startswith("/") and uid == int(bot_row.get("owner_id", 0)):
            has_reply = "reply_to_message" in msg
            reply_has_photo = bool(msg.get("reply_to_message", {}).get("photo")) if has_reply else False
            logger.info(f"[DEBUG-CMD] cmd={cmd} uid={uid} chat={chat_id} has_reply={has_reply} reply_has_photo={reply_has_photo} has_caption={bool(msg.get('caption'))} has_photo={bool(msg.get('photo'))}")

        if not uid:
            return "OK", 200

        # LIVEGRAM: detect admin reply in admin group → route back to user
        # Skip livegram if admin has a pending_inputs action (e.g. setting custom message)
        admin_gid = bot_row.get("admin_group_id")
        if admin_gid and chat_id == int(admin_gid) and bot_row.get("livegram") and msg.get("reply_to_message"):
            if (bot_id, uid) not in pending_inputs:
                if livegram_handle_admin_reply(bot_row, msg):
                    return "OK", 200

        # LIVEGRAM: forward user messages to admin group (fire-and-forget, don't return)
        if bot_row.get("livegram") and not require_admin(bot_row, uid):
            _chat_type = msg.get("chat", {}).get("type", "private")
            scope = bot_row.get("livegram_scope") or "private"
            is_admin_grp = admin_gid and chat_id == int(admin_gid)
            if not is_admin_grp:
                if _chat_type == "private" or scope == "all":
                    livegram_forward_to_admin(bot_row, msg)

        # IMPORTANT: handle /start with referral BEFORE creating user row without upline
        if text_msg and text_msg.startswith("/start"):
            handle_start(bot_row, chat_id, from_user, text_msg)
            return "OK", 200

        # Contact share can arrive without prior /start; ensure row exists inside handle_contact()
        if msg.get("contact"):
            handle_contact(bot_row, msg)
            return "OK", 200

        # For all other updates, ensure user row exists
        upsert_user(bot_id, from_user, None)


        state = get_user_state(bot_id, uid)
        if state:
            # Auto-expire states older than 30 minutes
            state_age = state.get("created_at")
            if state_age:
                from datetime import timezone
                age_secs = (datetime.now(timezone.utc) - state_age).total_seconds() if hasattr(state_age, 'tzinfo') and state_age.tzinfo else (datetime.utcnow() - state_age).total_seconds()
                if age_secs > 1800:  # 30 min
                    clear_user_state(bot_id, uid)
                    state = None
            # Skip state handling for commands (let them fall through to command handlers)
            if state and text_msg and text_msg.startswith("/"):
                clear_user_state(bot_id, uid)
                state = None
        if state:
            if state.get("state") == "await_withdraw" and text_msg:
                process_withdraw(bot_row, chat_id, from_user, text_msg)
                return "OK", 200
            if state.get("state") == "await_addbot_token" and text_msg:
                if not require_admin(bot_row, uid):
                    clear_user_state(bot_id, uid)
                    return "OK", 200
                handle_addbot_receive_token(bot_row, chat_id, uid, text_msg)
                return "OK", 200

        # commands
        if text_msg.startswith("/start"):
            handle_start(bot_row, chat_id, from_user, text_msg)
            return "OK", 200

        elif text_msg.startswith("/withdraw"):
            handle_withdraw_request(bot_row, chat_id, from_user)

        elif text_msg.startswith("/myid"):
            send_message(token, chat_id, f"🆔 Your Scanner ID: <code>{uid}</code>", parse_mode="HTML")

        elif text_msg.startswith("/mybots") and require_admin(bot_row, uid):
            send_mybots(bot_row, chat_id, int(bot_row["owner_id"]), page=0)

        elif text_msg.startswith(("/settings", "/setting")) and require_admin(bot_row, uid):
            send_or_edit_settings_panel(bot_row, chat_id, uid, page=1)

        elif text_msg.startswith("/addbot") and require_admin(bot_row, uid):
            handle_addbot_start(bot_row, chat_id, uid)

        elif text_msg.startswith("/clone"):
            if not is_owner(uid, bot_row):
                send_message(token, chat_id, "❌ Owner sahaja boleh /clone", parse_mode="HTML")
            else:
                parts = text_msg.split(maxsplit=1)
                if len(parts) < 2 or not parts[1].strip():
                    send_message(token, chat_id,
                        "📋 <b>CLONE DATA</b>\n\n"
                        "Copy setup/content dari bot lain ke bot ini.\n\n"
                        "<b>Cara guna:</b>\n"
                        "<code>/clone @usernamebot</code>\n\n"
                        "<b>Data yang di-clone:</b>\n"
                        "• Start message & loading message\n"
                        "• Scanner media & game list\n"
                        "• Custom commands/callbacks\n"
                        "• Bot settings (joinlock, referral, dll)\n\n"
                        "<b>Data yang TIDAK di-clone:</b>\n"
                        "• Users / contacts\n"
                        "• Referral balance\n"
                        "• Withdrawals & scan history",
                        parse_mode="HTML",
                    )
                else:
                    target_username = parts[1].strip()
                    source_bot = get_bot_by_username(target_username)
                    if not source_bot:
                        send_message(token, chat_id,
                            f"❌ Bot <code>{_h(target_username)}</code> tidak dijumpai dalam sistem.",
                            parse_mode="HTML",
                        )
                    elif str(source_bot["id"]) == bot_id:
                        send_message(token, chat_id, "⚠️ Tak boleh clone dari bot yang sama.", parse_mode="HTML")
                    elif int(source_bot["owner_id"]) != uid:
                        send_message(token, chat_id, "❌ Kau bukan owner bot source tu.", parse_mode="HTML")
                    else:
                        # Prevent duplicate clone runs (Telegram may retry)
                        import time as _t
                        _clone_key = f"{bot_id}"
                        _now = _t.time()
                        if _clone_key in _clone_in_progress and (_now - _clone_in_progress[_clone_key]) < 180:
                            send_message(token, chat_id, "⏳ Clone masih berjalan... sila tunggu.", parse_mode="HTML")
                            return "OK", 200
                        _clone_in_progress[_clone_key] = _now
                        send_message(token, chat_id, "⏳ Cloning data... sila tunggu (media sedang di-transfer).", parse_mode="HTML")
                        # Run clone in background thread to avoid webhook timeout
                        import threading
                        def _do_clone():
                            try:
                                result = clone_bot_data(str(source_bot["id"]), bot_id, chat_id=chat_id)
                                if "error" in result:
                                    send_message(token, chat_id, f"❌ Error: {result['error']}", parse_mode="HTML")
                                else:
                                    src_name = source_bot.get("bot_username") or str(source_bot["id"])[:8]
                                    send_message(token, chat_id,
                                        f"✅ <b>CLONE BERJAYA!</b>\n\n"
                                        f"Data dari <code>@{_h(src_name)}</code> telah di-copy.\n\n"
                                        f"📊 <b>Summary:</b>\n"
                                        f"• Scanner media: <b>{result['media']}</b> provider\n"
                                        f"• Game list: <b>{result['games']}</b> games\n"
                                        f"• Custom actions: <b>{result['actions']}</b> items\n"
                                        f"• Media re-uploaded: <b>{result.get('reuploaded', 0)}</b> files\n"
                                        f"• Bot settings: ✅ updated\n\n"
                                        f"🔍 <b>Debug:</b>\n"
                                        f"• Source ID: <code>{result.get('source_id','?')}</code>\n"
                                        f"• Source user: <code>{result.get('source_user','?')}</code>\n"
                                        f"• Start text len: {result.get('start_text_len',0)}\n\n"
                                        f"Taip /settings untuk check.",
                                        parse_mode="HTML",
                                    )
                            except Exception as e:
                                logger.error("Clone error: %s", e)
                                send_message(token, chat_id, f"❌ Clone error: {e}", parse_mode="HTML")
                            finally:
                                _clone_in_progress.pop(_clone_key, None)
                        threading.Thread(target=_do_clone, daemon=True).start()
                    return "OK", 200

        # NEW admin commands (owner only untuk add/del)
        elif text_msg.startswith("/admins") and require_admin(bot_row, uid):
            rows = list_admins(bot_id)
            if not rows:
                send_message(token, chat_id, "👥 <b>ADMINS</b>\nTiada admin untuk bot ni.", parse_mode="HTML")
            else:
                lines = []
                for r in rows:
                    exp = r["expiry_at"]
                    exp_txt = "PERMANENT" if exp is None else str(exp)
                    lines.append(f"• <code>{r['admin_user_id']}</code> | exp: <code>{exp_txt}</code> | by: <code>{r['added_by']}</code>")
                send_message(token, chat_id, "👥 <b>ADMINS</b>\n" + "\n".join(lines), parse_mode="HTML")

        elif text_msg.startswith("/addadmin"):
            if not require_admin(bot_row, uid):
                send_message(token, chat_id, "❌ Owner/Admin sahaja boleh /addadmin", parse_mode="HTML")
            else:
                parts = text_msg.split()
                if len(parts) < 2 or not parts[1].isdigit():
                    send_message(token, chat_id, "Format: <code>/addadmin 123456789</code> atau <code>/addadmin 123456789 30</code>", parse_mode="HTML")
                else:
                    target = int(parts[1])
                    days = int(parts[2]) if len(parts) >= 3 and parts[2].isdigit() else ADMIN_DEFAULT_DAYS
                    add_admin(bot_id, target, uid, days)
                    send_message(token, chat_id, f"✅ Admin added: <code>{target}</code> (expiry {days} hari).", parse_mode="HTML")

        elif text_msg.startswith("/deladmin"):
            if not is_owner(uid, bot_row):
                send_message(token, chat_id, "❌ Owner sahaja boleh /deladmin", parse_mode="HTML")
            else:
                parts = text_msg.split()
                if len(parts) < 2 or not parts[1].isdigit():
                    send_message(token, chat_id, "Format: <code>/deladmin 123456789</code>", parse_mode="HTML")
                else:
                    target = int(parts[1])
                    ok = del_admin(bot_id, target)
                    send_message(token, chat_id, ("✅ Admin removed" if ok else "⚠️ Admin tak jumpa") + f": <code>{target}</code>", parse_mode="HTML")

        elif require_admin(bot_row, uid):
            # Direct /setstart or /setloading command bypasses pending inputs
            if text_msg.startswith(("/setstart", "/setloading")) and msg.get("reply_to_message"):
                # Clear any pending input first
                pending_inputs.pop((bot_id, uid), None)
                rep = msg["reply_to_message"]
                mt, mid, txt = save_content_from_reply(rep)
                extra = "\n".join(text_msg.split("\n")[1:]).strip()
                final_txt = (txt + ("\n" + extra if extra else "")).strip()

                col_txt = "start_text" if text_msg.startswith("/setstart") else "loading_text"
                col_type = "start_media_type" if text_msg.startswith("/setstart") else "loading_media_type"
                col_file = "start_media_file_id" if text_msg.startswith("/setstart") else "loading_media_file_id"

                with engine.begin() as conn:
                    conn.execute(
                        text(f"UPDATE bots SET {col_txt}=:t, {col_type}=:mt, {col_file}=:mf WHERE id=:i"),
                        {"t": final_txt, "mt": mt, "mf": mid, "i": bot_id},
                    )
                logger.info(f"[CMD] /{col_txt} updated by uid={uid} bot_id={bot_id}")
                send_message(token, chat_id, f"✅ {col_txt} Updated.", parse_mode="HTML")
                return "OK", 200

            #  PENDING INPUTS HANDLER (interactive flows from settings panel)
            if (bot_id, uid) in pending_inputs:
                action, ts = pending_inputs[(bot_id, uid)]
                
                # Timeout after 5 minutes
                if time.time() - ts > 300:
                    del pending_inputs[(bot_id, uid)]
                    send_message(token, chat_id, "⏱ Input timeout. Please try again.", parse_mode="HTML")
                    return "OK", 200
                
                if action == "setshare":
                    try:
                        amt = float(text_msg.strip())
                        if amt <= 0:
                            raise ValueError("Amount must be positive")
                        with engine.begin() as conn:
                            conn.execute(text("UPDATE bots SET affiliate_amount=:a WHERE id=:i"), {"a": amt, "i": bot_id})
                        del pending_inputs[(bot_id, uid)]
                        send_message(token, chat_id, f"✅ Share commission updated: <b>RM{amt:.2f}</b> per click", parse_mode="HTML")
                    except Exception:
                        send_message(token, chat_id, "❌ Invalid amount. Must be a positive number (e.g. 1.00)", parse_mode="HTML")
                    return "OK", 200
                
                elif action == "setminwd":
                    try:
                        amt = float(text_msg.strip())
                        if amt <= 0:
                            raise ValueError("Amount must be positive")
                        with engine.begin() as conn:
                            conn.execute(text("UPDATE bots SET min_withdraw_amount=:a WHERE id=:i"), {"a": amt, "i": bot_id})
                        del pending_inputs[(bot_id, uid)]
                        send_message(token, chat_id, f"✅ Min withdraw updated: <b>RM{amt:.2f}</b>", parse_mode="HTML")
                    except Exception:
                        send_message(token, chat_id, "❌ Invalid amount. Must be a positive number (e.g. 30.00)", parse_mode="HTML")
                    return "OK", 200
                
                elif action == "editstart":
                    rep = msg
                    mt, mid, txt = save_content_from_reply(rep)
                    with engine.begin() as conn:
                        conn.execute(text("""
                            UPDATE bots
                            SET start_text=:t,
                                start_media_type=:mt,
                                start_media_file_id=:mf
                            WHERE id=:i
                        """), {"t": txt, "mt": mt, "mf": mid, "i": bot_id})
                    del pending_inputs[(bot_id, uid)]
                    send_message(token, chat_id, "✅ START message updated!", parse_mode="HTML")
                    return "OK", 200
                
                elif action == "editloading":
                    rep = msg
                    mt, mid, txt = save_content_from_reply(rep)
                    with engine.begin() as conn:
                        conn.execute(text("""
                            UPDATE bots
                            SET loading_message=:t,
                                loading_media_type=:mt,
                                loading_media_file_id=:mf
                            WHERE id=:i
                        """), {"t": txt, "mt": mt, "mf": mid, "i": bot_id})
                    del pending_inputs[(bot_id, uid)]
                    send_message(token, chat_id, "✅ LOADING message updated!", parse_mode="HTML")
                    return "OK", 200
                
                elif action == "scanlimit":
                    input_val = text_msg.strip().lower()
                    if input_val in ("off", "0", "unlimited", "none"):
                        with engine.begin() as conn:
                            conn.execute(text("UPDATE bots SET scan_limit_per_day=NULL WHERE id=:i"), {"i": bot_id})
                        del pending_inputs[(bot_id, uid)]
                        send_message(token, chat_id, "✅ Scan limit set to UNLIMITED", parse_mode="HTML")
                    elif input_val.isdigit():
                        lim = int(input_val)
                        with engine.begin() as conn:
                            conn.execute(text("UPDATE bots SET scan_limit_per_day=:l WHERE id=:i"), {"l": lim, "i": bot_id})
                        del pending_inputs[(bot_id, uid)]
                        send_message(token, chat_id, f"✅ Scan limit set to <b>{lim}</b> scans/day", parse_mode="HTML")
                    else:
                        send_message(token, chat_id, "❌ Invalid input. Use a number or 'off'", parse_mode="HTML")
                    return "OK", 200
                
                elif action == "scanlimitmsg":
                    rep = msg
                    mt, mid, txt = save_content_from_reply(rep)
                    with engine.begin() as conn:
                        conn.execute(text("""
                            UPDATE bots
                            SET scan_limit_message=:t,
                                scan_limit_media_type=:mt,
                                scan_limit_media_file_id=:mf
                            WHERE id=:i
                        """), {"t": txt, "mt": mt, "mf": mid, "i": bot_id})
                    del pending_inputs[(bot_id, uid)]
                    send_message(token, chat_id, "✅ Custom scan limit message updated!", parse_mode="HTML")
                    return "OK", 200
                
                # Withdrawal Messages
                elif action == "withdrawalrequest":
                    rep = msg
                    mt, mid, txt = save_content_from_reply(rep)
                    with engine.begin() as conn:
                        conn.execute(text("UPDATE bots SET withdrawal_request_message=:t, withdrawal_request_media_type=:mt, withdrawal_request_media_file_id=:mf WHERE id=:i"), 
                                   {"t": txt, "mt": mt, "mf": mid, "i": bot_id})
                    del pending_inputs[(bot_id, uid)]
                    send_message(token, chat_id, "✅ Withdrawal REQUEST message updated!", parse_mode="HTML")
                    return "OK", 200
                
                elif action == "withdrawalapprove":
                    rep = msg
                    mt, mid, txt = save_content_from_reply(rep)
                    with engine.begin() as conn:
                        conn.execute(text("UPDATE bots SET withdrawal_approve_message=:t, withdrawal_approve_media_type=:mt, withdrawal_approve_media_file_id=:mf WHERE id=:i"),
                                   {"t": txt, "mt": mt, "mf": mid, "i": bot_id})
                    del pending_inputs[(bot_id, uid)]
                    send_message(token, chat_id, "✅ Withdrawal APPROVE message updated!", parse_mode="HTML")
                    return "OK", 200
                
                elif action == "withdrawalreject":
                    rep = msg
                    mt, mid, txt = save_content_from_reply(rep)
                    with engine.begin() as conn:
                        conn.execute(text("UPDATE bots SET withdrawal_reject_message=:t, withdrawal_reject_media_type=:mt, withdrawal_reject_media_file_id=:mf WHERE id=:i"),
                                   {"t": txt, "mt": mt, "mf": mid, "i": bot_id})
                    del pending_inputs[(bot_id, uid)]
                    send_message(token, chat_id, "✅ Withdrawal REJECT message updated!", parse_mode="HTML")
                    return "OK", 200
                
                elif action == "withdrawalfailed":
                    rep = msg
                    mt, mid, txt = save_content_from_reply(rep)
                    with engine.begin() as conn:
                        conn.execute(text("UPDATE bots SET withdrawal_failed_message=:t, withdrawal_failed_media_type=:mt, withdrawal_failed_media_file_id=:mf WHERE id=:i"),
                                   {"t": txt, "mt": mt, "mf": mid, "i": bot_id})
                    del pending_inputs[(bot_id, uid)]
                    send_message(token, chat_id, "✅ Withdrawal FAILED message updated!", parse_mode="HTML")
                    return "OK", 200
                
                elif action == "withdrawalsubmitted":
                    rep = msg
                    mt, mid, txt = save_content_from_reply(rep)
                    with engine.begin() as conn:
                        conn.execute(text("UPDATE bots SET withdrawal_submitted_message=:t, withdrawal_submitted_media_type=:mt, withdrawal_submitted_media_file_id=:mf WHERE id=:i"),
                                   {"t": txt, "mt": mt, "mf": mid, "i": bot_id})
                    del pending_inputs[(bot_id, uid)]
                    send_message(token, chat_id, "✅ Withdrawal SUBMITTED message updated!", parse_mode="HTML")
                    return "OK", 200
                
                # Verification Messages
                elif action == "joinmsg":
                    rep = msg
                    mt, mid, txt = save_content_from_reply(rep)
                    with engine.begin() as conn:
                        conn.execute(text("UPDATE bots SET join_message=:t, join_message_media_type=:mt, join_message_media_file_id=:mf WHERE id=:i"),
                                   {"t": txt, "mt": mt, "mf": mid, "i": bot_id})
                    del pending_inputs[(bot_id, uid)]
                    send_message(token, chat_id, "✅ JoinLock message updated!", parse_mode="HTML")
                    return "OK", 200

                elif action == "joinedmsg":
                    rep = msg
                    mt, mid, txt = save_content_from_reply(rep)
                    with engine.begin() as conn:
                        conn.execute(text("UPDATE bots SET joined_message=:t, joined_message_media_type=:mt, joined_message_media_file_id=:mf WHERE id=:i"),
                                   {"t": txt, "mt": mt, "mf": mid, "i": bot_id})
                    del pending_inputs[(bot_id, uid)]
                    send_message(token, chat_id, "✅ Joined message updated!", parse_mode="HTML")
                    return "OK", 200
                
                elif action == "contactmsg":
                    rep = msg
                    mt, mid, txt = save_content_from_reply(rep)
                    with engine.begin() as conn:
                        conn.execute(text("UPDATE bots SET contact_message=:t, contact_message_media_type=:mt, contact_message_media_file_id=:mf WHERE id=:i"),
                                   {"t": txt, "mt": mt, "mf": mid, "i": bot_id})
                    del pending_inputs[(bot_id, uid)]
                    send_message(token, chat_id, "✅ Contact request message updated!", parse_mode="HTML")
                    return "OK", 200
                
                elif action == "pendingmsg":
                    rep = msg
                    mt, mid, txt = save_content_from_reply(rep)
                    with engine.begin() as conn:
                        conn.execute(text("UPDATE bots SET pending_message=:t, pending_message_media_type=:mt, pending_message_media_file_id=:mf WHERE id=:i"),
                                   {"t": txt, "mt": mt, "mf": mid, "i": bot_id})
                    del pending_inputs[(bot_id, uid)]
                    send_message(token, chat_id, "✅ Pending message updated!", parse_mode="HTML")
                    return "OK", 200
                
                elif action == "verifiedmsg":
                    rep = msg
                    mt, mid, txt = save_content_from_reply(rep)
                    with engine.begin() as conn:
                        conn.execute(text("UPDATE bots SET verified_message=:t, verified_message_media_type=:mt, verified_message_media_file_id=:mf WHERE id=:i"),
                                   {"t": txt, "mt": mt, "mf": mid, "i": bot_id})
                    del pending_inputs[(bot_id, uid)]
                    send_message(token, chat_id, "✅ Verified message updated!", parse_mode="HTML")
                    return "OK", 200
                
                elif action == "rejectedmsg":
                    rep = msg
                    mt, mid, txt = save_content_from_reply(rep)
                    with engine.begin() as conn:
                        conn.execute(text("UPDATE bots SET rejected_message=:t, rejected_message_media_type=:mt, rejected_message_media_file_id=:mf WHERE id=:i"),
                                   {"t": txt, "mt": mt, "mf": mid, "i": bot_id})
                    del pending_inputs[(bot_id, uid)]
                    send_message(token, chat_id, "✅ Rejected message updated!", parse_mode="HTML")
                    return "OK", 200
                
                elif action == "groupcontactmsg":
                    rep = msg
                    mt, mid, txt = save_content_from_reply(rep)
                    with engine.begin() as conn:
                        conn.execute(text("UPDATE bots SET group_contact_message=:t, group_contact_message_media_type=:mt, group_contact_message_media_file_id=:mf WHERE id=:i"),
                                   {"t": txt, "mt": mt, "mf": mid, "i": bot_id})
                    del pending_inputs[(bot_id, uid)]
                    send_message(token, chat_id, "✅ Group contact message updated!", parse_mode="HTML")
                    return "OK", 200

            if text_msg.startswith("/broadcast") and msg.get("reply_to_message"):
                import threading
                threading.Thread(
                    target=handle_broadcast_optimized,
                    args=(bot_row, chat_id, uid, text_msg, msg["reply_to_message"]),
                    daemon=True,
                ).start()
                return "OK", 200

            elif text_msg.startswith("/setjoinmsg"):
                if not msg.get("reply_to_message"):
                    send_message(token, chat_id, "❌ Reply ke mesej yang kau nak jadikan <b>JoinLock message</b>, kemudian tulis <code>/setjoinmsg</code>", parse_mode="HTML")
                else:
                    mt, mid, txt = save_content_from_reply(msg["reply_to_message"])
                    txt = merge_command_extra(txt, text_msg, "/setjoinmsg")
                    with engine.begin() as conn:
                        conn.execute(text("UPDATE bots SET join_message=:t, join_message_media_type=:mt, join_message_media_file_id=:mf WHERE id=:i"), {"t": txt, "mt": mt, "mf": mid, "i": bot_id})
                    send_message(token, chat_id, "✅ join_message updated.", parse_mode="HTML")

            elif text_msg.startswith("/setjoinedmsg"):
                if not msg.get("reply_to_message"):
                    send_message(token, chat_id, "❌ Reply ke mesej yang kau nak jadikan <b>Joined message</b>, kemudian tulis <code>/setjoinedmsg</code>", parse_mode="HTML")
                else:
                    mt, mid, txt = save_content_from_reply(msg["reply_to_message"])
                    txt = merge_command_extra(txt, text_msg, "/setjoinedmsg")
                    with engine.begin() as conn:
                        conn.execute(text("UPDATE bots SET joined_message=:t, joined_message_media_type=:mt, joined_message_media_file_id=:mf WHERE id=:i"), {"t": txt, "mt": mt, "mf": mid, "i": bot_id})
                    send_message(token, chat_id, "✅ joined_message updated.", parse_mode="HTML")

            # JOINLOCK set list
            elif text_msg.startswith("/setjoin"):
                if msg.get("reply_to_message"):
                    raw = (msg["reply_to_message"].get("text") or msg["reply_to_message"].get("caption") or "").strip()
                else:
                    parts = text_msg.split(maxsplit=1)
                    raw = parts[1].strip() if len(parts) > 1 else ""
                with engine.begin() as conn:
                    conn.execute(text("UPDATE bots SET join_targets=:t WHERE id=:i"), {"t": raw, "i": bot_id})
                send_message(token, chat_id, "✅ join_targets updated.", parse_mode="HTML")

            elif text_msg.startswith("/setcontactmsg"):
                if not msg.get("reply_to_message"):
                    send_message(token, chat_id, "❌ Reply ke mesej yang kau nak jadikan <b>Contact Request message</b>, kemudian tulis <code>/setcontactmsg</code>", parse_mode="HTML")
                else:
                    mt, mid, txt = save_content_from_reply(msg["reply_to_message"])
                    with engine.begin() as conn:
                        conn.execute(text("UPDATE bots SET contact_message=:t, contact_message_media_type=:mt, contact_message_media_file_id=:mf WHERE id=:i"), {"t": txt, "mt": mt, "mf": mid, "i": bot_id})
                    send_message(token, chat_id, "✅ contact_message updated.", parse_mode="HTML")

            elif text_msg.startswith("/setpendingmsg"):
                if not msg.get("reply_to_message"):
                    send_message(token, chat_id, "❌ Reply ke mesej yang kau nak jadikan <b>Pending message</b>, kemudian tulis <code>/setpendingmsg</code>", parse_mode="HTML")
                else:
                    mt, mid, txt = save_content_from_reply(msg["reply_to_message"])
                    with engine.begin() as conn:
                        conn.execute(text("UPDATE bots SET pending_message=:t, pending_message_media_type=:mt, pending_message_media_file_id=:mf WHERE id=:i"), {"t": txt, "mt": mt, "mf": mid, "i": bot_id})
                    send_message(token, chat_id, "✅ pending_message updated.", parse_mode="HTML")

            elif text_msg.startswith("/setverifiedmsg"):
                if not msg.get("reply_to_message"):
                    send_message(token, chat_id, "❌ Reply ke mesej yang kau nak jadikan <b>Verified message</b>, kemudian tulis <code>/setverifiedmsg</code>", parse_mode="HTML")
                else:
                    mt, mid, txt = save_content_from_reply(msg["reply_to_message"])
                    with engine.begin() as conn:
                        conn.execute(text("UPDATE bots SET verified_message=:t, verified_message_media_type=:mt, verified_message_media_file_id=:mf WHERE id=:i"), {"t": txt, "mt": mt, "mf": mid, "i": bot_id})
                    send_message(token, chat_id, "✅ verified_message updated.", parse_mode="HTML")

            elif text_msg.startswith("/setrejectedmsg"):
                if not msg.get("reply_to_message"):
                    send_message(token, chat_id, "❌ Reply ke mesej yang kau nak jadikan <b>Rejected message</b>, kemudian tulis <code>/setrejectedmsg</code>", parse_mode="HTML")
                else:
                    mt, mid, txt = save_content_from_reply(msg["reply_to_message"])
                    with engine.begin() as conn:
                        conn.execute(text("UPDATE bots SET rejected_message=:t, rejected_message_media_type=:mt, rejected_message_media_file_id=:mf WHERE id=:i"), {"t": txt, "mt": mt, "mf": mid, "i": bot_id})
                    send_message(token, chat_id, "✅ rejected_message updated.", parse_mode="HTML")

            elif text_msg.startswith("/setgroupcontactmsg"):
                if not msg.get("reply_to_message"):
                    send_message(token, chat_id, "❌ Reply ke mesej yang kau nak jadikan <b>Group Contact message</b>, kemudian tulis <code>/setgroupcontactmsg</code>", parse_mode="HTML")
                else:
                    mt, mid, txt = save_content_from_reply(msg["reply_to_message"])
                    with engine.begin() as conn:
                        conn.execute(text("UPDATE bots SET group_contact_message=:t, group_contact_message_media_type=:mt, group_contact_message_media_file_id=:mf WHERE id=:i"), {"t": txt, "mt": mt, "mf": mid, "i": bot_id})
                    send_message(token, chat_id, "✅ group_contact_message updated.", parse_mode="HTML")

            elif text_msg.startswith("/setwithdrawmsg") and msg.get("reply_to_message"):
                raw = (msg["reply_to_message"].get("text") or msg["reply_to_message"].get("caption") or "").strip()
                with engine.begin() as conn:
                    conn.execute(text("UPDATE bots SET withdrawal_prompt=:t WHERE id=:i"), {"t": raw, "i": bot_id})
                send_message(token, chat_id, "✅ withdrawal_prompt updated.", parse_mode="HTML")

            elif text_msg.startswith("/setwithdrawalmsg") and msg.get("reply_to_message"):
                # Custom message to user when withdrawal APPROVED (supports text/media from reply)
                rep = msg["reply_to_message"]
                mt, mid, txt = save_content_from_reply(rep)
                raw = (txt or "").strip()
                with engine.begin() as conn:
                    conn.execute(text("""
                        UPDATE bots
                        SET withdrawal_approve_message=:t,
                            withdrawal_approve_media_type=:mt,
                            withdrawal_approve_media_file_id=:mf
                        WHERE id=:i
                    """), {"t": raw, "mt": mt, "mf": mid, "i": bot_id})
                send_message(token, chat_id, "✅ withdrawal APPROVE message updated.", parse_mode="HTML")

            elif text_msg.startswith("/setwithdrawalreject") and msg.get("reply_to_message"):
                # Custom message to user when withdrawal REJECTED (supports text/media from reply)
                rep = msg["reply_to_message"]
                mt, mid, txt = save_content_from_reply(rep)
                raw = (txt or "").strip()
                with engine.begin() as conn:
                    conn.execute(text("""
                        UPDATE bots
                        SET withdrawal_reject_message=:t,
                            withdrawal_reject_media_type=:mt,
                            withdrawal_reject_media_file_id=:mf
                        WHERE id=:i
                    """), {"t": raw, "mt": mt, "mf": mid, "i": bot_id})
                send_message(token, chat_id, "✅ withdrawal REJECT message updated.", parse_mode="HTML")


            elif cmd == "/setscanlimit":
                logger.info(f"[CMD] /setscanlimit uid={uid} chat={chat_id} text={text_msg}")
                # /setscanlimit 20            -> set global daily scan limit
                # /setscanlimit off|0         -> disable limit (unlimited)
                # /setscanlimit 20 @username  -> set override for a user
                # /setscanlimit reset         -> reset today's usage (all users)
                # /setscanlimit reset @user   -> reset today's usage for a user
                # /setscanlimit del @user     -> delete override for a user
                parts = (text_msg or "").split()
                arg1 = parts[1].strip() if len(parts) >= 2 else ""
                arg2 = parts[2].strip() if len(parts) >= 3 else ""

                def _resolve_target_user_id(conn, raw: str) -> Optional[int]:
                    if not raw:
                        return None
                    raw = raw.strip()
                    if raw.isdigit():
                        return int(raw)
                    return _find_user_id_by_username(conn, bot_id, raw)

                if not arg1:
                    cur_lim = bot_row.get("scan_limit_per_day")
                    cur_txt = "UNLIMITED" if (cur_lim is None or int(cur_lim or 0) <= 0) else str(int(cur_lim))
                    send_message(
                        token,
                        chat_id,
                        "📌 <b>Scan Limit</b>\n"
                        f"Global: <code>{cur_txt}</code> / hari\n\n"
                        "Cara:\n"
                        "• <code>/setscanlimit 20</code>\n"
                        "• <code>/setscanlimit off</code>\n"
                        "• <code>/setscanlimit 20 @username</code>\n"
                        "• <code>/setscanlimit reset</code> / <code>/setscanlimit reset @username</code>\n"
                        "• <code>/setscanlimit del @username</code>",
                        parse_mode="HTML",
                    )
                    return "OK", 200

                a1 = arg1.lower()

                if a1 in ("off", "0", "unlimited", "none"):
                    with engine.begin() as conn:
                        conn.execute(text("UPDATE bots SET scan_limit_per_day=NULL WHERE id=:i"), {"i": bot_id})
                    send_message(token, chat_id, "✅ Scan limit OFF (unlimited).", parse_mode="HTML")
                    return "OK", 200

                if a1 == "reset":
                    day = _today_local_date()
                    with engine.begin() as conn:
                        tgt_id = _resolve_target_user_id(conn, arg2) if arg2 else None
                        if tgt_id:
                            res = conn.execute(
                                text("DELETE FROM scan_daily_usage WHERE bot_id=:b AND user_id=:u AND day=:d"),
                                {"b": bot_id, "u": int(tgt_id), "d": day},
                            )
                            deleted = int(getattr(res, "rowcount", 0) or 0)
                            send_message(token, chat_id, f"✅ Reset scan usage hari ini untuk <code>{tgt_id}</code> (deleted {deleted}).", parse_mode="HTML")
                        else:
                            res = conn.execute(
                                text("DELETE FROM scan_daily_usage WHERE bot_id=:b AND day=:d"),
                                {"b": bot_id, "d": day},
                            )
                            deleted = int(getattr(res, "rowcount", 0) or 0)
                            send_message(token, chat_id, f"✅ Reset scan usage hari ini (all users). (deleted {deleted})", parse_mode="HTML")
                    return "OK", 200

                if a1 == "del":
                    if not arg2:
                        send_message(token, chat_id, "Format: <code>/setscanlimit del @username</code>", parse_mode="HTML")
                        return "OK", 200
                    with engine.begin() as conn:
                        tgt_id = _resolve_target_user_id(conn, arg2)
                        if not tgt_id:
                            send_message(token, chat_id, "❌ Username tak jumpa dalam DB. Pastikan user pernah /start bot.", parse_mode="HTML")
                            return "OK", 200
                        res = conn.execute(
                            text("DELETE FROM scan_limit_overrides WHERE bot_id=:b AND user_id=:u"),
                            {"b": bot_id, "u": int(tgt_id)},
                        )
                        deleted = int(getattr(res, "rowcount", 0) or 0)
                    send_message(token, chat_id, ("✅ Override removed." if deleted else "⚠️ Override not found.") + f" user_id=<code>{tgt_id}</code>", parse_mode="HTML")
                    return "OK", 200

                # set limit number
                if not re.match(r"^\d+$", a1):
                    send_message(token, chat_id, "Format: <code>/setscanlimit 20</code> atau <code>/setscanlimit off</code>", parse_mode="HTML")
                    return "OK", 200

                lim_i = int(a1)
                if lim_i < 0:
                    lim_i = 0

                if arg2:
                    with engine.begin() as conn:
                        tgt_id = _resolve_target_user_id(conn, arg2)
                        if not tgt_id:
                            send_message(token, chat_id, "❌ Username tak jumpa dalam DB. Pastikan user pernah /start bot.", parse_mode="HTML")
                            return "OK", 200
                        conn.execute(text('''
                            INSERT INTO scan_limit_overrides (bot_id, user_id, limit_per_day)
                            VALUES (:b, :u, :l)
                            ON CONFLICT (bot_id, user_id) DO UPDATE SET
                              limit_per_day=excluded.limit_per_day,
                              updated_at=NOW()
                        '''), {"b": bot_id, "u": int(tgt_id), "l": int(lim_i)})
                    send_message(token, chat_id, f"✅ Set scan limit user <code>{tgt_id}</code>: <b>{lim_i}</b>/hari", parse_mode="HTML")
                else:
                    with engine.begin() as conn:
                        conn.execute(text("UPDATE bots SET scan_limit_per_day=:l WHERE id=:i"), {"l": int(lim_i), "i": bot_id})
                    send_message(token, chat_id, f"✅ Set scan limit GLOBAL: <b>{lim_i}</b>/hari", parse_mode="HTML")
                return "OK", 200

            elif text_msg.startswith("/setscanlimitmsg"):
                if not msg.get("reply_to_message"):
                    send_message(token, chat_id, "❌ Sila REPLY pada media/text yang nak dijadikan scan limit message.", parse_mode="HTML")
                    return "OK", 200
                logger.info(f"[CMD] /setscanlimitmsg uid={uid} chat={chat_id} text={text_msg}")
                # Custom message when user hits daily scan limit (supports text/media from reply)
                rep = msg["reply_to_message"]
                mt, mid, txt = save_content_from_reply(rep)
                raw = (txt or "").strip()
                with engine.begin() as conn:
                    conn.execute(text('''
                        UPDATE bots
                        SET scan_limit_message=:t,
                            scan_limit_message_media_type=:mt,
                            scan_limit_message_media_file_id=:mf
                        WHERE id=:i
                    '''), {"t": raw, "mt": mt, "mf": mid, "i": bot_id})
                send_message(token, chat_id, "✅ scan_limit message updated.", parse_mode="HTML")



            elif text_msg.startswith("/setshareamt"):
                # /setshareamt 1.00  (RM per 1 click share)
                parts = (text_msg or "").split()
                if len(parts) < 2:
                    send_message(token, chat_id, "Format: <code>/setshareamt 1.00</code>", parse_mode="HTML")
                else:
                    try:
                        amt = float(parts[1])
                        if amt <= 0:
                            raise ValueError("amt<=0")
                        with engine.begin() as conn:
                            conn.execute(text("UPDATE bots SET affiliate_amount=:a WHERE id=:i"), {"a": amt, "i": bot_id})
                        send_message(token, chat_id, f"✅ Share commission set: <b>RM{amt:.2f}</b> per 1 click.", parse_mode="HTML")
                    except Exception:
                        send_message(token, chat_id, "❌ Amount tak sah. Contoh: <code>/setshareamt 1.00</code>", parse_mode="HTML")

            elif text_msg.startswith("/setminwithdraw"):
                # /setminwithdraw 30.00  (min balance to request withdrawal)
                parts = (text_msg or "").split()
                if len(parts) < 2:
                    send_message(token, chat_id, "Format: <code>/setminwithdraw 30.00</code>", parse_mode="HTML")
                else:
                    try:
                        amt = float(parts[1])
                        if amt <= 0:
                            raise ValueError("amt<=0")
                        with engine.begin() as conn:
                            conn.execute(text("UPDATE bots SET min_withdraw_amount=:a WHERE id=:i"), {"a": amt, "i": bot_id})
                        send_message(token, chat_id, f"✅ Minimum withdraw set: <b>RM{amt:.2f}</b>", parse_mode="HTML")
                    except Exception:
                        send_message(token, chat_id, "❌ Amount tak sah. Contoh: <code>/setminwithdraw 30.00</code>", parse_mode="HTML")

            elif text_msg.startswith("/getrates"):
                b2 = get_bot_by_id(bot_id) or bot_row
                share_amt = get_bot_affiliate_amount(b2)
                min_wd = get_bot_min_withdraw(b2)
                send_message(
                    token,
                    chat_id,
                    f"⚙️ <b>BOT RATES</b>\n• 1 click share: <b>RM{share_amt:.2f}</b>\n• Min withdraw: <b>RM{min_wd:.2f}</b>",
                    parse_mode="HTML",
                )
            elif text_msg.startswith("/setlockbot"):
                val = "on" in text_msg.lower()
                with engine.begin() as conn:
                    conn.execute(text("UPDATE bots SET lock_bot=:v WHERE id=:i"), {"v": val, "i": bot_id})
                send_message(token, chat_id, f"🔒 PhoneLock: {val}", parse_mode="HTML")

            elif text_msg.startswith("/setadmingroup"):
                with engine.begin() as conn:
                    conn.execute(text("UPDATE bots SET admin_group_id=:g WHERE id=:i"), {"g": chat_id, "i": bot_id})
                send_message(token, chat_id, "✅ Group Admin Disimpan.", parse_mode="HTML")

            elif text_msg.startswith(("/setstart", "/setloading")):
                rep = msg.get("reply_to_message")
                if rep:
                    mt, mid, txt = save_content_from_reply(rep)
                    extra = "\n".join(text_msg.split("\n")[1:]).strip()
                    final_txt = (txt + ("\n" + extra if extra else "")).strip()

                    col_txt = "start_text" if text_msg.startswith("/setstart") else "loading_text"
                    col_type = "start_media_type" if text_msg.startswith("/setstart") else "loading_media_type"
                    col_file = "start_media_file_id" if text_msg.startswith("/setstart") else "loading_media_file_id"

                    with engine.begin() as conn:
                        conn.execute(
                            text(f"UPDATE bots SET {col_txt}=:t, {col_type}=:mt, {col_file}=:mf WHERE id=:i"),
                            {"t": final_txt, "mt": mt, "mf": mid, "i": bot_id},
                        )
                    send_message(token, chat_id, f"✅ {col_txt} Updated.", parse_mode="HTML")
                else:
                    cmd_name = "/setstart" if text_msg.startswith("/setstart") else "/setloading"
                    send_message(token, chat_id, f"❌ Sila <b>REPLY</b> kepada content (text/gambar/video) yang nak dijadikan {cmd_name} message.", parse_mode="HTML")

            elif text_msg.startswith(("/addscanner", "/setscannermedia")):
                if (not is_owner(uid, bot_row)) and (not is_admin(uid, bot_id)):
                    tg_send_message(token, chat_id, "❌ Command ini untuk OWNER/ADMIN sahaja.", parse_mode="HTML")
                    return jsonify({"ok": True})
                _is_set_cmd = text_msg.startswith("/setscannermedia")
                parts = text_msg.split(maxsplit=1)
                provider = norm_provider(parts[1] if len(parts) > 1 else "")
                if not provider:
                    _cmd_name = "/setscannermedia" if _is_set_cmd else "/addscanner"
                    tg_send_message(token, chat_id, f"❌ Format: {_cmd_name} <provider>\nContoh: {_cmd_name} jili", parse_mode="HTML")
                    return jsonify({"ok": True})
                if not msg.get("reply_to_message"):
                    tg_send_message(token, chat_id, "❌ Sila reply pada MEDIA (gambar/video/gif/document) yang nak dijadikan scanner.", parse_mode="HTML")
                    return jsonify({"ok": True})

                rmsg = msg["reply_to_message"]
                media_type = None
                file_id = None

                if rmsg.get("photo"):
                    media_type = "photo"
                    file_id = rmsg["photo"][-1].get("file_id")
                elif rmsg.get("video"):
                    media_type = "video"
                    file_id = rmsg["video"].get("file_id")
                elif rmsg.get("animation"):
                    media_type = "animation"
                    file_id = rmsg["animation"].get("file_id")
                elif rmsg.get("document"):
                    media_type = "document"
                    file_id = rmsg["document"].get("file_id")

                if not media_type or not file_id:
                    tg_send_message(token, chat_id, "❌ Media tu kena photo / video / gif (animation) / document.", parse_mode="HTML")
                    return jsonify({"ok": True})

                with engine.begin() as conn:
                    upsert_scanner_media(conn, bot_id, provider, media_type, file_id)

                if _is_set_cmd:
                    tg_send_message(token, chat_id, f"✅ Scanner media dikemaskini untuk <b>{html.escape(provider)}</b>.", parse_mode="HTML")
                else:
                    tg_send_message(token, chat_id, f"✅ Scanner media ditambah untuk <b>{html.escape(provider)}</b>.\n\nSeterusnya: /addgames {html.escape(provider)} (reply file txt).", parse_mode="HTML")
                return jsonify({"ok": True})

            elif text_msg.startswith("/addgames") or text_msg.startswith("/updategames"):
                if (not is_owner(uid, bot_row)) and (not is_admin(uid, bot_id)):
                    tg_send_message(token, chat_id, "❌ Command ini untuk OWNER/ADMIN sahaja.", parse_mode="HTML")
                    return jsonify({"ok": True})
                is_update = text_msg.startswith("/updategames")
                parts = text_msg.split(maxsplit=1)
                provider = norm_provider(parts[1] if len(parts) > 1 else "")
                if not provider:
                    tg_send_message(token, chat_id, "❌ Format: /addgames <provider> (reply file txt)\nContoh: /addgames jili", parse_mode="HTML")
                    return jsonify({"ok": True})

                raw = ""
                if msg.get("reply_to_message"):
                    rmsg = msg["reply_to_message"]
                    # Prefer document txt
                    if rmsg.get("document") and rmsg["document"].get("file_id"):
                        file_id = rmsg["document"]["file_id"]
                        try:
                            # download content using getFile
                            r = requests.get(f"https://api.telegram.org/bot{token}/getFile", params={"file_id": file_id}, timeout=20)
                            j = r.json()
                            file_path = j.get("result", {}).get("file_path")
                            if not file_path:
                                raise RuntimeError("file_path missing")
                            fr = requests.get(f"https://api.telegram.org/file/bot{token}/{file_path}", timeout=30)
                            fr.raise_for_status()
                            raw = fr.text
                        except Exception as e:
                            logger.exception("addgames download failed: %s", e)
                            tg_send_message(token, chat_id, "❌ Tak berjaya baca file txt. Pastikan file tu text/plain dan kecil (contoh bawah 1MB).", parse_mode="HTML")
                            return jsonify({"ok": True})
                    elif rmsg.get("text"):
                        raw = rmsg.get("text") or ""
                else:
                    # allow /addgames provider <paste list>
                    raw = parts[1] if len(parts) > 1 else ""

                games = parse_games_text(raw)
                if not games:
                    tg_send_message(token, chat_id, "❌ Games tiada. Sila reply file .txt (1 baris 1 game) atau reply text list game.", parse_mode="HTML")
                    return jsonify({"ok": True})

                with engine.begin() as conn:
                    # Both addgames and updategames now MERGE —
                    # insert new games without deleting existing ones.
                    # Use /clearscan <provider> to wipe a provider's list first if needed.
                    conn.execute(
                        text(
                            """
                            INSERT INTO scanner_games (bot_id, provider, game)
                            VALUES (:bot_id, :provider, :game)
                            ON CONFLICT DO NOTHING
                            """
                        ),
                        [{"bot_id": bot_id, "provider": provider, "game": g} for g in games],
                    )

                verb = "dikemaskini" if is_update else "ditambah"
                tg_send_message(token, chat_id, f"✅ List games <b>{html.escape(provider)}</b> {verb}: <b>{len(games)}</b> item.\n\nNota: Duplicate auto buang. Kalau kurang 20, bot akan paparkan semua.", parse_mode="HTML")
                return jsonify({"ok": True})



            
            elif text_msg.startswith("/clearscan"):
                # OWNER / ADMIN sahaja - padam list game provider dalam DB
                if (not is_owner(uid, bot_row)) and (not is_admin(uid, bot_id)):
                    tg_send_message(token, chat_id, "❌ Command ini untuk OWNER/ADMIN sahaja.", parse_mode="HTML")
                    return jsonify({"ok": True})

                parts = text_msg.split(maxsplit=1)
                arg = (parts[1].strip() if len(parts) > 1 else "")
                if not arg:
                    tg_send_message(
                        token,
                        chat_id,
                        "❌ Format:\n<code>/clearscan &lt;provider|all&gt;</code>\nContoh: <code>/clearscan jili</code> atau <code>/clearscan all</code>",
                        parse_mode="HTML",
                    )
                    return jsonify({"ok": True})

                arg_norm = norm_provider(arg)
                with engine.begin() as conn:
                    if arg_norm == "all":
                        res = conn.execute(text("DELETE FROM scanner_games WHERE bot_id=:b"), {"b": bot_id})
                        deleted = int(getattr(res, "rowcount", 0) or 0)
                        tg_send_message(token, chat_id, f"✅ Clear scan: semua provider dibuang. (<b>{deleted}</b> item)", parse_mode="HTML")
                    else:
                        res = conn.execute(
                            text("DELETE FROM scanner_games WHERE bot_id=:b AND provider=:p"),
                            {"b": bot_id, "p": arg_norm},
                        )
                        deleted = int(getattr(res, "rowcount", 0) or 0)
                        tg_send_message(token, chat_id, f"✅ Clear scan: provider <b>{html.escape(arg_norm)}</b> dibuang. (<b>{deleted}</b> item)", parse_mode="HTML")
                return jsonify({"ok": True})

            elif text_msg.startswith("/setcallback"):
                rep = msg.get("reply_to_message")
                parts = text_msg.split()
                if rep and len(parts) >= 2:
                    key = parts[1].strip()
                    delay = 0
                    if "delay=" in text_msg:
                        try:
                            delay = int(re.search(r"delay=(\d+)", text_msg).group(1))
                        except Exception:
                            delay = 0

                    mt, mid, txt = save_content_from_reply(rep)
                    extra = "\n".join(text_msg.split("\n")[1:]).strip()
                    final_txt = (txt + ("\n" + extra if extra else "")).strip()

                    actions_upsert(bot_id, key, mt or "text", final_txt, mid, delay)
                    send_message(token, chat_id, f"✅ Callback '{key}' Saved.", parse_mode="HTML")

            # NEW: /setcommand
            elif text_msg.startswith("/setcommand"):
                rep = msg.get("reply_to_message")
                parts = text_msg.split()
                if rep and len(parts) >= 2:
                    raw = parts[1].strip().lstrip("/")
                    cmd = re.sub(r"[^a-zA-Z0-9_]", "", raw).lower()
                    if not cmd:
                        send_message(token, chat_id, "❌ Command name invalid. Contoh: /setcommand hello", parse_mode="HTML")
                        return "OK", 200

                    delay = 0
                    if "delay=" in text_msg:
                        try:
                            delay = int(re.search(r"delay=(\d+)", text_msg).group(1))
                        except Exception:
                            delay = 0

                    mt, mid, txt = save_content_from_reply(rep)
                    extra = "\n".join(text_msg.split("\n")[1:]).strip()
                    final_txt = (txt + ("\n" + extra if extra else "")).strip()

                    key = f"cmd:{cmd}"
                    actions_upsert(bot_id, key, mt or "text", final_txt, mid, delay)
                    send_message(token, chat_id, f"✅ Command '/{cmd}' Saved.", parse_mode="HTML")
                else:
                    send_message(token, chat_id, "Cara: reply content + /setcommand hello", parse_mode="HTML")

            elif text_msg.startswith("/delcallback"):
                parts = text_msg.split()
                key = parts[1].strip() if len(parts) > 1 else ""
                if key:
                    ok = delete_callback(bot_id, key)
                    send_message(token, chat_id, f"{'🗑 Deleted' if ok else '⚠️ Not found'}: <code>{key}</code>", parse_mode="HTML")

            elif text_msg.startswith(("/approve", "/reject")):
                rep = msg.get("reply_to_message")
                if not rep:
                    send_message(token, chat_id, "⚠️ Reply pada mesej request (withdraw / contact) dulu baru guna /approve /reject.", parse_mode="HTML")
                    return "OK", 200

                rep_txt = rep.get("text") or rep.get("caption") or ""

                # Premium manual approval by UID
                uid_match = re.search(r"UID:\s*<code>(\d+)</code>", rep_txt)
                if uid_match and bot_row.get("manual_approval"):
                    target_uid = int(uid_match.group(1))
                    is_app = text_msg.startswith("/approve")
                    with engine.begin() as conn:
                        if is_app:
                            conn.execute(text("UPDATE users SET is_premium=TRUE, premium_until=NULL WHERE bot_id=:b AND user_id=:u"),
                                         {"b": bot_id, "u": target_uid})
                        else:
                            conn.execute(text("UPDATE users SET is_premium=FALSE, premium_until=NULL WHERE bot_id=:b AND user_id=:u"),
                                         {"b": bot_id, "u": target_uid})

                    if is_app:
                        msg_user = bot_row.get("verified_message") or (
                            "🎉 <b>PREMIUM AKTIF, BOSSKU!</b>\n"
                            "Akses kau dah unlock ✅\n"
                            "Sekarang boleh guna semua menu premium 🔥"
                        )
                        send_message(token, target_uid, msg_user, parse_mode="HTML")
                        send_message(token, chat_id, "✅ Premium Approved.", parse_mode="HTML")
                    else:
                        msg_user = bot_row.get("rejected_message") or (
                            "❌ <b>PREMIUM DITOLAK</b>\n"
                            "Bossku, admin tolak request. Kalau silap, boleh try semula."
                        )
                        send_message(token, target_uid, msg_user, parse_mode="HTML")
                        send_message(token, chat_id, "❌ Premium Rejected.", parse_mode="HTML")
                    return "OK", 200

                # Withdraw approval by ID
                rid_match = re.search(r"ID:\s*<code>([0-9a-fA-F-]+)</code>", rep_txt)
                if rid_match:
                    rid = rid_match.group(1)
                    is_app = text_msg.startswith("/approve")

                    with engine.begin() as conn:
                        wd = conn.execute(
                            text("SELECT * FROM withdrawals WHERE id=:i"),
                            {"i": rid},
                        ).mappings().first()

                        if not wd:
                            send_message(token, chat_id, "⚠️ Withdrawal ID tak jumpa.", parse_mode="HTML")
                            return "OK", 200

                        if wd["status"] != "PENDING":
                            send_message(token, chat_id, "⚠️ Withdrawal dah diproses sebelum ni.", parse_mode="HTML")
                            return "OK", 200

                        if is_app:
                            # /approve 50  (amount wajib)
                            parts = (text_msg or "").split()
                            if len(parts) < 2:
                                send_message(token, chat_id, "Format: <code>/approve 50</code>", parse_mode="HTML")
                                return "OK", 200
                            try:
                                amt = float(parts[1])
                            except Exception:
                                send_message(token, chat_id, "Format: <code>/approve 50</code>", parse_mode="HTML")
                                return "OK", 200

                            if amt <= 0:
                                send_message(token, chat_id, "❌ Amount tak sah.", parse_mode="HTML")
                                return "OK", 200

                            # Lock user row, check balance
                            u = conn.execute(
                                text("SELECT balance FROM users WHERE bot_id=:b AND user_id=:u FOR UPDATE"),
                                {"b": bot_id, "u": wd["user_id"]},
                            ).mappings().first()
                            bal_before = float((u or {}).get("balance") or 0)

                            if bal_before < amt:
                                send_message(
                                    token,
                                    chat_id,
                                    f"❌ Balance tak cukup untuk approve.\nBal user: RM{bal_before:.2f}\nApprove: RM{amt:.2f}",
                                    parse_mode="HTML",
                                )
                                return "OK", 200

                            # Deduct balance + mark approved
                            conn.execute(
                                text("UPDATE users SET balance=balance-:a WHERE bot_id=:b AND user_id=:u"),
                                {"a": amt, "b": bot_id, "u": wd["user_id"]},
                            )
                            conn.execute(
                                text("""
                                    UPDATE withdrawals
                                    SET status='APPROVED',
                                        approved_amount=:a,
                                        processed_at=NOW(),
                                        processed_by=:by
                                    WHERE id=:i
                                """),
                                {"a": amt, "i": rid, "by": uid},
                            )

                            bal_after = bal_before - amt
                            msg_user = (
                                "✅ <b>WITHDRAW BERJAYA</b>\n"
                                f"Jumlah: <b>RM{amt:.2f}</b>\n"
                                f"Baki sekarang: <b>RM{bal_after:.2f}</b>\n\n"
                                "Bossku, duit sedang diproses 😘"
                            )
                            send_message(token, int(wd["user_id"]), msg_user, parse_mode="HTML")
                            send_message(token, chat_id, f"✅ Withdraw Approved. (Baki user: RM{bal_after:.2f})", parse_mode="HTML")
                            return "OK", 200

                        # Reject
                        conn.execute(
                            text("""
                                UPDATE withdrawals
                                SET status='REJECTED',
                                    processed_at=NOW(),
                                    processed_by=:by
                                WHERE id=:i
                            """),
                            {"i": rid, "by": uid},
                        )
                        send_message(
                            token,
                            int(wd["user_id"]),
                            "❌ <b>WITHDRAW DITOLAK</b>\nRequest ditolak. Sila semak detail & try lagi.",
                            parse_mode="HTML",
                        )
                        send_message(token, chat_id, "❌ Withdraw Rejected.", parse_mode="HTML")
                    return "OK", 200
# Dynamic command triggers (after builtins)
        if text_msg.startswith("/"):
            cmd = parse_command_name(text_msg)
            if cmd:
                act = actions_get(bot_id, f"cmd:{cmd}")
                if act:
                    if int(act.get("delay_seconds") or 0) > 0:
                        time.sleep(int(act["delay_seconds"]))

                    urow = get_user_row(bot_id, uid) or {"user_id": uid, "first_name": from_user.get("first_name"), "username": from_user.get("username")}
                    if not ensure_access(bot_row, chat_id, uid, urow):
                        return "OK", 200

                    txt = render_placeholders(act.get("text") or "", bot_row.get("bot_username") or "", urow)

                    # scan placeholders ({count}/{limit}/{remaining}/{reset})

                    with engine.connect() as _c:

                        txt = apply_scan_placeholders(_c, txt, bot_row, bot_id, int((urow or {}).get("user_id") or uid))
                    share_q = make_share_query(bot_row.get("bot_username") or "", urow)
                    txt, markup = parse_buttons(txt, share_inline_query=share_q)

                    if act["type"] != "text" and act.get("media_file_id"):
                        send_media(token, chat_id, act["type"], act["media_file_id"], caption=txt, reply_markup=markup)
                    else:
                        send_message(token, chat_id, txt, reply_markup=markup)

        # Catch-all: if user in private chat hasn't verified contact, re-prompt
        chat_type = msg.get("chat", {}).get("type", "private")
        if chat_type == "private" and not require_admin(bot_row, uid):
            urow = get_user_row(bot_id, uid)
            if bot_row.get("lock_bot") and (not urow or not urow.get("is_verified")):
                ensure_contact_verified(bot_row, chat_id, urow)
                return "OK", 200

        return "OK", 200

    # callback_query
    if "callback_query" in update:
        cq = update["callback_query"]
        msg = cq.get("message") or {}
        chat_id = msg.get("chat", {}).get("id")
        message_id = msg.get("message_id")
        data = cq.get("data", "")
        from_user = cq.get("from") or {}
        uid = from_user.get("id")

        if not chat_id or not uid:
            return "OK", 200

        user_row, _ = upsert_user(bot_id, from_user, None)

        # Gate recheck for joinlock
        if data == "gate:recheck":
            answer_callback(token, cq["id"], "Checking…", show_alert=False)
            bot_row2 = get_bot_by_id(bot_id) or bot_row
            if ensure_access(bot_row2, chat_id, uid, user_row):
                send_joined_message(bot_row2, chat_id, user_row)
            return "OK", 200

        # Admin approve buttons (manual approval)
        elif data.startswith("adm:"):
            if not require_admin(bot_row, uid):
                answer_callback(token, cq["id"], "No access", show_alert=True)
                return "OK", 200

            parts = data.split(":")
            action = parts[1] if len(parts) > 1 else ""

            # Settings categories navigation
            if action == "cat":
                cat = parts[2] if len(parts) > 2 else "home"
                bot_row2 = get_bot_by_id(bot_id) or bot_row
                send_or_edit_settings_panel(bot_row2, chat_id, uid, page=1, edit_ctx={"message_id": message_id}, cat=cat)
                answer_callback(token, cq["id"])
            # (scan limit gate removed from adm:* handlers; handled in cb:* scanner path)

            if action == "home":
                bot_row2 = get_bot_by_id(bot_id) or bot_row
                send_or_edit_settings_panel(bot_row2, chat_id, uid, page=1, edit_ctx={"message_id": message_id}, cat="home")
                answer_callback(token, cq["id"])
                return "OK", 200

            if action == "full":
                p = int(parts[2]) if len(parts) > 2 and parts[2].isdigit() else 1
                bot_row2 = get_bot_by_id(bot_id) or bot_row
                send_or_edit_settings_panel(bot_row2, chat_id, uid, page=p, edit_ctx={"message_id": message_id})
                answer_callback(token, cq["id"])
                return "OK", 200

            target_uid = int(parts[2]) if len(parts) > 2 and parts[2].isdigit() else 0
            if not target_uid:
                answer_callback(token, cq["id"], "Invalid target", show_alert=True)
                return "OK", 200

            # Protect: admin/owner tak patut jadi target premium manual approve
            if require_admin(bot_row, int(target_uid)):
                answer_callback(token, cq["id"], "Target ialah admin/owner (skip).", show_alert=True)
                return "OK", 200

            if action == "ap":
                with engine.begin() as conn:
                    conn.execute(
                        text("UPDATE users SET is_premium=TRUE, premium_until=NULL WHERE bot_id=:b AND user_id=:u"),
                        {"b": bot_id, "u": target_uid},
                    )

                msg_user = bot_row.get("verified_message") or (
                    "🎉 <b>PREMIUM AKTIF, BOSSKU!</b>\n"
                    "Akses kau dah unlock ✅\n"
                    "Sekarang boleh guna semua menu premium 🔥"
                )
                _vmt = bot_row.get("verified_message_media_type")
                _vmf = bot_row.get("verified_message_media_file_id")
                if _vmt and _vmf:
                    send_media(token, target_uid, _vmt, _vmf, caption=msg_user, parse_mode="HTML")
                else:
                    send_message(token, target_uid, msg_user, parse_mode="HTML")

                # Update mesej admin dengan nama approver + lock button
                try:
                    _m = cq.get("message") or {}
                    _chat_id = (_m.get("chat") or {}).get("id")
                    _msg_id = _m.get("message_id")
                    if _chat_id and _msg_id:
                        _from = cq.get("from") or {}
                        _admin_id = _from.get("id")
                        _admin_name = _from.get("first_name") or "Admin"
                        _admin_user = _from.get("username")
                        _who = f"@{_admin_user}" if _admin_user else f"<a href='tg://user?id={_admin_id}'>{html.escape(_admin_name)}</a>"
                        _stamp = now_local_str("%Y-%m-%d %H:%M:%S")
                        
                        # Check if message is media (has caption) or text
                        _is_media = _m.get("photo") or _m.get("video") or _m.get("animation") or _m.get("document")
                        _cur = _m.get("caption") if _is_media else _m.get("text") or ""
                        
                        if ("<b>APPROVED</b>" not in _cur) and ("<b>REJECTED</b>" not in _cur):
                            _cur = _cur + f"\n\n✅ <b>APPROVED</b>\nBy: {_who}\nAt: {_stamp}"
                        
                        # Use appropriate edit method based on message type
                        if _is_media:
                            edit_caption(token, _chat_id, _msg_id, _cur, reply_markup={"inline_keyboard": []}, parse_mode="HTML")
                        else:
                            edit_message(token, _chat_id, _msg_id, _cur, reply_markup={"inline_keyboard": []}, parse_mode="HTML")
                except Exception as e:
                    logger.exception(f"Failed to update admin premium message (approve): {e}")

                answer_callback(token, cq["id"], "Approved ✅", show_alert=False)

            elif action == "rj":
                with engine.begin() as conn:
                    conn.execute(
                        text("UPDATE users SET is_premium=FALSE, premium_until=NULL WHERE bot_id=:b AND user_id=:u"),
                        {"b": bot_id, "u": target_uid},
                    )

                msg_user = bot_row.get("rejected_message") or (
                    "❌ <b>PREMIUM DITOLAK</b>\n"
                    "Bossku, admin tolak request. Kalau silap, boleh try semula."
                )
                _rmt = bot_row.get("rejected_message_media_type")
                _rmf = bot_row.get("rejected_message_media_file_id")
                if _rmt and _rmf:
                    send_media(token, target_uid, _rmt, _rmf, caption=msg_user, parse_mode="HTML")
                else:
                    send_message(token, target_uid, msg_user, parse_mode="HTML")

                # Update mesej admin dengan nama rej actor + lock button
                try:
                    _m = cq.get("message") or {}
                    _chat_id = (_m.get("chat") or {}).get("id")
                    _msg_id = _m.get("message_id")
                    if _chat_id and _msg_id:
                        _from = cq.get("from") or {}
                        _admin_id = _from.get("id")
                        _admin_name = _from.get("first_name") or "Admin"
                        _admin_user = _from.get("username")
                        _who = f"@{_admin_user}" if _admin_user else f"<a href='tg://user?id={_admin_id}'>{html.escape(_admin_name)}</a>"
                        _stamp = now_local_str("%Y-%m-%d %H:%M:%S")
                        
                        # Check if message is media (has caption) or text
                        _is_media = _m.get("photo") or _m.get("video") or _m.get("animation") or _m.get("document")
                        _cur = _m.get("caption") if _is_media else _m.get("text") or ""
                        
                        if ("<b>APPROVED</b>" not in _cur) and ("<b>REJECTED</b>" not in _cur):
                            _cur = _cur + f"\n\n❌ <b>REJECTED</b>\nBy: {_who}\nAt: {_stamp}"
                        
                        # Use appropriate edit method based on message type
                        if _is_media:
                            edit_caption(token, _chat_id, _msg_id, _cur, reply_markup={"inline_keyboard": []}, parse_mode="HTML")
                        else:
                            edit_message(token, _chat_id, _msg_id, _cur, reply_markup={"inline_keyboard": []}, parse_mode="HTML")
                except Exception as e:
                    logger.exception(f"Failed to update admin premium message (reject): {e}")

                answer_callback(token, cq["id"], "Rejected ❌", show_alert=False)

            return "OK", 200


        # Withdrawal approve/reject buttons
        elif data.startswith("wd:"):
            if not require_admin(bot_row, uid):
                answer_callback(token, cq["id"], "No access", show_alert=True)
                return "OK", 200

            # data: wd:ap:<uuid> OR wd:rj:<uuid>
            parts = data.split(":", 2)
            action = parts[1] if len(parts) > 1 else ""
            wid = parts[2] if len(parts) > 2 else ""

            if not re.match(r"^[0-9a-fA-F-]{20,}$", wid):
                answer_callback(token, cq["id"], "Invalid ID", show_alert=True)
                return "OK", 200

            with engine.begin() as conn:
                wd = conn.execute(
                    text("SELECT * FROM withdrawals WHERE id=:i FOR UPDATE"),
                    {"i": wid},
                ).mappings().first()

                if not wd:
                    answer_callback(token, cq["id"], "WD not found", show_alert=True)
                    return "OK", 200

                if wd["status"] != "PENDING":
                    answer_callback(token, cq["id"], f"Already {wd['status']}", show_alert=True)
                    return "OK", 200

                u = conn.execute(
                    text("SELECT balance FROM users WHERE bot_id=:b AND user_id=:u FOR UPDATE"),
                    {"b": bot_id, "u": wd["user_id"]},
                ).mappings().first()
                bal_before = float((u or {}).get("balance") or 0)

                if action == "ap":
                    # Use amount locked at request time
                    amt = float(wd.get("request_amount") or 0)
                    if amt <= 0:
                        amt = bal_before  # fallback for old records without request_amount
                    if amt <= 0:
                        answer_callback(token, cq["id"], "Amount tak sah.", show_alert=True)
                        return "OK", 200
                    if bal_before < amt:
                        answer_callback(token, cq["id"], f"Balance tak cukup (RM{bal_before:.2f}).", show_alert=True)
                        return "OK", 200

                    conn.execute(
                        text("UPDATE users SET balance=balance-:a WHERE bot_id=:b AND user_id=:u"),
                        {"a": amt, "b": bot_id, "u": wd["user_id"]},
                    )
                    conn.execute(
                        text("""
                            UPDATE withdrawals
                            SET status='APPROVED',
                                approved_amount=:a,
                                processed_at=NOW(),
                                processed_by=:by
                            WHERE id=:i
                        """),
                        {"a": amt, "i": wid, "by": uid},
                    )

                    bal_after = bal_before - amt
                    bot_latest = get_bot_by_id(bot_id) or bot_row
                    tpl = bot_latest.get("withdrawal_approve_message") or (
                        "✅ <b>WITHDRAW BERJAYA</b>\n"
                        "Jumlah: <b>{amount}</b>\n"
                        "Baki sekarang: <b>{balance_after}</b>\n\n"
                        "Bossku, duit sedang diproses 😘"
                    )
                    msg_user = render_withdrawal_template(tpl, amt, bal_before, bal_after)

                    mt = bot_latest.get("withdrawal_approve_media_type")
                    mf = bot_latest.get("withdrawal_approve_media_file_id")
                    if mt and mf:
                        send_media(token, int(wd["user_id"]), mt, mf, caption=msg_user, parse_mode="HTML")
                    else:
                        send_message(token, int(wd["user_id"]), msg_user, parse_mode="HTML")

                    # Update the admin/group message so you can SEE it was approved (and lock the buttons)
                    try:
                        _m = cq.get("message") or {}
                        _chat_id = (_m.get("chat") or {}).get("id")
                        _msg_id = _m.get("message_id")
                        if _chat_id and _msg_id:
                            _from = cq.get("from") or {}
                            _admin_id = _from.get("id")
                            _admin_name = _from.get("first_name") or "Admin"
                            _admin_user = _from.get("username")
                            _who = f"@{_admin_user}" if _admin_user else f"<a href='tg://user?id={_admin_id}'>{html.escape(_admin_name)}</a>"
                            _stamp = now_local_str("%Y-%m-%d %H:%M:%S")
                            _cur = _m.get("text") or ""
                            if ("<b>APPROVED</b>" not in _cur) and ("<b>REJECTED</b>" not in _cur):
                                _cur = _cur + f"\n\n✅ <b>APPROVED</b>\nBy: {_who}\nAt: {_stamp}"
                            # remove inline keyboard (avoid double approve/reject)
                            edit_message(token, _chat_id, _msg_id, _cur, reply_markup={"inline_keyboard": []}, parse_mode="HTML")
                    except Exception:
                        logger.exception("Failed to update admin withdrawal message (approve)")
                    answer_callback(token, cq["id"], "Approved ✅", show_alert=False)

                elif action == "rj":
                    conn.execute(
                        text("""
                            UPDATE withdrawals
                            SET status='REJECTED',
                                processed_at=NOW(),
                                processed_by=:by
                            WHERE id=:i
                        """),
                        {"i": wid, "by": uid},
                    )

                    bot_latest = get_bot_by_id(bot_id) or bot_row
                    tpl = bot_latest.get("withdrawal_reject_message") or (
                        "❌ <b>WITHDRAW DITOLAK</b>\n"
                        "Request ditolak. Sila semak detail & cuba lagi."
                    )
                    msg_user = render_withdrawal_template(tpl, 0.0, bal_before, bal_before)

                    mt = bot_latest.get("withdrawal_reject_media_type")
                    mf = bot_latest.get("withdrawal_reject_media_file_id")
                    if mt and mf:
                        send_media(token, int(wd["user_id"]), mt, mf, caption=msg_user, parse_mode="HTML")
                    else:
                        send_message(token, int(wd["user_id"]), msg_user, parse_mode="HTML")

                    # Update the admin/group message so you can SEE it was rejected (and lock the buttons)
                    try:
                        _m = cq.get("message") or {}
                        _chat_id = (_m.get("chat") or {}).get("id")
                        _msg_id = _m.get("message_id")
                        if _chat_id and _msg_id:
                            _from = cq.get("from") or {}
                            _admin_id = _from.get("id")
                            _admin_name = _from.get("first_name") or "Admin"
                            _admin_user = _from.get("username")
                            _who = f"@{_admin_user}" if _admin_user else f"<a href='tg://user?id={_admin_id}'>{html.escape(_admin_name)}</a>"
                            _stamp = now_local_str("%Y-%m-%d %H:%M:%S")
                            _cur = _m.get("text") or ""
                            if ("<b>APPROVED</b>" not in _cur) and ("<b>REJECTED</b>" not in _cur):
                                _cur = _cur + f"\n\n❌ <b>REJECTED</b>\nBy: {_who}\nAt: {_stamp}"
                            edit_message(token, _chat_id, _msg_id, _cur, reply_markup={"inline_keyboard": []}, parse_mode="HTML")
                    except Exception:
                        logger.exception("Failed to update admin withdrawal message (reject)")
                    answer_callback(token, cq["id"], "Rejected ❌", show_alert=False)

                else:
                    answer_callback(token, cq["id"], "Unknown action", show_alert=True)

            return "OK", 200

        # your existing cb:
        elif data.startswith("cb:"):
            # Callback actions stored in DB (premium mode: edit in-place)
            raw = data.split(":", 1)[1].strip()
            delay_override = None
            if ";d=" in raw:
                key_part, d_part = raw.split(";d=", 1)
                raw = key_part.strip()
                mdel = re.match(r"^(\d+)", (d_part or "").strip())
                if mdel:
                    try:
                        delay_override = int(mdel.group(1))
                    except Exception:
                        delay_override = None
            key = raw.strip()
            # normalize scan_* keys -> provider (e.g., scan_jili -> jili)
            if key.startswith("scan_"):
                key = key.split("scan_", 1)[1]

            # ===== DAILY SCAN LIMIT GATE (read-only check, no increment) =====
            try:
                with engine.connect() as _conn_gate:
                    _media_gate = None
                    _games_gate = None
                    try:
                        _media_gate = get_scanner_media(_conn_gate, bot_id, key)
                    except Exception:
                        _media_gate = None
                    try:
                        _games_gate = get_scanner_games(_conn_gate, bot_id, key) or []
                    except Exception:
                        _games_gate = []
                    _is_scanner = bool(_media_gate) or bool(_games_gate)

                    if _is_scanner:
                        # fetch latest bot config (in case just updated)
                        _bot_latest = get_bot_by_id(bot_id) or bot_row
                        # READ-ONLY check: do NOT increment here, only check current usage
                        _lim = get_scan_limit_for_user(_conn_gate, _bot_latest, bot_id, int(uid))
                        if _lim is not None:
                            try:
                                _lim_int = int(_lim)
                            except Exception:
                                _lim_int = 0
                            if _lim_int > 0:
                                _used_today, _, _, _ = scan_daily_get_stats(_conn_gate, _bot_latest, bot_id, int(uid))
                                if _used_today >= _lim_int:
                                    # BLOCKED: over daily limit
                                    _urow_gate = user_row or get_user_row(bot_id, uid) or {"user_id": uid, "first_name": (from_user.get("first_name") or "")}
                                    _tpl = (_bot_latest.get("scan_limit_message") or "").strip() or "🚫 Had scan harian anda telah habis.\nLimit: {limit}/hari\nCuba semula esok."
                                    _msg = render_placeholders(_tpl, _bot_latest.get("bot_username") or "", _urow_gate)
                                    _msg = apply_scan_placeholders(_conn_gate, _msg, _bot_latest, bot_id, int(uid))
                                    _remaining = max(0, _lim_int - _used_today)

                                    _msg = (_msg
                                            .replace("{limit}", str(_lim_int))
                                            .replace("{used}", str(_used_today))
                                            .replace("{remaining}", str(_remaining)))

                                    # Parse buttons from scan limit message (supports !1link, !1web, etc)
                                    _msg, _parsed_kb = parse_buttons(_msg)
                                    _kb_lim_rows = [[{"text": "⬅️ Kembali", "callback_data": "cb:menuscanner"}]]
                                    if _parsed_kb and _parsed_kb.get("inline_keyboard"):
                                        _kb_lim_rows = _parsed_kb["inline_keyboard"] + _kb_lim_rows
                                    _kb_lim = {"inline_keyboard": _kb_lim_rows}

                                    # always show alert + send new message (do NOT edit old media message)
                                    # Strip HTML for popup (answerCallbackQuery = plain text only)
                                    _popup_txt = re.sub(r'<[^>]+>', '', _msg)[:180] if _msg else "Limit harian habis"
                                    try:
                                        answer_callback(token, cq["id"], _popup_txt, show_alert=True)
                                    except Exception:
                                        try:
                                            answer_callback(token, cq["id"], "Limit harian habis", show_alert=True)
                                        except Exception:
                                            pass
                                    try:
                                        send_message(token, chat_id, _msg or "Limit harian habis.", reply_markup=_kb_lim, parse_mode="HTML")
                                    except Exception:
                                        # fallback without parse mode
                                        send_message(token, chat_id, _msg or "Limit harian habis.", reply_markup=_kb_lim)
                                    return "OK", 200
            except Exception as _gate_err:
                logger.error("SCAN LIMIT GATE error (bot=%s user=%s key=%s): %s", bot_id, uid, key, _gate_err, exc_info=True)
            # ===== END DAILY SCAN LIMIT GATE =====

            act = actions_get(bot_id, key)

            # answer quickly to stop Telegram spinner
            answer_callback(token, cq["id"])

            if not act:
                # Scanner fallback: if key matches a provider that has scanner media + games, run scanner.
                try:
                    with engine.begin() as conn:
                        media = get_scanner_media(conn, bot_id, key)
                        games = get_scanner_games(conn, bot_id, key) if media else []
                        if media and games:
                            # gate + cooldown
                            urow_gate = user_row or get_user_row(bot_id, uid) or {"user_id": uid}
                            if not ensure_access(bot_row, chat_id, uid, urow_gate):
                                return "OK", 200
                            # daily scan limit (per day)
                            allowed, used_after, lim = scan_daily_touch_or_block(conn, bot_row, bot_id, int(uid))
                            if not allowed:
                                bot_latest = get_bot_by_id(bot_id) or bot_row
                                urow_lim = urow_gate
                                tpl = bot_latest.get("scan_limit_message") or (
                                    "❌ <b>LIMIT SCAN HARI INI HABIS</b>\\n"
                                    "Anda dah capai limit scan untuk hari ini. Cuba lagi esok."
                                )
                                txt_lim = render_placeholders(tpl, bot_latest.get("bot_username") or "", urow_lim)
                                if lim is not None and lim > 0:
                                    txt_lim = (txt_lim + f"\\n\\n📌 Used: <b>{used_after}/{lim}</b>").strip()
                                # Parse buttons from scan limit message (supports !1link, !1web, etc)
                                txt_lim, _parsed_kb2 = parse_buttons(txt_lim)
                                _kb_lim2_rows = [[{"text": "⬅️ Kembali", "callback_data": "cb:menuscanner"}]]
                                if _parsed_kb2 and _parsed_kb2.get("inline_keyboard"):
                                    _kb_lim2_rows = _parsed_kb2["inline_keyboard"] + _kb_lim2_rows
                                kb_lim = {"inline_keyboard": _kb_lim2_rows}

                                mt_lim = bot_latest.get("scan_limit_message_media_type")
                                mf_lim = bot_latest.get("scan_limit_message_media_file_id")
                                try:
                                    if mt_lim and mf_lim:
                                        # try edit as media (if current message is media); else send as new media
                                        try:
                                            edit_media(token, chat_id, message_id, mt_lim, mf_lim, caption=txt_lim, reply_markup=kb_lim, parse_mode="HTML")
                                        except Exception:
                                            send_media(token, chat_id, mt_lim, mf_lim, caption=txt_lim, reply_markup=kb_lim, parse_mode="HTML")
                                    else:
                                        edit_message(token, chat_id, message_id, txt_lim or " ", reply_markup=kb_lim, parse_mode="HTML")
                                except Exception:
                                    send_message(token, chat_id, txt_lim or " ", reply_markup=kb_lim, parse_mode="HTML")
                                return "OK", 200


                            firstname = (from_user.get("first_name") or "").strip()
                            _mid = str((urow_gate or {}).get("member_id") or "")
                            bot_latest = get_bot_by_id(bot_id) or bot_row
                            run_scanner_flow(
                                token,
                                chat_id,
                                message_id,
                                firstname,
                                key,
                                media,
                                games,
                                member_id=_mid,
                                cache_key=(bot_id, uid, key),
                                bot_row=bot_latest,
                            )
                            return "OK", 200
                except Exception as e:
                    logger.exception("scanner fallback error: %s", e)
                return "OK", 200

            # Ensure user passes gate for any callback
            if not ensure_access(bot_row, chat_id, uid, user_row):
                return "OK", 200

            delay = int(delay_override if delay_override is not None else (act.get("delay_seconds") or 0))

            # If no message_id to edit, fallback to sending new message
            if not message_id:
                urow = get_user_row(bot_id, uid) or {"user_id": uid}
                txt = render_placeholders(act.get("text") or "", bot_row.get("bot_username") or "", urow)
                # scan placeholders ({count}/{limit}/{remaining}/{reset})
                with engine.connect() as _c:
                    txt = apply_scan_placeholders(_c, txt, bot_row, bot_id, int((urow or {}).get("user_id") or uid))
                share_q = make_share_query(bot_row.get("bot_username") or "", urow)
                txt, markup = parse_buttons(txt, share_inline_query=share_q)
                send_message(token, chat_id, txt or " ", reply_markup=markup, parse_mode="HTML")
                return "OK", 200

            # No delay -> jump directly to callback result (fast chaining)
            if delay <= 0:
                urow = get_user_row(bot_id, uid) or {"user_id": uid}
                txt = render_placeholders(act.get("text") or "", bot_row.get("bot_username") or "", urow)
                # scan placeholders ({count}/{limit}/{remaining}/{reset})
                with engine.connect() as _c:
                    txt = apply_scan_placeholders(_c, txt, bot_row, bot_id, int((urow or {}).get("user_id") or uid))
                share_q = make_share_query(bot_row.get("bot_username") or "", urow)
                txt, markup = parse_buttons(txt, share_inline_query=share_q)

                if act["type"] != "text" and act.get("media_file_id"):
                    # Prefer media edit when action includes media
                    try:
                        edit_media(token, chat_id, message_id, act["type"], act["media_file_id"], caption=txt, reply_markup=markup, parse_mode="HTML")
                    except Exception:
                        # fallback: delete + resend
                        delete_message(token, chat_id, message_id)
                        send_message(token, chat_id, txt or " ", reply_markup=markup, parse_mode="HTML")
                else:
                    ok = edit_message(token, chat_id, message_id, txt or " ", reply_markup=markup, parse_mode="HTML")
                    if ok is None:
                        # Current message is media — can't editMessageText
                        delete_message(token, chat_id, message_id)
                        send_message(token, chat_id, txt or " ", reply_markup=markup, parse_mode="HTML")
                return "OK", 200

            # Delay > 0 -> cinematic LOADING then edit back to result using Cloud Tasks
            urow = get_user_row(bot_id, uid) or {
                "user_id": uid,
                "first_name": from_user.get("first_name") or "",
                "username": from_user.get("username"),
                "balance": 0,
                "shared_count": 0,
                "member_id": "000000",
            }
            # Step 1: edit current message into LOADING
            edit_loading_message(bot_row, chat_id, message_id, urow)

            # Step 2: queue delayed action
            if can_use_tasks_action():
                enqueue_action_task(
                    {"bot_id": bot_id, "chat_id": chat_id, "user_id": uid, "message_id": message_id, "key": key},
                    delay_seconds=delay,
                )
            else:
                # Fallback (less reliable on Cloud Run): background timer
                import threading

                def _later():
                    act2 = actions_get(bot_id, key)
                    if not act2:
                        return
                    u2 = get_user_row(bot_id, uid) or {"user_id": uid}
                    if not ensure_access(bot_row, chat_id, uid, u2):
                        return
                    t2 = render_placeholders(act2.get("text") or "", bot_row.get("bot_username") or "", u2)
                    q2 = make_share_query(bot_row.get("bot_username") or "", u2)
                    t2, mk2 = parse_buttons(t2, share_inline_query=q2)
                    try:
                        if act2["type"] != "text" and act2.get("media_file_id"):
                            edit_media(token, chat_id, message_id, act2["type"], act2["media_file_id"], caption=t2, reply_markup=mk2, parse_mode="HTML")
                        else:
                            ok = edit_message(token, chat_id, message_id, t2 or " ", reply_markup=mk2, parse_mode="HTML")
                            if ok is None:
                                delete_message(token, chat_id, message_id)
                                send_message(token, chat_id, t2 or " ", reply_markup=mk2, parse_mode="HTML")
                    except Exception:
                        send_message(token, chat_id, t2 or " ", reply_markup=mk2, parse_mode="HTML")

                threading.Timer(delay, _later).start()

            return "OK", 200
        elif data == "req_withdraw":
            # Block withdrawal request if balance not enough (show popup alert)
            bot_id_ = str(bot_row["id"])
            uid_int = int(from_user.get("id") or 0)
            min_wd = get_bot_min_withdraw(bot_row)
            with engine.connect() as conn:
                urow0 = conn.execute(
                    text("SELECT balance FROM users WHERE bot_id=:b AND user_id=:u"),
                    {"b": bot_id_, "u": uid_int},
                ).mappings().first()
            bal0 = float((urow0 or {}).get("balance") or 0)

            if bal0 < float(min_wd):
                _fail_msg = build_withdraw_insufficient_msg(float(min_wd), float(bal0), bot_row)
                _fail_popup = re.sub(r'<[^>]+>', '', _fail_msg)[:180]
                answer_callback(token, cq["id"], _fail_popup, show_alert=True)
                return "OK", 200

            handle_withdraw_request(bot_row, chat_id, from_user)
            answer_callback(token, cq["id"])
            return "OK", 200

        # SETTINGS UI
        elif data.startswith("st:"):
            if data == "st:noop":
                answer_callback(token, cq["id"])
                return "OK", 200

            if not require_admin(bot_row, uid):
                answer_callback(token, cq["id"], "No access", show_alert=True)
                return "OK", 200

            parts = data.split(":")
            action = parts[1] if len(parts) > 1 else ""

            # Category navigation: st:cat:content, st:cat:security, etc.
            if action == "cat":
                cat_name = parts[2] if len(parts) > 2 else ""
                answer_callback(token, cq["id"])
                bot_row2 = get_bot_by_id(bot_id) or bot_row
                send_or_edit_settings_panel(bot_row2, chat_id, uid, page=1, edit_ctx={"message_id": message_id}, cat=cat_name)
                return "OK", 200

            if action in ("lock", "join", "manual", "inplace"):
                val = (parts[2] == "on") if len(parts) > 2 else False
                if action == "join" and val:
                    bot_row_latest = get_bot_by_id(bot_id) or bot_row
                    if not parse_join_targets(bot_row_latest.get("join_targets")):
                        answer_callback(token, cq["id"], "Set join targets dulu guna /setjoin @channel", show_alert=True)
                        send_message(token, chat_id, "⚠️ <b>JoinLock perlukan join target</b>\n\nSet dulu contoh:\n<code>/setjoin @channelanda</code>\natau\n<code>/setjoin @channel1,@channel2</code>", parse_mode="HTML")
                        return "OK", 200
                col = {
                    "lock": "lock_bot",
                    "join": "join_lock",
                    "manual": "manual_approval",
                    "inplace": "inplace_callbacks",
                }[action]
                with engine.begin() as conn:
                    conn.execute(text(f"UPDATE bots SET {col}=:v WHERE id=:i"), {"v": val, "i": bot_id})
                answer_callback(token, cq["id"], f"{col} set: {val}")
                bot_row2 = get_bot_by_id(bot_id) or bot_row
                send_or_edit_settings_panel(bot_row2, chat_id, uid, page=1, edit_ctx={"message_id": message_id})
                return "OK", 200

            if action == "livegram":
                sub = parts[2] if len(parts) > 2 else ""
                if sub == "scope":
                    new_scope = parts[3] if len(parts) > 3 else "private"
                    with engine.begin() as conn:
                        conn.execute(text("UPDATE bots SET livegram_scope=:v WHERE id=:i"), {"v": new_scope, "i": bot_id})
                    answer_callback(token, cq["id"], f"Livegram scope: {new_scope}")
                else:
                    val = (sub == "on")
                    with engine.begin() as conn:
                        conn.execute(text("UPDATE bots SET livegram=:v WHERE id=:i"), {"v": val, "i": bot_id})
                    answer_callback(token, cq["id"], f"Livegram {'ON ✅' if val else 'OFF ❌'}")
                bot_row2 = get_bot_by_id(bot_id) or bot_row
                send_or_edit_settings_panel(bot_row2, chat_id, uid, page=1, edit_ctx={"message_id": message_id}, cat="utils")
                return "OK", 200

            if action == "admingroup":
                if int(chat_id) >= 0:
                    answer_callback(token, cq["id"], "Tekan button ni dalam GROUP (bukan PM).", show_alert=True)
                    return "OK", 200
                with engine.begin() as conn:
                    conn.execute(text("UPDATE bots SET admin_group_id=:g WHERE id=:i"), {"g": chat_id, "i": bot_id})
                answer_callback(token, cq["id"], "Admin group saved ✅")
                bot_row2 = get_bot_by_id(bot_id) or bot_row
                send_or_edit_settings_panel(bot_row2, chat_id, uid, page=1, edit_ctx={"message_id": message_id})
                return "OK", 200

            if action == "financial":
                sub = parts[2] if len(parts) > 2 else ""
                
                if sub == "viewrates":
                    bot_row2 = get_bot_by_id(bot_id) or bot_row
                    share_amt = get_bot_affiliate_amount(bot_row2)
                    min_wd = get_bot_min_withdraw(bot_row2)
                    msg = (
                        "💰 <b>FINANCIAL RATES</b>\n"
                        "━━━━━━━━━━━━━━━━━━\n"
                        f"• 1 click share: <b>RM{share_amt:.2f}</b>\n"
                        f"• Min withdraw: <b>RM{min_wd:.2f}</b>\n"
                        "\n"
                        "Use buttons to edit 📝"
                    )
                    send_message(token, chat_id, msg, parse_mode="HTML")
                    answer_callback(token, cq["id"])
                    return "OK", 200
                
                elif sub == "setshare":
                    pending_inputs[(bot_id, uid)] = ("setshare", time.time())
                    msg = (
                        "✏️ <b>Edit Share Amount</b>\n\n"
                        "Reply with amount per click (RM)\n"
                        "Contoh: <code>1.00</code> or <code>0.50</code>"
                    )
                    send_message(token, chat_id, msg, parse_mode="HTML")
                    answer_callback(token, cq["id"])
                    return "OK", 200
                
                elif sub == "setminwd":
                    pending_inputs[(bot_id, uid)] = ("setminwd", time.time())
                    msg = (
                        "✏️ <b>Edit Min Withdraw</b>\n\n"
                        "Reply with minimum withdrawal amount (RM)\n"
                        "Contoh: <code>30.00</code> or <code>50.00</code>"
                    )
                    send_message(token, chat_id, msg, parse_mode="HTML")
                    answer_callback(token, cq["id"])
                    return "OK", 200

            if action == "content":
                sub = parts[2] if len(parts) > 2 else ""
                
                if sub == "editstart":
                    pending_inputs[(bot_id, uid)] = ("editstart", time.time())
                    msg = (
                        "✏️ <b>Edit START Message</b>\n\n"
                        "Reply to this message with your new START message.\n"
                        "You can send text or media (photo/video/gif)."
                    )
                    send_message(token, chat_id, msg, parse_mode="HTML")
                    answer_callback(token, cq["id"])
                    return "OK", 200
                
                elif sub == "editloading":
                    pending_inputs[(bot_id, uid)] = ("editloading", time.time())
                    msg = (
                        "✏️ <b>Edit LOADING Message</b>\n\n"
                        "Reply to this message with your new LOADING message.\n"
                        "You can send text or media (photo/video/gif)."
                    )
                    send_message(token, chat_id, msg, parse_mode="HTML")
                    answer_callback(token, cq["id"])
                    return "OK", 200

            if action == "scan":
                sub = parts[2] if len(parts) > 2 else ""

                if sub == "dur":
                    raw_dur = parts[3] if len(parts) > 3 else ""
                    try:
                        dur = int(raw_dur)
                    except Exception:
                        answer_callback(token, cq["id"], "Invalid duration", show_alert=True)
                        return "OK", 200

                    if dur not in SCAN_DURATION_PRESETS:
                        answer_callback(token, cq["id"], "Preset tak valid", show_alert=True)
                        return "OK", 200

                    with engine.begin() as conn:
                        conn.execute(text("UPDATE bots SET scanner_duration_seconds=:d WHERE id=:i"), {"d": dur, "i": bot_id})

                    answer_callback(token, cq["id"], f"Scanner loading set: {dur}s ({scanner_loading_text_count(dur)} text)")
                    bot_row2 = get_bot_by_id(bot_id) or bot_row
                    send_or_edit_settings_panel(bot_row2, chat_id, uid, page=1, edit_ctx={"message_id": message_id}, cat="scanner")
                    return "OK", 200
                
                if sub == "setglobal":
                    bot_row2 = get_bot_by_id(bot_id) or bot_row
                    cur_lim = bot_row2.get("scan_limit_per_day")
                    cur_txt = "UNLIMITED" if (cur_lim is None or int(cur_lim or 0) <= 0) else str(int(cur_lim))
                    
                    pending_inputs[(bot_id, uid)] = ("scanlimit", time.time())
                    msg = (
                        "⚙️ <b>Set Global Scan Limit</b>\n\n"
                        f"Current: <b>{cur_txt}</b> scans/day\n\n"
                        "Reply with:\n"
                        "• Number (e.g. <code>20</code>) for limit\n"
                        "• <code>off</code> for unlimited"
                    )
                    send_message(token, chat_id, msg, parse_mode="HTML")
                    answer_callback(token, cq["id"])
                    return "OK", 200
                
                elif sub == "viewusage":
                    # Show scan usage stats
                    from datetime import date
                    today = date.today().isoformat()
                    with engine.connect() as conn:
                        usage = conn.execute(
                            text("SELECT user_id, count FROM scan_daily_usage WHERE bot_id=:b AND day=:d ORDER BY count DESC LIMIT 10"),
                            {"b": bot_id, "d": today}
                        ).mappings().all()
                    
                    if usage:
                        lines = [f"• <code>{u['user_id']}</code>: <b>{u['count']}</b> scans" for u in usage]
                        usage_txt = "\n".join(lines)
                    else:
                        usage_txt = "<i>No scans today yet.</i>"
                    
                    msg = (
                        "📊 <b>Scan Usage Today</b>\n"
                        "━━━━━━━━━━━━━━━━━━\n"
                        f"{usage_txt}\n\n"
                        "<i>Top 10 users shown</i>"
                    )
                    send_message(token, chat_id, msg, parse_mode="HTML")
                    answer_callback(token, cq["id"])
                    return "OK", 200
                
                elif sub == "reset":
                    from datetime import date
                    today = date.today().isoformat()
                    with engine.begin() as conn:
                        res = conn.execute(
                            text("DELETE FROM scan_daily_usage WHERE bot_id=:b AND day=:d"),
                            {"b": bot_id, "d": today}
                        )
                        deleted = int(getattr(res, "rowcount", 0) or 0)
                    msg = f"✅ Reset scan usage for today (deleted {deleted} records)"
                    send_message(token, chat_id, msg, parse_mode="HTML")
                    answer_callback(token, cq["id"])
                    return "OK", 200
                
                elif sub == "editmsg":
                    pending_inputs[(bot_id, uid)] = ("scanlimitmsg", time.time())
                    msg = (
                        "✏️ <b>Custom Scan Limit Message</b>\n\n"
                        "Reply to this message with your custom message\n"
                        "shown when user exceeds daily scan limit.\n\n"
                        "You can send text or media (photo/video/gif)."
                    )
                    send_message(token, chat_id, msg, parse_mode="HTML")
                    answer_callback(token, cq["id"])
                    return "OK", 200

                elif sub == "editjoined":
                    pending_inputs[(bot_id, uid)] = ("joinedmsg", time.time())
                    msg = (
                        "✏️ <b>Edit Joined Message</b>\n\n"
                        "Reply to this message with your message after user passes JoinLock.\n"
                        "You can send text or media (photo/video/gif).\n\n"
                        "Support placeholder macam {firstname}, {username}, {member_id}, [balance], [share], [link]."
                    )
                    send_message(token, chat_id, msg, parse_mode="HTML")
                    answer_callback(token, cq["id"])
                    return "OK", 200

            if action == "withdrawal":
                sub = parts[2] if len(parts) > 2 else ""
                if sub == "editrequest":
                    pending_inputs[(bot_id, uid)] = ("withdrawalrequest", time.time())
                    send_message(token, chat_id, "✏️ <b>Edit Withdrawal Request Message</b>\n\nReply with message shown when user REQUESTS withdrawal.\n\nYou can send text or media.", parse_mode="HTML")
                    answer_callback(token, cq["id"])
                    return "OK", 200
                elif sub == "editapprove":
                    pending_inputs[(bot_id, uid)] = ("withdrawalapprove", time.time())
                    send_message(token, chat_id, "✏️ <b>Edit Withdrawal Approve Message</b>\n\nReply with message when approved.\nPlaceholders: <code>{amount}</code>, <code>{balance_after}</code>\n\nYou can send text or media.", parse_mode="HTML")
                    answer_callback(token, cq["id"])
                    return "OK", 200
                elif sub == "editreject":
                    pending_inputs[(bot_id, uid)] = ("withdrawalreject", time.time())
                    send_message(token, chat_id, "✏️ <b>Edit Withdrawal Reject Message</b>\n\nReply with message when rejected.\n\nYou can send text or media.", parse_mode="HTML")
                    answer_callback(token, cq["id"])
                    return "OK", 200
                elif sub == "editfailed":
                    pending_inputs[(bot_id, uid)] = ("withdrawalfailed", time.time())
                    send_message(token, chat_id, "✏️ <b>Edit Withdrawal Failed Message</b>\n\nReply with message when balance insufficient.\nPlaceholders: <code>{min_withdraw}</code>, <code>{balance}</code>\n\nYou can send text or media.", parse_mode="HTML")
                    answer_callback(token, cq["id"])
                    return "OK", 200
                elif sub == "editsubmitted":
                    pending_inputs[(bot_id, uid)] = ("withdrawalsubmitted", time.time())
                    send_message(token, chat_id, "✏️ <b>Edit Withdrawal Submitted Message</b>\n\nReply with message when withdrawal request submitted.\nPlaceholders: <code>{balance}</code>\n\nYou can send text or media.", parse_mode="HTML")
                    answer_callback(token, cq["id"])
                    return "OK", 200

            if action == "verify":
                sub = parts[2] if len(parts) > 2 else ""
                if sub == "editjoin":
                    pending_inputs[(bot_id, uid)] = ("joinmsg", time.time())
                    send_message(token, chat_id, "✏️ <b>Edit JoinLock Message</b>\n\nReply with message when user needs to join channels/groups.\n\nYou can send text or media.", parse_mode="HTML")
                    answer_callback(token, cq["id"])
                    return "OK", 200
                elif sub == "editcontact":
                    pending_inputs[(bot_id, uid)] = ("contactmsg", time.time())
                    send_message(token, chat_id, "✏️ <b>Edit Contact Request Message</b>\n\nReply with message when requesting phone.\n\nYou can send text or media.", parse_mode="HTML")
                    answer_callback(token, cq["id"])
                    return "OK", 200
                elif sub == "editpending":
                    pending_inputs[(bot_id, uid)] = ("pendingmsg", time.time())
                    send_message(token, chat_id, "✏️ <b>Edit Pending Message</b>\n\nReply with message while verification pending.\n\nYou can send text or media.", parse_mode="HTML")
                    answer_callback(token, cq["id"])
                    return "OK", 200
                elif sub == "editverified":
                    pending_inputs[(bot_id, uid)] = ("verifiedmsg", time.time())
                    send_message(token, chat_id, "✏️ <b>Edit Verified Message</b>\n\nReply with message when VERIFIED.\n\nYou can send text or media.", parse_mode="HTML")
                    answer_callback(token, cq["id"])
                    return "OK", 200
                elif sub == "editrejected":
                    pending_inputs[(bot_id, uid)] = ("rejectedmsg", time.time())
                    send_message(token, chat_id, "✏️ <b>Edit Rejected Message</b>\n\nReply with message when REJECTED.\n\nYou can send text or media.", parse_mode="HTML")
                    answer_callback(token, cq["id"])
                    return "OK", 200
                elif sub == "editgroupcontact":
                    pending_inputs[(bot_id, uid)] = ("groupcontactmsg", time.time())
                    send_message(token, chat_id, "✏️ <b>Edit Group Contact Message</b>\n\nReply with message when group contacts shared.\n\nYou can send text or media.", parse_mode="HTML")
                    answer_callback(token, cq["id"])
                    return "OK", 200

            if action == "preview":
                which = parts[2] if len(parts) > 2 else ""
                if which == "start":
                    preview_start(bot_row, chat_id, uid)
                elif which == "loading":
                    preview_loading(bot_row, chat_id, uid)
                elif which == "joined":
                    preview_joined(bot_row, chat_id, uid)
                answer_callback(token, cq["id"])
                return "OK", 200

            if action == "help":
                topic = parts[2] if len(parts) > 2 else ""
                if topic == "all":
                    send_message(token, chat_id, settings_help_all(), parse_mode="HTML")
                else:
                    send_message(token, chat_id, "ℹ️ Select a specific 'How' button for detailed help.", parse_mode="HTML")
                answer_callback(token, cq["id"])
                return "OK", 200

            if action == "how":
                topic = parts[2] if len(parts) > 2 else ""
                send_message(token, chat_id, settings_how(topic), parse_mode="HTML")
                answer_callback(token, cq["id"])
                return "OK", 200

            if action == "placeholders":
                send_message(token, chat_id, HELP_PLACEHOLDERS_FULL, parse_mode="HTML")
                answer_callback(token, cq["id"])
                return "OK", 200

            if action == "cbpage":
                try:
                    page = int(parts[2])
                except Exception:
                    page = 1
                bot_row2 = get_bot_by_id(bot_id) or bot_row
                send_or_edit_settings_panel(bot_row2, chat_id, uid, page=page, edit_ctx={"message_id": message_id})
                answer_callback(token, cq["id"])
                return "OK", 200

            if action == "refresh":
                try:
                    page = int(parts[2])
                except Exception:
                    page = 1
                bot_row2 = get_bot_by_id(bot_id) or bot_row
                send_or_edit_settings_panel(bot_row2, chat_id, uid, page=page, edit_ctx={"message_id": message_id})
                answer_callback(token, cq["id"])
                return "OK", 200

            if action == "export":
                which = parts[2] if len(parts) > 2 else "all"
                export_users_excel(bot_row, chat_id, target=("verified" if which == "verified" else "all"))
                answer_callback(token, cq["id"], "Export sent ✅")
                return "OK", 200

            if action == "mybots":
                try:
                    page = int(parts[2])
                except Exception:
                    page = 0
                send_mybots(bot_row, chat_id, int(bot_row["owner_id"]), page=page)
                answer_callback(token, cq["id"])
                return "OK", 200

            if action == "cbdelmenu":
                cb_total, cb_rows = get_callbacks_page(bot_id, 1, SETTINGS_CB_PAGE_SIZE)
                if not cb_rows:
                    answer_callback(token, cq["id"], "No callbacks", show_alert=True)
                    return "OK", 200
                kb = {"inline_keyboard": []}
                for r in cb_rows[:10]:
                    kb["inline_keyboard"].append([
                        {"text": f"🗑 {r['key']}", "callback_data": f"st:cbdel:{r['key']}"}
                    ])
                kb["inline_keyboard"].append([
                    {"text": "⬅️ Back Panel", "callback_data": "st:refresh:1"}
                ])
                send_message(token, chat_id, "🗑 <b>Delete Callback</b>\nPilih key untuk delete:", reply_markup=kb, parse_mode="HTML")
                answer_callback(token, cq["id"])
                return "OK", 200

            if action == "cbdel":
                key = parts[2] if len(parts) > 2 else ""
                if not key:
                    answer_callback(token, cq["id"], "Missing key", show_alert=True)
                    return "OK", 200
                ok = delete_callback(bot_id, key)
                answer_callback(token, cq["id"], "Deleted ✅" if ok else "Not found ⚠️", show_alert=False)
                bot_row2 = get_bot_by_id(bot_id) or bot_row
                send_or_edit_settings_panel(bot_row2, chat_id, uid, page=1, edit_ctx={"message_id": message_id})
                return "OK", 200

            answer_callback(token, cq["id"])
            return "OK", 200

        answer_callback(token, cq["id"])
        return "OK", 200

    return "OK", 200


@app.post("/webhook")
def webhook_alias():
    return telegram_webhook()



def get_admin_target_chat_id(bot_row: dict) -> int:
    """Return chat_id for admin notifications (group if set, else owner)."""
    try:
        return int(bot_row.get("admin_group_id") or bot_row.get("owner_id") or 0)
    except Exception:
        return 0


def build_premium_approval_keyboard(uid: int):
    return {
        "inline_keyboard": [[
            {"text": "✅ Approve Premium", "callback_data": f"adm:ap:{uid}"},
            {"text": "❌ Reject Premium", "callback_data": f"adm:rj:{uid}"},
        ]]
    }


def send_premium_request_to_admin(bot_row: dict, uid: int, user_row: dict):
    """Send manual premium approval request to admin target."""
    token = bot_row["token"]
    admin_chat = get_admin_target_chat_id(bot_row)
    if not admin_chat:
        return

    fn = (user_row or {}).get("first_name") or ""
    un = (user_row or {}).get("username") or ""
    phone = (user_row or {}).get("phone") or ""
    member_id = (user_row or {}).get("member_id") or ""
    bal = float((user_row or {}).get("balance") or 0)

    header = "🔔 <b>PREMIUM REQUEST (MANUAL)</b>\n"
    body = (
        f"👤 Nama: <b>{html.escape(fn) or '-'}</b>\n"
        f"🔖 Username: <code>{html.escape(un) if un else '-'}</code>\n"
        f"🆔 UID: <code>{uid}</code>\n"
        f"📞 Phone: <code>{html.escape(phone) if phone else '-'}</code>\n"
        f"🎫 Member ID: <code>{html.escape(member_id) if member_id else '-'}</code>\n"
        f"💰 Balance: <b>RM{bal:.2f}</b>\n\n"
        "Tekan button di bawah untuk approve/reject.\n"
        "Atau reply mesej ini dengan <code>/approve</code> atau <code>/reject</code>."
    )
    kb = build_premium_approval_keyboard(uid)
    send_message(token, admin_chat, header + body, reply_markup=kb, parse_mode="HTML")

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.getenv("PORT", "8080")))
